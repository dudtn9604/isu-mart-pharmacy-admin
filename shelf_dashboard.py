"""
Shelf 위치별 성과 트래킹 대시보드
매대 배치 관리 + 위치별 매출 성과 분석 + SKU 치수 관리
Streamlit 포트 8502 — 단면 기준
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from datetime import date, timedelta
from io import BytesIO

import json as _json
from pathlib import Path

from shelf_config import SHELF_CONFIGS, get_total_locations
from shelf_data import (
    init_db,
    get_all_locations,
    get_current_placements,
    get_vacant_locations,
    get_placement_history,
    get_product_placement_history,
    get_all_placements,
    add_placement,
    end_placement,
    swap_placement,
    bulk_add_placements,
    delete_placement,
    fetch_sales_for_placements,
    fetch_sales_for_placement_history,
    get_all_dimensions,
    get_dimension,
    upsert_dimension,
    bulk_upsert_dimensions,
    recommend_locations,
    predict_shelf_demand,
    get_fixture_positions,
    update_fixture_position,
    swap_fixture_positions,
    bulk_update_fixture_positions,
    set_fixture_tiers_enabled,
    get_fixture_tier_status,
    get_showcard_history,
    save_showcard,
)

# ──────────────────────────────────────
# 페이지 설정
# ──────────────────────────────────────
st.set_page_config(
    page_title="매대 위치별 성과 트래킹",
    page_icon="🗄️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# DB 초기화
init_db()


# ──────────────────────────────────────
# 에디터 저장 API (포트 8503) — 별도 프로세스
# ──────────────────────────────────────
LAYOUT_FILE = Path(__file__).parent / "shelf_layout.json"
FOREON_LAYOUT_FILE = Path(__file__).parent / "foreon_layout.json"


def _ensure_layout_api():
    """포트 8503 API 서버가 살아있는지 확인 후, 없으면 subprocess로 시작"""
    import socket, subprocess, sys
    try:
        s = socket.create_connection(("localhost", 8503), timeout=0.5)
        s.close()
        return  # 이미 살아있음
    except (ConnectionRefusedError, OSError, socket.timeout):
        pass
    # 별도 프로세스로 layout_api_server.py 시작
    api_script = Path(__file__).parent / "layout_api_server.py"
    subprocess.Popen(
        [sys.executable, str(api_script)],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        start_new_session=True,
    )


_ensure_layout_api()


# ──────────────────────────────────────
# Supabase 상품 목록 캐시
# ──────────────────────────────────────
def load_product_list():
    try:
        from supabase_client import is_supabase_configured, fetch_products
        if not is_supabase_configured():
            st.warning("Supabase 설정이 완료되지 않았습니다.")
            return pd.DataFrame()
        return fetch_products()
    except Exception as e:
        st.warning(f"상품 목록 로드 실패: {e}")
        return pd.DataFrame()


@st.cache_data(ttl=3600)
def load_sale_cost_records(date_from: str, date_to: str):
    try:
        from supabase_client import is_supabase_configured, fetch_sale_cost_records
        if not is_supabase_configured():
            return pd.DataFrame()
        return fetch_sale_cost_records(date_from=date_from, date_to=date_to)
    except Exception:
        return pd.DataFrame()


# ──────────────────────────────────────
# 사이드바 메뉴
# ──────────────────────────────────────
st.sidebar.title("🗄️ 매대 성과 트래킹")
st.sidebar.markdown("---")

menu = st.sidebar.radio(
    "메뉴",
    ["🗺️ 매장 배치도", "✏️ 배치 관리", "📊 위치별 성과 분석", "📐 SKU 치수 관리", "🛒 교차판매 분석", "🏷️ 쇼카드 제작", "🏪 포레온 시뮬레이션"],
    label_visibility="collapsed",
)

# ======================================================================
# 매장 배치도
# ======================================================================
STORE_W = 12215  # mm
STORE_H = 17848  # mm

# 매대 물리 크기 (mm)
FIXTURE_DIMS = {
    "A": {"w": 900, "d": 360},
    "B": {"w": 930, "d": 360},
    "C": {"w": 636, "d": 360},
}

# 타입별 색상
TYPE_COLORS = {
    "A": "#4A90D9",  # 파랑
    "B": "#50C878",  # 초록
    "C": "#FF8C00",  # 주황
}
TYPE_COLORS_LIGHT = {
    "A": "rgba(74,144,217,0.3)",
    "B": "rgba(80,200,120,0.3)",
    "C": "rgba(255,140,0,0.3)",
}


def draw_floor_plan(fixtures_df, selected_id=None):
    """매장 평면도 Plotly Figure 생성"""
    fig = go.Figure()

    # 매장 외벽
    fig.add_shape(type="rect", x0=0, y0=0, x1=STORE_W, y1=STORE_H,
                  line=dict(color="black", width=2), fillcolor="rgba(245,245,245,0.5)")

    # 고정 요소
    fixed = [
        {"name": "AUTO DOOR", "x": 4200, "y": STORE_H - 500, "w": 2200, "h": 500, "c": "#DDD"},
        {"name": "POS", "x": 3500, "y": 2200, "w": 2200, "h": 700, "c": "#E8D5B7"},
        {"name": "조제실", "x": 5800, "y": 200, "w": 3200, "h": 2200, "c": "#D4E6F1"},
        {"name": "창고", "x": 9200, "y": 8000, "w": 1200, "h": 3000, "c": "#E8E8E8"},
        {"name": "프로모션 존", "x": 3200, "y": 6200, "w": 2200, "h": 1800, "c": "#FCE4EC"},
        {"name": "대기 공간", "x": 6800, "y": 5800, "w": 1600, "h": 1400, "c": "#E8F5E9"},
        {"name": "냉장고", "x": 10800, "y": 3000, "w": 1200, "h": 5000, "c": "#B3E5FC"},
        {"name": "약품 수납장", "x": 1500, "y": 200, "w": 3800, "h": 1600, "c": "#F3E5F5"},
    ]

    for f in fixed:
        fig.add_shape(type="rect",
                      x0=f["x"], y0=f["y"], x1=f["x"] + f["w"], y1=f["y"] + f["h"],
                      line=dict(color="#999", width=1), fillcolor=f["c"])
        fig.add_annotation(x=f["x"] + f["w"] / 2, y=f["y"] + f["h"] / 2,
                           text=f["name"], showarrow=False,
                           font=dict(size=9, color="#666"))

    # 매대 그리기
    for _, fx in fixtures_df.iterrows():
        stype = fx["shelf_type"]
        fno = fx["fixture_no"]
        x = fx["x_pos"]
        y = fx["y_pos"]
        orient = fx["orientation"]

        dims = FIXTURE_DIMS[stype]
        if orient == "V":
            dx, dy = dims["d"], dims["w"]
        else:
            dx, dy = dims["w"], dims["d"]

        is_selected = (selected_id is not None and
                       fx["shelf_type"] == selected_id[0] and
                       fx["fixture_no"] == selected_id[1])

        fill = TYPE_COLORS[stype] if is_selected else TYPE_COLORS_LIGHT[stype]
        border_w = 3 if is_selected else 1

        fig.add_shape(
            type="rect",
            x0=x, y0=y, x1=x + dx, y1=y + dy,
            line=dict(color=TYPE_COLORS[stype], width=border_w),
            fillcolor=fill,
        )

        label = fx.get("custom_label") or f"{stype}-{fno}"
        fig.add_annotation(
            x=x + dx / 2, y=y + dy / 2,
            text=label, showarrow=False,
            font=dict(size=7, color="black"),
        )

    # 범례 (수동)
    for stype, color in TYPE_COLORS.items():
        name = SHELF_CONFIGS[stype]["name"]
        fig.add_trace(go.Scatter(
            x=[None], y=[None], mode="markers",
            marker=dict(size=12, color=color, symbol="square"),
            name=f"{stype} ({name}) — {SHELF_CONFIGS[stype]['count']}대",
            showlegend=True,
        ))

    fig.update_layout(
        width=900, height=1100,
        xaxis=dict(range=[-200, STORE_W + 200], scaleanchor="y", scaleratio=1,
                   showgrid=False, zeroline=False, showticklabels=False),
        yaxis=dict(range=[-200, STORE_H + 200],
                   showgrid=False, zeroline=False, showticklabels=False),
        margin=dict(l=20, r=20, t=40, b=20),
        title="매장 배치도 (mm 단위)",
        legend=dict(orientation="h", yanchor="bottom", y=1.02),
        plot_bgcolor="white",
    )
    return fig


if menu == "🗺️ 매장 배치도":
    st.title("🗺️ 매장 배치도")

    # ── 상품 검색 ──
    _search_placements = get_current_placements()
    if not _search_placements.empty:
        _search_product_names = sorted(_search_placements["product_name"].dropna().unique().tolist())
        _search_product = st.selectbox(
            "🔍 상품 검색 — 배치도에서 위치 찾기",
            _search_product_names, index=None,
            placeholder="상품명을 입력하세요...",
            key="floor_product_search",
        )
        if _search_product:
            _matched = _search_placements[_search_placements["product_name"] == _search_product]
            _highlight_fixtures = set()
            _search_details = []
            for _, r in _matched.iterrows():
                fid = f"{r['shelf_type']}-{int(r['fixture_no'])}"
                _highlight_fixtures.add(fid)
                ps = int(r.get("position_start") or 1)
                pe = int(r.get("position_end") or 1)
                pos_str = f"{ps}번" if ps == pe else f"{ps}~{pe}번"
                _search_details.append(f"**{fid}** / {int(r['tier'])}단 / {pos_str}")
            st.info(f"📍 **{_search_product}** 위치: " + " | ".join(_search_details))
            _highlight_json = _json.dumps(list(_highlight_fixtures), ensure_ascii=False)
        else:
            _highlight_json = "[]"
    else:
        _highlight_json = "[]"

    st.markdown("---")

    # ── 배치 현황 요약 ──
    _all_locations = get_all_locations()
    _all_placements = get_current_placements() if _search_placements.empty else _search_placements
    _total_locs = len(_all_locations)
    _placed_locs = _all_placements["shelf_location_id"].nunique() if not _all_placements.empty else 0
    _active_skus = _all_placements["product_name"].nunique() if not _all_placements.empty else 0

    # 공간 활용률 계산: 배치 상품 가로 합계 / 매대 가로 길이
    _space_util = 0.0
    if not _all_placements.empty:
        _dims = get_all_dimensions()
        _width_map = {}
        if not _dims.empty:
            _width_map = _dims.set_index("product_name")["width"].dropna().to_dict()

        _tier_utils = []
        for (stype, fno, tier), grp in _all_placements.groupby(["shelf_type", "fixture_no", "tier"]):
            shelf_width = SHELF_CONFIGS.get(stype, {}).get("width", 0)
            if shelf_width <= 0:
                continue
            used_width = sum(_width_map.get(name, 0) for name in grp["product_name"])
            _tier_utils.append(min(used_width / shelf_width * 100, 100.0))

        if _tier_utils:
            _space_util = sum(_tier_utils) / len(_tier_utils)

    _k1, _k2, _k3, _k4 = st.columns(4)
    _k1.metric("총 선반", f"{_total_locs}개")
    _k2.metric("배치된 수", f"{_placed_locs}개")
    _k3.metric("공간 활용률", f"{_space_util:.1f}%",
               help="배치된 선반의 가로 공간 중 상품이 차지하는 비율 (평균)")
    _k4.metric("활용 SKU 수", f"{_active_skus}개")

    st.markdown("---")

    # 저장된 레이아웃 파일에서 로드 (있으면), 없으면 DB에서 로드
    _saved_layout = None
    if LAYOUT_FILE.exists():
        try:
            with open(str(LAYOUT_FILE), "r", encoding="utf-8") as f:
                _saved_layout = _json.load(f)
        except Exception:
            pass

    if _saved_layout and _saved_layout.get("fixtures"):
        editor_fixtures = _saved_layout["fixtures"]
        editor_facilities_data = _saved_layout.get("facilities", [])
    else:
        fixtures_df = get_fixture_positions()
        if fixtures_df.empty:
            st.warning("배치도 데이터가 없습니다. DB를 재초기화해 주세요.")
            st.stop()

        editor_fixtures = []
        for _, row in fixtures_df.iterrows():
            editor_fixtures.append({
                "id": f"{row['shelf_type']}-{int(row['fixture_no'])}",
                "type": row["shelf_type"],
                "no": int(row["fixture_no"]),
                "x": float(row["x_pos"]),
                "y": float(row["y_pos"]),
                "orient": row["orientation"],
                "zone": row["zone"] or "상단",
                "label": row.get("custom_label", "") or "",
            })

        editor_facilities_data = [
            {"id": "fac-1", "name": "POS", "x": 3500, "y": STORE_H-2900, "w": 2200, "h": 700, "label": ""},
            {"id": "fac-2", "name": "조제실", "x": 5800, "y": STORE_H-2400, "w": 3200, "h": 2200, "label": ""},
            {"id": "fac-3", "name": "창고", "x": 9200, "y": STORE_H-11000, "w": 1200, "h": 3000, "label": ""},
            {"id": "fac-4", "name": "프로모션 존", "x": 3200, "y": STORE_H-8000, "w": 2200, "h": 1800, "label": ""},
            {"id": "fac-5", "name": "대기 공간", "x": 6800, "y": STORE_H-7200, "w": 1600, "h": 1400, "label": ""},
            {"id": "fac-6", "name": "냉장고", "x": 10800, "y": STORE_H-8000, "w": 1200, "h": 5000, "label": ""},
            {"id": "fac-7", "name": "약품 수납장", "x": 1500, "y": STORE_H-1800, "w": 3800, "h": 1600, "label": ""},
        ]

    initial_data_json = _json.dumps(editor_fixtures, ensure_ascii=False)
    facilities_json = _json.dumps(editor_facilities_data, ensure_ascii=False)

    st.caption("💡 맵 줌: **Ctrl+스크롤** (Mac: ⌘+스크롤) | 드래그로 이동")

    # ── 임베디드 에디터 HTML ──
    editor_html = f"""
    <div id="editor-root" style="width:100%;height:480px;position:relative;background:#f8f8f8;border:1px solid #ddd;border-radius:8px;overflow:hidden;">
      <div id="toolbar" style="height:44px;background:#fff;border-bottom:1px solid #ddd;display:flex;align-items:center;padding:0 10px;gap:6px;font-size:13px;">
        <button onclick="addFixture('A')" style="padding:4px 10px;border:1px solid #4A90D9;color:#4A90D9;border-radius:4px;background:#fff;cursor:pointer;">+A</button>
        <button onclick="addFixture('B')" style="padding:4px 10px;border:1px solid #50C878;color:#50C878;border-radius:4px;background:#fff;cursor:pointer;">+B</button>
        <button onclick="addFixture('C')" style="padding:4px 10px;border:1px solid #FF8C00;color:#FF8C00;border-radius:4px;background:#fff;cursor:pointer;">+C</button>
        <span style="width:1px;height:24px;background:#ddd;margin:0 2px;"></span>
        <select id="addFac" onchange="addFacilityFromSelect(this)" style="padding:3px 6px;border:1px solid #ccc;border-radius:4px;font-size:12px;">
          <option value="">+시설물</option>
          <option value="조제실">조제실</option><option value="창고">창고</option>
          <option value="프로모션 존">프로모션 존</option><option value="냉장고">냉장고</option>
          <option value="약품 수납장">약품 수납장</option><option value="POS">POS</option>
          <option value="대기 공간">대기 공간</option><option value="기타">기타</option>
        </select>
        <span style="width:1px;height:24px;background:#ddd;margin:0 2px;"></span>
        <button onclick="rotateSelected()" style="padding:4px 10px;border:1px solid #ccc;border-radius:4px;background:#fff;cursor:pointer;">회전(R)</button>
        <button onclick="deleteSelected()" style="padding:4px 10px;border:1px solid #e74c3c;color:#e74c3c;border-radius:4px;background:#fff;cursor:pointer;">삭제(Del)</button>
        <button onclick="undoAction()" style="padding:4px 10px;border:1px solid #ccc;border-radius:4px;background:#fff;cursor:pointer;">되돌리기</button>
        <span style="width:1px;height:24px;background:#ddd;margin:0 2px;"></span>
        <label style="font-size:11px;">스냅:</label>
        <select id="snapSel" onchange="snapGrid=+this.value" style="padding:2px 4px;border:1px solid #ccc;border-radius:4px;font-size:11px;">
          <option value="0">없음</option><option value="50">50</option><option value="100" selected>100</option><option value="200">200</option>
        </select>
        <span style="flex:1;"></span>
        <button id="saveBtn" onclick="saveToStreamlit()" style="padding:4px 14px;border:1px solid #4A90D9;background:#4A90D9;color:#fff;border-radius:4px;cursor:pointer;font-weight:bold;">DB에 저장</button>
        <span id="statusText" style="font-size:11px;color:#888;margin-left:8px;"></span>
      </div>
      <svg id="svg" style="display:block;"></svg>
    </div>
    <style>
      .fixture {{ cursor: move; }}
      .fixture:hover rect {{ stroke-width: 2.5; }}
      .fixture.selected rect {{ stroke-width: 3; filter: drop-shadow(0 0 4px rgba(0,0,0,0.3)); }}
      .fixture text {{ pointer-events: none; user-select: none; }}
      .facility {{ cursor: move; }}
      .facility:hover rect {{ stroke-width: 2.5; }}
      .facility.selected rect {{ stroke-width: 3; stroke-dasharray: 6,3; filter: drop-shadow(0 0 4px rgba(0,0,0,0.3)); }}
      .facility text {{ pointer-events: none; user-select: none; }}
      .fixed-element rect {{ pointer-events: none; }}
      .fixed-element text {{ pointer-events: none; user-select: none; }}
    </style>
    <script>
    const STORE_W = {STORE_W}, STORE_H = {STORE_H};
    const TYPES = {{
      A: {{ name:'기본매대', w:900, d:360, color:'#4A90D9', light:'rgba(74,144,217,0.25)' }},
      B: {{ name:'연결매대', w:930, d:360, color:'#50C878', light:'rgba(80,200,120,0.25)' }},
      C: {{ name:'엔드캡매대', w:636, d:360, color:'#FF8C00', light:'rgba(255,140,0,0.25)' }},
    }};
    const TRULY_FIXED = [{{ name:'AUTO DOOR', x:4200, y:0, w:2200, h:500, c:'#DDD' }}];
    const FACILITY_TYPES = {{
      '조제실':{{w:3200,h:2200,c:'#D4E6F1',border:'#5B9BD5'}},
      '창고':{{w:1200,h:3000,c:'#E8E8E8',border:'#999'}},
      '프로모션 존':{{w:2200,h:1800,c:'#FCE4EC',border:'#E91E63'}},
      '냉장고':{{w:1200,h:5000,c:'#B3E5FC',border:'#03A9F4'}},
      '약품 수납장':{{w:3800,h:1600,c:'#F3E5F5',border:'#9C27B0'}},
      'POS':{{w:2200,h:700,c:'#E8D5B7',border:'#A0522D'}},
      '대기 공간':{{w:1600,h:1400,c:'#E8F5E9',border:'#4CAF50'}},
      '기타':{{w:1000,h:1000,c:'#F5F5F5',border:'#757575'}},
    }};

    let fixtures = {initial_data_json};
    let facilities = {facilities_json};
    const highlightFixtures = new Set({_highlight_json});
    // 다중 선택: [{{type:'fixture'|'facility', id:'A-1'}}, ...]
    let selection = [];
    let snapGrid = 100, undoStack = [];
    let scale = 1, panX = 0, panY = 0;
    let isPanning = false, panStartX = 0, panStartY = 0;
    let isDragging = false, dragItems = [], dragStartPositions = [], dragOffX = 0, dragOffY = 0;
    let placingType = null, placingFacility = null;
    // 범위 선택 (마키)
    let isMarquee = false, marqueeX0 = 0, marqueeY0 = 0, marqueeX1 = 0, marqueeY1 = 0;
    // 리사이즈
    let isResizing = false, resizeFac = null, resizeHandle = '', resizeStartX = 0, resizeStartY = 0;
    let resizeOrigX = 0, resizeOrigY = 0, resizeOrigW = 0, resizeOrigH = 0;
    const HANDLE_SIZE = 8;

    function isSelected(type, id) {{ return selection.some(s => s.type === type && s.id === id); }}
    function getSelectedItems() {{
      return selection.map(s => {{
        if (s.type === 'fixture') return {{ ...fixtures.find(f => f.id === s.id), _stype: 'fixture' }};
        else return {{ ...facilities.find(f => f.id === s.id), _stype: 'facility' }};
      }}).filter(Boolean);
    }}

    const svgNS = 'http://www.w3.org/2000/svg';
    const svg = document.getElementById('svg');
    const root = document.getElementById('editor-root');

    function toSVG(mx, my) {{ return [mx * scale + panX, my * scale + panY]; }}
    function fromSVG(sx, sy) {{ return [(sx - panX) / scale, (sy - panY) / scale]; }}
    function snap(v) {{ return snapGrid > 0 ? Math.round(v / snapGrid) * snapGrid : v; }}

    function render() {{
      svg.innerHTML = '';
      const ww = root.clientWidth, wh = root.clientHeight - 44;
      svg.setAttribute('width', ww); svg.setAttribute('height', wh);

      // 배경
      const bg = document.createElementNS(svgNS, 'rect');
      bg.setAttribute('width', ww); bg.setAttribute('height', wh);
      bg.setAttribute('fill', '#f8f8f8');
      svg.appendChild(bg);

      // 매장 외벽
      const [sx, sy] = toSVG(0, 0);
      const sw = STORE_W * scale, sh = STORE_H * scale;
      addRect(svg, sx, sy, sw, sh, '#333', 2, 'rgba(255,255,255,0.9)');

      // 그리드
      if (scale > 0.04) {{
        const gridSize = scale > 0.08 ? 1000 : 2000;
        const gridG = document.createElementNS(svgNS, 'g');
        gridG.setAttribute('opacity', '0.15');
        for (let gx = 0; gx <= STORE_W; gx += gridSize) {{
          const [lx] = toSVG(gx, 0);
          const line = document.createElementNS(svgNS, 'line');
          line.setAttribute('x1', lx); line.setAttribute('y1', sy);
          line.setAttribute('x2', lx); line.setAttribute('y2', sy + sh);
          line.setAttribute('stroke', '#999'); line.setAttribute('stroke-width', 0.5);
          gridG.appendChild(line);
        }}
        for (let gy = 0; gy <= STORE_H; gy += gridSize) {{
          const [, ly] = toSVG(0, gy);
          const line = document.createElementNS(svgNS, 'line');
          line.setAttribute('x1', sx); line.setAttribute('y1', ly);
          line.setAttribute('x2', sx + sw); line.setAttribute('y2', ly);
          line.setAttribute('stroke', '#999'); line.setAttribute('stroke-width', 0.5);
          gridG.appendChild(line);
        }}
        svg.appendChild(gridG);
      }}

      // 고정 요소 (AUTO DOOR)
      TRULY_FIXED.forEach(f => {{
        const g = document.createElementNS(svgNS, 'g');
        g.setAttribute('class', 'fixed-element');
        const [fx, fy] = toSVG(f.x, f.y);
        addRect(g, fx, fy, f.w * scale, f.h * scale, '#aaa', 0.5, f.c);
        addText(g, fx + f.w * scale / 2, fy + f.h * scale / 2, f.name, Math.max(8, 10 * scale / 0.065), '#666');
        svg.appendChild(g);
      }});

      // 시설물
      facilities.forEach(fac => {{
        const g = document.createElementNS(svgNS, 'g');
        const isSel = isSelected('facility', fac.id);
        g.setAttribute('class', 'facility' + (isSel ? ' selected' : ''));
        const ft = FACILITY_TYPES[fac.name] || FACILITY_TYPES['기타'];
        const fw = fac.w * scale, fh = fac.h * scale;
        const [fx, fy] = toSVG(fac.x, fac.y);
        addRect(g, fx, fy, fw, fh, isSel ? ft.border : '#aaa', isSel ? 3 : 1, ft.c);
        addText(g, fx + fw / 2, fy + fh / 2, fac.label || fac.name, Math.max(8, 10 * scale / 0.065), '#555');
        // 리사이즈 핸들 (선택된 시설물, 단일 선택 시)
        if (isSel && selection.length === 1) {{
          const hs = HANDLE_SIZE;
          const handles = [
            {{ name: 'nw', cx: fx,      cy: fy,      cursor: 'nw-resize' }},
            {{ name: 'ne', cx: fx + fw, cy: fy,      cursor: 'ne-resize' }},
            {{ name: 'sw', cx: fx,      cy: fy + fh, cursor: 'sw-resize' }},
            {{ name: 'se', cx: fx + fw, cy: fy + fh, cursor: 'se-resize' }},
            {{ name: 'n',  cx: fx + fw/2, cy: fy,      cursor: 'n-resize' }},
            {{ name: 's',  cx: fx + fw/2, cy: fy + fh, cursor: 's-resize' }},
            {{ name: 'w',  cx: fx,        cy: fy + fh/2, cursor: 'w-resize' }},
            {{ name: 'e',  cx: fx + fw,   cy: fy + fh/2, cursor: 'e-resize' }},
          ];
          handles.forEach(h => {{
            const hr = document.createElementNS(svgNS, 'rect');
            hr.setAttribute('x', h.cx - hs/2); hr.setAttribute('y', h.cy - hs/2);
            hr.setAttribute('width', hs); hr.setAttribute('height', hs);
            hr.setAttribute('fill', '#fff'); hr.setAttribute('stroke', ft.border);
            hr.setAttribute('stroke-width', 1.5); hr.setAttribute('rx', 2);
            hr.style.cursor = h.cursor;
            hr.addEventListener('mousedown', ev => {{ ev.stopPropagation(); startResize(ev, fac, h.name); }});
            g.appendChild(hr);
          }});
          // 크기 표시 라벨
          addText(g, fx + fw / 2, fy + fh + 14, Math.round(fac.w) + ' × ' + Math.round(fac.h) + ' mm', 9, '#999');
        }}
        g.addEventListener('mousedown', e => onFacilityMouseDown(e, fac));
        svg.appendChild(g);
      }});

      // 매대
      fixtures.forEach(fx => {{
        const g = document.createElementNS(svgNS, 'g');
        const isSel = isSelected('fixture', fx.id);
        const isHL = highlightFixtures.has(fx.id);
        g.setAttribute('class', 'fixture' + (isSel ? ' selected' : ''));
        const t = TYPES[fx.type];
        const dx = (fx.orient === 'V' ? t.d : t.w) * scale;
        const dy = (fx.orient === 'V' ? t.w : t.d) * scale;
        const [rx, ry] = toSVG(fx.x, fx.y);
        if (isHL) {{
          // 하이라이트: 글로우 배경
          const glow = document.createElementNS(svgNS, 'rect');
          glow.setAttribute('x', rx - 3); glow.setAttribute('y', ry - 3);
          glow.setAttribute('width', dx + 6); glow.setAttribute('height', dy + 6);
          glow.setAttribute('rx', 4);
          glow.setAttribute('fill', 'none'); glow.setAttribute('stroke', '#FF4444');
          glow.setAttribute('stroke-width', 3); glow.setAttribute('opacity', 0.8);
          const anim = document.createElementNS(svgNS, 'animate');
          anim.setAttribute('attributeName', 'opacity');
          anim.setAttribute('values', '0.4;1;0.4'); anim.setAttribute('dur', '1.5s');
          anim.setAttribute('repeatCount', 'indefinite');
          glow.appendChild(anim);
          g.appendChild(glow);
          addRect(g, rx, ry, dx, dy, '#FF4444', 2.5, '#FF444433');
        }} else {{
          addRect(g, rx, ry, dx, dy, t.color, isSel ? 3 : 1.5, isSel ? t.color : t.light);
        }}
        const label = fx.label || fx.id;
        const fontSize = Math.max(6, Math.min(11, Math.min(dx, dy) * 0.4));
        addText(g, rx + dx / 2, ry + dy / 2, label, fontSize, isHL ? '#FF4444' : (isSel ? '#fff' : '#333'));
        g.addEventListener('mousedown', e => onFixtureMouseDown(e, fx));
        svg.appendChild(g);
      }});

      // 범위 선택 마키 그리기
      if (isMarquee) {{
        const [sx0, sy0] = toSVG(Math.min(marqueeX0, marqueeX1), Math.min(marqueeY0, marqueeY1));
        const mw = Math.abs(marqueeX1 - marqueeX0) * scale;
        const mh = Math.abs(marqueeY1 - marqueeY0) * scale;
        addRect(svg, sx0, sy0, mw, mh, '#4A90D9', 1.5, 'rgba(74,144,217,0.12)');
      }}

      updateStatus();
    }}

    function addRect(parent, x, y, w, h, stroke, sw, fill) {{
      const r = document.createElementNS(svgNS, 'rect');
      r.setAttribute('x', x); r.setAttribute('y', y);
      r.setAttribute('width', w); r.setAttribute('height', h);
      r.setAttribute('stroke', stroke); r.setAttribute('stroke-width', sw);
      r.setAttribute('fill', fill || 'none'); r.setAttribute('rx', 2);
      parent.appendChild(r);
    }}
    function addText(parent, x, y, text, size, color) {{
      const t = document.createElementNS(svgNS, 'text');
      t.setAttribute('x', x); t.setAttribute('y', y);
      t.setAttribute('text-anchor', 'middle'); t.setAttribute('dominant-baseline', 'central');
      t.setAttribute('font-size', size); t.setAttribute('fill', color);
      t.setAttribute('font-family', '-apple-system, sans-serif');
      t.textContent = text;
      parent.appendChild(t);
    }}

    // ── 이벤트 ──
    function startDragSelected(e, clickedItem, clickedType) {{
      const rect = root.getBoundingClientRect();
      const [mx, my] = fromSVG(e.clientX - rect.left, e.clientY - rect.top - 44);
      dragOffX = mx; dragOffY = my; isDragging = false;
      // 선택된 모든 아이템의 시작 위치 저장
      dragItems = [];
      dragStartPositions = [];
      selection.forEach(s => {{
        let item;
        if (s.type === 'fixture') item = fixtures.find(f => f.id === s.id);
        else item = facilities.find(f => f.id === s.id);
        if (item) {{
          dragItems.push(item);
          dragStartPositions.push({{ x: item.x, y: item.y }});
        }}
      }});
      saveUndo();
    }}

    function onFixtureMouseDown(e, fx) {{
      e.stopPropagation();
      if (e.shiftKey) {{
        // Shift+클릭: 선택 토글
        if (isSelected('fixture', fx.id)) {{
          selection = selection.filter(s => !(s.type === 'fixture' && s.id === fx.id));
        }} else {{
          selection.push({{ type: 'fixture', id: fx.id }});
        }}
        render(); return;
      }}
      // 이미 선택된 아이템이면 드래그 시작 (선택 유지)
      if (!isSelected('fixture', fx.id)) {{
        selection = [{{ type: 'fixture', id: fx.id }}];
      }}
      render();
      startDragSelected(e, fx, 'fixture');
    }}
    function onFacilityMouseDown(e, fac) {{
      e.stopPropagation();
      if (e.shiftKey) {{
        if (isSelected('facility', fac.id)) {{
          selection = selection.filter(s => !(s.type === 'facility' && s.id === fac.id));
        }} else {{
          selection.push({{ type: 'facility', id: fac.id }});
        }}
        render(); return;
      }}
      if (!isSelected('facility', fac.id)) {{
        selection = [{{ type: 'facility', id: fac.id }}];
      }}
      render();
      startDragSelected(e, fac, 'facility');
    }}

    // 선택된 아이템들의 바운딩 박스 계산
    function getSelectionBBox() {{
      let minX = Infinity, minY = Infinity, maxX = -Infinity, maxY = -Infinity;
      selection.forEach(s => {{
        let item, w, h;
        if (s.type === 'fixture') {{
          item = fixtures.find(f => f.id === s.id);
          if (!item) return;
          const t = TYPES[item.type];
          w = item.orient === 'V' ? t.d : t.w;
          h = item.orient === 'V' ? t.w : t.d;
        }} else {{
          item = facilities.find(f => f.id === s.id);
          if (!item) return;
          w = item.w; h = item.h;
        }}
        minX = Math.min(minX, item.x);
        minY = Math.min(minY, item.y);
        maxX = Math.max(maxX, item.x + w);
        maxY = Math.max(maxY, item.y + h);
      }});
      return {{ x0: minX, y0: minY, x1: maxX, y1: maxY }};
    }}

    svg.addEventListener('mousedown', e => {{
      if (placingType) {{
        const rect = root.getBoundingClientRect();
        const [mx, my] = fromSVG(e.clientX - rect.left, e.clientY - rect.top - 44);
        placeNewFixture(snap(mx), snap(my)); return;
      }}
      if (placingFacility) {{
        const rect = root.getBoundingClientRect();
        const [mx, my] = fromSVG(e.clientX - rect.left, e.clientY - rect.top - 44);
        placeNewFacility(snap(mx), snap(my)); return;
      }}
      const rect = root.getBoundingClientRect();
      const [mx, my] = fromSVG(e.clientX - rect.left, e.clientY - rect.top - 44);
      if (e.shiftKey) {{
        // Shift+드래그: 범위 선택 (마키)
        isMarquee = true;
        marqueeX0 = mx; marqueeY0 = my;
        marqueeX1 = mx; marqueeY1 = my;
      }} else if (selection.length > 0) {{
        // 선택된 아이템이 있을 때: 바운딩 박스 안이면 그룹 이동, 밖이면 팬
        const bbox = getSelectionBBox();
        const pad = 300;
        if (mx >= bbox.x0 - pad && mx <= bbox.x1 + pad && my >= bbox.y0 - pad && my <= bbox.y1 + pad) {{
          dragOffX = mx; dragOffY = my; isDragging = false;
          dragItems = []; dragStartPositions = [];
          selection.forEach(s => {{
            let item;
            if (s.type === 'fixture') item = fixtures.find(f => f.id === s.id);
            else item = facilities.find(f => f.id === s.id);
            if (item) {{ dragItems.push(item); dragStartPositions.push({{ x: item.x, y: item.y }}); }}
          }});
          saveUndo();
        }} else {{
          selection = []; render();
          isPanning = true;
          panStartX = e.clientX - rect.left - panX;
          panStartY = e.clientY - rect.top - 44 - panY;
        }}
      }} else {{
        // 선택 없음: 화면 이동 (팬)
        isPanning = true;
        panStartX = e.clientX - rect.left - panX;
        panStartY = e.clientY - rect.top - 44 - panY;
      }}
    }});

    window.addEventListener('mousemove', e => {{
      const rect = root.getBoundingClientRect();
      const [mx, my] = fromSVG(e.clientX - rect.left, e.clientY - rect.top - 44);
      if (isResizing && resizeFac) {{
        const dx = snap(mx - resizeStartX);
        const dy = snap(my - resizeStartY);
        const MIN_SIZE = 200; // 최소 크기 200mm
        const h = resizeHandle;
        let newX = resizeOrigX, newY = resizeOrigY, newW = resizeOrigW, newH = resizeOrigH;
        if (h.includes('e')) newW = Math.max(MIN_SIZE, resizeOrigW + dx);
        if (h.includes('w')) {{ newW = Math.max(MIN_SIZE, resizeOrigW - dx); newX = resizeOrigX + resizeOrigW - newW; }}
        if (h.includes('s')) newH = Math.max(MIN_SIZE, resizeOrigH + dy);
        if (h.includes('n')) {{ newH = Math.max(MIN_SIZE, resizeOrigH - dy); newY = resizeOrigY + resizeOrigH - newH; }}
        resizeFac.x = newX; resizeFac.y = newY;
        resizeFac.w = newW; resizeFac.h = newH;
        render();
      }}
      if (dragItems.length > 0) {{
        isDragging = true;
        const dx = snap(mx - dragOffX);
        const dy = snap(my - dragOffY);
        dragItems.forEach((item, i) => {{
          item.x = Math.max(0, Math.min(STORE_W - 100, dragStartPositions[i].x + dx));
          item.y = Math.max(0, Math.min(STORE_H - 100, dragStartPositions[i].y + dy));
        }});
        render();
      }}
      if (isMarquee) {{
        marqueeX1 = mx; marqueeY1 = my;
        // 마키 영역 내 아이템 선택
        const x0 = Math.min(marqueeX0, marqueeX1), x1 = Math.max(marqueeX0, marqueeX1);
        const y0 = Math.min(marqueeY0, marqueeY1), y1 = Math.max(marqueeY0, marqueeY1);
        selection = [];
        fixtures.forEach(fx => {{
          const t = TYPES[fx.type];
          const fw = fx.orient === 'V' ? t.d : t.w;
          const fh = fx.orient === 'V' ? t.w : t.d;
          if (fx.x + fw > x0 && fx.x < x1 && fx.y + fh > y0 && fx.y < y1)
            selection.push({{ type: 'fixture', id: fx.id }});
        }});
        facilities.forEach(fac => {{
          if (fac.x + fac.w > x0 && fac.x < x1 && fac.y + fac.h > y0 && fac.y < y1)
            selection.push({{ type: 'facility', id: fac.id }});
        }});
        render();
      }}
      if (isPanning) {{
        panX = e.clientX - rect.left - panStartX;
        panY = e.clientY - rect.top - 44 - panStartY;
        render();
      }}
    }});

    window.addEventListener('mouseup', () => {{
      dragItems = []; dragStartPositions = [];
      isPanning = false; isMarquee = false;
      isResizing = false; resizeFac = null;
    }});

    svg.addEventListener('wheel', e => {{
      if (!e.ctrlKey && !e.metaKey) return;  // 일반 스크롤은 페이지 스크롤로 통과
      e.preventDefault();
      const rect = root.getBoundingClientRect();
      const [mx, my] = fromSVG(e.clientX - rect.left, e.clientY - rect.top - 44);
      const factor = e.deltaY < 0 ? 1.15 : 1 / 1.15;
      const newScale = Math.max(0.02, Math.min(0.2, scale * factor));
      panX = (e.clientX - rect.left) - mx * newScale;
      panY = (e.clientY - rect.top - 44) - my * newScale;
      scale = newScale;
      render();
    }}, {{ passive: false }});

    document.addEventListener('keydown', e => {{
      if (e.target.tagName === 'INPUT' || e.target.tagName === 'SELECT') return;
      if (e.key === 'Delete' || e.key === 'Backspace') {{ deleteSelected(); e.preventDefault(); }}
      if (e.key === 'r' || e.key === 'R') rotateSelected();
      if ((e.key === 'z' || e.key === 'Z') && (e.metaKey || e.ctrlKey)) {{ undoAction(); e.preventDefault(); }}
      if ((e.key === 'a' || e.key === 'A') && (e.metaKey || e.ctrlKey)) {{ selectAll(); e.preventDefault(); }}
      if (e.key === 'Escape') {{ placingType = null; placingFacility = null; selection = []; render(); }}
    }});

    // ── 선택/조작 ──
    function rotateSelected() {{
      if (selection.length === 0) return;
      saveUndo();
      selection.forEach(s => {{
        if (s.type === 'fixture') {{
          const fx = fixtures.find(f => f.id === s.id);
          if (fx) fx.orient = fx.orient === 'V' ? 'H' : 'V';
        }} else {{
          const fac = facilities.find(f => f.id === s.id);
          if (fac) {{ const tmp = fac.w; fac.w = fac.h; fac.h = tmp; }}
        }}
      }});
      render();
    }}

    function deleteSelected() {{
      if (selection.length === 0) return;
      saveUndo();
      const fxIds = new Set(selection.filter(s => s.type === 'fixture').map(s => s.id));
      const facIds = new Set(selection.filter(s => s.type === 'facility').map(s => s.id));
      fixtures = fixtures.filter(f => !fxIds.has(f.id));
      facilities = facilities.filter(f => !facIds.has(f.id));
      selection = [];
      render();
    }}

    // ── 리사이즈 ──
    function startResize(e, fac, handle) {{
      isResizing = true;
      resizeFac = fac;
      resizeHandle = handle;
      const rect = root.getBoundingClientRect();
      const [mx, my] = fromSVG(e.clientX - rect.left, e.clientY - rect.top - 44);
      resizeStartX = mx; resizeStartY = my;
      resizeOrigX = fac.x; resizeOrigY = fac.y;
      resizeOrigW = fac.w; resizeOrigH = fac.h;
      saveUndo();
    }}

    // Ctrl+A: 전체 선택
    function selectAll() {{
      selection = [];
      fixtures.forEach(f => selection.push({{ type: 'fixture', id: f.id }}));
      facilities.forEach(f => selection.push({{ type: 'facility', id: f.id }}));
      render();
    }}

    // ── 추가 ──
    function addFixture(type) {{
      placingType = type; placingFacility = null;
      document.getElementById('statusText').textContent = type + ' 배치 중 — 클릭으로 위치 지정 (Esc 취소)';
    }}
    function placeNewFixture(x, y) {{
      if (!placingType) return; saveUndo();
      const type = placingType;
      const existing = fixtures.filter(f => f.type === type);
      const no = existing.length > 0 ? Math.max(...existing.map(f => f.no)) + 1 : 1;
      fixtures.push({{ id: type+'-'+no, type, no, x, y, orient:'V', zone:'상단', label:'' }});
      placingType = null;
      selection = [{{ type: 'fixture', id: type+'-'+no }}]; render();
    }}
    function addFacilityFromSelect(sel) {{
      const name = sel.value; if (!name) return; sel.value = '';
      placingFacility = name; placingType = null;
      document.getElementById('statusText').textContent = name + ' 배치 중 — 클릭으로 위치 지정 (Esc 취소)';
    }}
    function placeNewFacility(x, y) {{
      if (!placingFacility) return; saveUndo();
      const name = placingFacility;
      const ft = FACILITY_TYPES[name] || FACILITY_TYPES['기타'];
      const maxNo = facilities.length > 0 ? Math.max(...facilities.map(f => parseInt(f.id.split('-')[1]) || 0)) : 0;
      const id = 'fac-' + (maxNo + 1);
      facilities.push({{ id, name, x, y, w: ft.w, h: ft.h, label: '' }});
      placingFacility = null;
      selection = [{{ type: 'facility', id }}]; render();
    }}

    // ── Undo ──
    function saveUndo() {{
      undoStack.push(JSON.stringify({{ fixtures, facilities }}));
      if (undoStack.length > 50) undoStack.shift();
    }}
    function undoAction() {{
      if (undoStack.length === 0) return;
      const state = JSON.parse(undoStack.pop());
      fixtures = state.fixtures || []; facilities = state.facilities || [];
      selection = []; render();
    }}

    function updateStatus() {{
      if (!placingType && !placingFacility) {{
        const a = fixtures.filter(f=>f.type==='A').length;
        const b = fixtures.filter(f=>f.type==='B').length;
        const c = fixtures.filter(f=>f.type==='C').length;
        let msg = 'A:'+a+' B:'+b+' C:'+c+' (총 '+(a+b+c)+'대)';
        if (selection.length > 0) msg += ' | 선택: ' + selection.length + '개';
        msg += ' | 드래그:화면이동 Shift+드래그:범위선택 Shift+클릭:추가선택 Ctrl+A:전체';
        document.getElementById('statusText').textContent = msg;
      }}
    }}

    // ── DB 저장 (API 호출) ──
    function saveToStreamlit() {{
      const btn = document.getElementById('saveBtn');
      btn.textContent = '저장 중...';
      btn.style.background = '#888';

      const data = JSON.stringify({{
        fixtures: fixtures.map(f => ({{ id:f.id, type:f.type, no:f.no, x:Math.round(f.x), y:Math.round(f.y), orient:f.orient, zone:f.zone||'', label:f.label||'' }})),
        facilities: facilities.map(f => ({{ id:f.id, name:f.name, x:Math.round(f.x), y:Math.round(f.y), w:Math.round(f.w), h:Math.round(f.h), label:f.label||'' }})),
      }});

      fetch('http://localhost:8503/save-layout', {{
        method: 'POST',
        headers: {{ 'Content-Type': 'application/json' }},
        body: data,
      }})
      .then(r => r.json())
      .then(res => {{
        btn.textContent = '저장 완료! (' + (res.count || 0) + '대)';
        btn.style.background = '#27ae60';
        setTimeout(() => {{ btn.textContent = 'DB에 저장'; btn.style.background = '#4A90D9'; }}, 2500);
      }})
      .catch(err => {{
        btn.textContent = '저장 실패!';
        btn.style.background = '#e74c3c';
        setTimeout(() => {{ btn.textContent = 'DB에 저장'; btn.style.background = '#4A90D9'; }}, 3000);
      }});
    }}

    // ── 초기화 ──
    function fitView() {{
      const ww = root.clientWidth, wh = root.clientHeight - 44;
      const sx = (ww - 40) / STORE_W, sy = (wh - 40) / STORE_H;
      scale = Math.min(sx, sy);
      panX = (ww - STORE_W * scale) / 2;
      panY = (wh - STORE_H * scale) / 2;
    }}

    fitView();
    render();
    window.addEventListener('resize', () => {{ fitView(); render(); }});
    </script>
    """

    # Streamlit에 에디터 임베드
    from streamlit.components.v1 import html as st_html

    st.info("드래그: 화면이동 | Shift+드래그: 범위선택 | Shift+클릭: 추가선택 | Ctrl+A: 전체 | R: 회전 | Del: 삭제 | 스크롤: 확대/축소")

    st_html(editor_html, height=540, scrolling=False)

    # 매대 목록 (하단)
    st.markdown("---")
    st.subheader("매대 목록")

    col_f1, col_f2 = st.columns(2)
    with col_f1:
        type_filter = st.selectbox("타입 필터", ["전체", "A", "B", "C"], key="fp_type_filter")

    # editor_fixtures 리스트를 DataFrame으로 변환
    _fx_df = pd.DataFrame(editor_fixtures)
    if not _fx_df.empty:
        _fx_df.rename(columns={"x": "X (mm)", "y": "Y (mm)", "orient": "방향", "zone": "존"}, inplace=True)
        _fx_df["라벨"] = _fx_df.apply(lambda r: r.get("label") or r.get("id", ""), axis=1)
        _fx_df["타입명"] = _fx_df["type"].map(lambda t: f"{t} ({SHELF_CONFIGS.get(t, {}).get('name', '')})")

        disp = _fx_df if type_filter == "전체" else _fx_df[_fx_df["type"] == type_filter]

        st.dataframe(
            disp[["라벨", "타입명", "X (mm)", "Y (mm)", "방향", "존"]],
            use_container_width=True, hide_index=True,
        )
        st.caption(f"표시: {len(disp)}대 / 전체: {len(_fx_df)}대")


# ======================================================================
# 탭 1: 배치 관리
# ======================================================================
elif menu == "✏️ 배치 관리":
    st.title("✏️ 배치 관리")

    tab_add, tab_end, tab_bulk, tab_tier_cfg = st.tabs(["➕ 배치 입력", "🔄 배치 종료/변경", "📤 일괄 입력", "⚙️ 매대 단 설정"])

    # --- 배치 입력 ---
    with tab_add:
        st.subheader("새 배치 입력")

        locations = get_all_locations()
        products_df = load_product_list()

        col1, col2 = st.columns(2)

        with col1:
            sel_type = st.selectbox("매대 타입", list(SHELF_CONFIGS.keys()),
                                    format_func=lambda x: f"{x} ({SHELF_CONFIGS[x]['name']})",
                                    key="add_type")
            type_locs = locations[locations["shelf_type"] == sel_type]

            fixture_options = sorted(type_locs["fixture_no"].unique())
            sel_fixture = st.selectbox(
                "매대 번호",
                fixture_options,
                format_func=lambda x: f"{sel_type}-{x}",
                key="add_fixture",
            )

            fixture_locs = type_locs[type_locs["fixture_no"] == sel_fixture]
            tier_face_options = fixture_locs[["id", "tier", "display_label"]].to_dict("records")
            sel_location = st.selectbox(
                "단",
                tier_face_options,
                format_func=lambda x: f"{x['tier']}단" + (" (무제한)" if SHELF_CONFIGS[sel_type]['tiers'][x['tier']-1] >= 999 else f" ({SHELF_CONFIGS[sel_type]['tiers'][x['tier']-1]}cm)"),
                key="add_tier",
            )

            pos_col1, pos_col2 = st.columns(2)
            with pos_col1:
                sel_pos_start = st.number_input("시작 위치 (왼쪽부터)", min_value=1, value=1, step=1, key="add_pos_start")
            with pos_col2:
                sel_pos_end = st.number_input("끝 위치", min_value=1, value=1, step=1, key="add_pos_end")

        with col2:
            if not products_df.empty:
                product_names = sorted(products_df["name"].dropna().unique().tolist())
                sel_product = st.selectbox("상품 선택", product_names, index=None,
                                           placeholder="상품명 검색/선택...", key="add_product")
                matched = products_df[products_df["name"] == sel_product] if sel_product else pd.DataFrame()
                auto_product_id = matched.iloc[0]["id"] if not matched.empty else None
                auto_category = matched.iloc[0].get("erp_category", "") if not matched.empty else ""
            else:
                sel_product = st.text_input("상품명 (직접 입력)", key="add_product_text")
                auto_product_id = None
                auto_category = st.text_input("카테고리", key="add_category_text")

            sel_start_date = st.date_input("배치 시작일", value=date.today(), key="add_start_date")
            sel_notes = st.text_input("메모 (선택)", key="add_notes")

        if st.button("배치 등록", type="primary", key="btn_add"):
            if sel_location and sel_product:
                pid = add_placement(
                    shelf_location_id=sel_location["id"],
                    product_name=sel_product,
                    start_date=sel_start_date,
                    product_id=auto_product_id,
                    erp_category=auto_category,
                    notes=sel_notes if sel_notes else None,
                    position_start=sel_pos_start,
                    position_end=sel_pos_end,
                )
                st.success(f"배치 등록 완료! (ID: {pid}) — {sel_location['display_label']} <- {sel_product}")
                st.rerun()
            else:
                st.warning("위치와 상품을 모두 선택해 주세요.")

        # --- 현재 배치 목록 (페이지네이션) ---
        st.markdown("---")
        st.subheader("현재 등록된 배치")
        _cur_placements = get_current_placements()
        if _cur_placements.empty:
            st.info("등록된 배치가 없습니다.")
        else:
            # 위치 라벨 생성
            def _placement_pos_label(row):
                ps = int(row.get("position_start") or 1)
                pe = int(row.get("position_end") or 1)
                pos = f"{ps}번" if ps == pe else f"{ps}~{pe}번"
                return f"{row['display_label']} / {pos}"
            _cur_placements["위치"] = _cur_placements.apply(_placement_pos_label, axis=1)
            _cur_placements = _cur_placements.sort_values("created_at", ascending=False)
            _show_cols = ["위치", "product_name", "erp_category", "start_date"]
            _show_df = _cur_placements[_show_cols].rename(columns={
                "product_name": "상품명", "erp_category": "카테고리", "start_date": "시작일"
            }).reset_index(drop=True)
            _page_size = 10
            _total = len(_show_df)
            if "placement_show_count" not in st.session_state:
                st.session_state.placement_show_count = _page_size
            _show_n = min(st.session_state.placement_show_count, _total)
            st.caption(f"총 {_total}건 중 {_show_n}건 표시")
            st.dataframe(_show_df.iloc[:_show_n], use_container_width=True, hide_index=True)
            if _show_n < _total:
                if st.button(f"더보기 (+{min(_page_size, _total - _show_n)}건)", key="btn_show_more_placements"):
                    st.session_state.placement_show_count = _show_n + _page_size
                    st.rerun()

    # --- 배치 종료/변경 ---
    with tab_end:
        st.subheader("기존 배치 종료")

        current = get_current_placements()
        if current.empty:
            st.info("현재 활성 배치가 없습니다.")
        else:
            placement_options = current[["id", "display_label", "product_name", "start_date", "position_start", "position_end"]].to_dict("records")
            def _end_format(x):
                pos_start = int(x.get("position_start") or 1)
                pos_end = int(x.get("position_end") or 1)
                if pos_start == pos_end:
                    pos_str = f" / {pos_start}번"
                else:
                    pos_str = f" / {pos_start}~{pos_end}번"
                return f"{x['display_label']}{pos_str} — {x['product_name']} (시작: {x['start_date']})"
            sel_placement = st.selectbox(
                "종료할 배치 선택",
                placement_options,
                format_func=_end_format,
                key="end_placement_select",
            )

            end_date_val = st.date_input("종료일", value=date.today(), key="end_date")

            col_end, col_del = st.columns(2)
            with col_end:
                if st.button("배치 종료", type="primary", key="btn_end"):
                    end_placement(sel_placement["id"], end_date_val)
                    st.success(f"배치 종료 완료! {sel_placement['display_label']} — {sel_placement['product_name']}")
                    st.rerun()
            with col_del:
                if st.button("배치 삭제 (잘못 입력 시)", key="btn_delete"):
                    delete_placement(sel_placement["id"])
                    st.success("배치 삭제 완료!")
                    st.rerun()

    # --- 일괄 입력 ---
    with tab_bulk:
        st.subheader("일괄 입력 (CSV/Excel)")

        st.markdown("##### 1. 템플릿 다운로드")
        locations = get_all_locations()
        products_df_bulk = load_product_list()

        # 위치 라벨 → ID 매핑 (업로드 시 사용)
        _loc_label_to_id = dict(zip(locations["display_label"], locations["id"]))

        template_df = pd.DataFrame({
            "위치": locations["display_label"],
            "상품명": "",
            "시작일": date.today().isoformat(),
            "시작위치": 1,
            "끝위치": 1,
            "메모": "",
        })

        buf = BytesIO()
        template_df.to_excel(buf, index=False, engine="openpyxl")
        buf.seek(0)

        # openpyxl로 상품 드롭다운 추가
        from openpyxl import load_workbook
        from openpyxl.worksheet.datavalidation import DataValidation

        wb = load_workbook(buf)
        ws = wb.active
        data_row_count = len(locations) + 1  # 헤더 포함 마지막 행

        # 위치 칼럼(A열) 보호: 드롭다운으로 고정
        loc_labels = locations["display_label"].tolist()
        ws_ref = wb.create_sheet("참조데이터")

        for idx, label in enumerate(loc_labels, start=1):
            ws_ref.cell(row=idx, column=1, value=label)

        loc_dv = DataValidation(
            type="list",
            formula1=f"참조데이터!$A$1:$A${len(loc_labels)}",
            allow_blank=False,
            showErrorMessage=True,
        )
        loc_dv.error = "유효하지 않은 위치입니다. 드롭다운에서 선택해 주세요."
        loc_dv.errorTitle = "위치 선택 오류"
        loc_dv.add(f"A2:A{data_row_count}")
        ws.add_data_validation(loc_dv)

        if not products_df_bulk.empty:
            product_names_sorted = sorted(products_df_bulk["name"].dropna().unique().tolist())

            # 상품 목록을 참조데이터 시트 B열에 기록
            for idx, pname in enumerate(product_names_sorted, start=1):
                ws_ref.cell(row=idx, column=2, value=pname)

            # 상품명 칼럼(B열)에 데이터 유효성 검사 (드롭다운) 추가
            product_count = len(product_names_sorted)
            dv = DataValidation(
                type="list",
                formula1=f"참조데이터!$B$1:$B${product_count}",
                allow_blank=True,
                showErrorMessage=True,
                showInputMessage=True,
            )
            dv.error = "상품목록에 없는 상품입니다. 드롭다운에서 선택해 주세요."
            dv.errorTitle = "상품 선택 오류"
            dv.prompt = "드롭다운에서 상품을 선택하세요"
            dv.promptTitle = "상품 선택"
            dv.add(f"B2:B{data_row_count}")
            ws.add_data_validation(dv)

        ws_ref.sheet_state = "hidden"

        buf2 = BytesIO()
        wb.save(buf2)
        buf2.seek(0)

        st.download_button(
            "템플릿 다운로드 (.xlsx)",
            data=buf2,
            file_name="shelf_placement_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        if not products_df_bulk.empty:
            st.caption(f"💡 상품명 칼럼에서 {len(product_names_sorted)}개 상품 중 선택 가능 (드롭다운)")
        else:
            st.caption("⚠️ 상품 목록을 불러올 수 없어 직접 입력이 필요합니다.")

        st.markdown("##### 2. 파일 업로드")
        uploaded = st.file_uploader("CSV 또는 Excel 파일", type=["csv", "xlsx"], key="bulk_upload")

        if uploaded:
            if uploaded.name.endswith(".csv"):
                upload_df = pd.read_csv(uploaded)
            else:
                upload_df = pd.read_excel(uploaded, sheet_name=0, engine="openpyxl")

            # 새 양식(한글 칼럼)과 구 양식(영문 칼럼) 모두 지원
            _col_map = {
                "위치": "display_label",
                "상품명": "product_name",
                "시작일": "start_date",
                "시작위치": "position_start",
                "끝위치": "position_end",
                "메모": "notes",
            }
            upload_df = upload_df.rename(columns=_col_map)

            # display_label → shelf_location_id 자동 매핑
            _locations_for_map = get_all_locations()
            _label_to_id = dict(zip(_locations_for_map["display_label"], _locations_for_map["id"]))

            if "display_label" in upload_df.columns and "shelf_location_id" not in upload_df.columns:
                upload_df["shelf_location_id"] = upload_df["display_label"].map(_label_to_id)
                _unmapped = upload_df[upload_df["shelf_location_id"].isna() & upload_df["display_label"].notna()]
                if not _unmapped.empty:
                    st.error(f"매핑 실패한 위치가 {len(_unmapped)}건 있습니다: {_unmapped['display_label'].tolist()[:5]}")

            valid = upload_df[upload_df["product_name"].notna() & (upload_df["product_name"] != "")]

            # 상품 목록에서 product_id, erp_category 자동 매핑
            _bulk_products = load_product_list()
            if not _bulk_products.empty and not valid.empty:
                _pmap = _bulk_products.set_index("name")
                valid = valid.copy()
                for col_name, map_field in [("product_id", "id"), ("erp_category", "erp_category")]:
                    def _resolve(row, field=map_field):
                        pname = row["product_name"]
                        if pname in _pmap.index:
                            return str(_pmap.loc[pname].get(field, ""))
                        return None
                    valid[col_name] = valid.apply(_resolve, axis=1)

            # 표시용 칼럼 정리
            _show_cols = ["display_label", "product_name", "start_date", "position_start", "position_end", "notes"]
            _show_cols = [c for c in _show_cols if c in valid.columns]
            st.dataframe(valid[_show_cols].head(20), use_container_width=True, hide_index=True)
            st.caption(f"유효 행: {len(valid)}개")

            if st.button("일괄 등록", type="primary", key="btn_bulk"):
                # shelf_location_id 매핑 확인
                if "shelf_location_id" not in valid.columns or valid["shelf_location_id"].isna().any():
                    st.error("일부 위치를 매핑할 수 없습니다. 위치 칼럼을 확인해 주세요.")
                else:
                    records = []
                    for _, row in valid.iterrows():
                        records.append({
                            "shelf_location_id": int(row["shelf_location_id"]),
                            "product_name": str(row["product_name"]),
                            "product_id": str(row.get("product_id", "")) if pd.notna(row.get("product_id")) else None,
                            "erp_category": str(row.get("erp_category", "")) if pd.notna(row.get("erp_category")) else None,
                            "start_date": str(row.get("start_date", date.today().isoformat())),
                            "position_start": int(row["position_start"]) if pd.notna(row.get("position_start")) else 1,
                            "position_end": int(row["position_end"]) if pd.notna(row.get("position_end")) else 1,
                            "notes": str(row["notes"]) if pd.notna(row.get("notes")) else None,
                        })
                    count = bulk_add_placements(records)
                    st.success(f"일괄 등록 완료! {count}건")
                    st.rerun()

    # --- 매대 단 설정 ---
    with tab_tier_cfg:
        st.subheader("매대별 사용 단 설정")
        st.caption("사용하지 않는 단을 비활성화하면 템플릿·배치율 계산에서 제외됩니다.")

        all_locs = get_all_locations(include_disabled=True)

        # ── 매대 타입별 활성 단 통계 ──
        st.markdown("##### 📊 매장 전체 단 활용 현황")
        stat_rows = []
        for stype, scfg in SHELF_CONFIGS.items():
            type_locs = all_locs[all_locs["shelf_type"] == stype]
            total = len(type_locs)
            active_locs = type_locs[type_locs["enabled"] == 1]
            active = len(active_locs)
            inactive = total - active
            fixture_count = type_locs["fixture_no"].nunique()
            tiers_per = len(scfg["tiers"])
            width_cm = scfg["width"]

            # 진열 길이: 단 수 × 매대 폭(cm) → m
            len_total = total * width_cm / 100
            len_active = active * width_cm / 100

            stat_rows.append({
                "타입": f"{stype} ({scfg['name']})",
                "매대 수": fixture_count,
                "단/매대": tiers_per,
                "폭(cm)": width_cm,
                "전체 단": total,
                "활성 단": active,
                "비활성 단": inactive,
                "활용률": f"{active / total * 100:.0f}%" if total > 0 else "-",
                "전체 길이(m)": round(len_total, 1),
                "활용 길이(m)": round(len_active, 1),
            })
        total_all = sum(r["전체 단"] for r in stat_rows)
        active_all = sum(r["활성 단"] for r in stat_rows)
        inactive_all = sum(r["비활성 단"] for r in stat_rows)
        len_total_all = sum(r["전체 길이(m)"] for r in stat_rows)
        len_active_all = sum(r["활용 길이(m)"] for r in stat_rows)
        stat_rows.append({
            "타입": "합계",
            "매대 수": sum(r["매대 수"] for r in stat_rows),
            "단/매대": "-",
            "폭(cm)": "-",
            "전체 단": total_all,
            "활성 단": active_all,
            "비활성 단": inactive_all,
            "활용률": f"{active_all / total_all * 100:.0f}%" if total_all > 0 else "-",
            "전체 길이(m)": round(len_total_all, 1),
            "활용 길이(m)": round(len_active_all, 1),
        })

        import pandas as _pd_stat
        stat_df = _pd_stat.DataFrame(stat_rows)
        st.dataframe(stat_df, use_container_width=True, hide_index=True)

        # ── 매대별 활성 단 한눈에 보기 ──
        st.markdown("##### 🗂️ 매대별 활성 단 현황")
        overview_type = st.selectbox(
            "타입 필터", ["전체"] + list(SHELF_CONFIGS.keys()),
            format_func=lambda x: x if x == "전체" else f"{x} ({SHELF_CONFIGS[x]['name']})",
            key="tier_overview_type",
        )
        overview_locs = all_locs if overview_type == "전체" else all_locs[all_locs["shelf_type"] == overview_type]
        overview_grp = overview_locs.groupby(["shelf_type", "fixture_no"]).agg(
            활성=("enabled", lambda x: sum(x == 1)),
            전체=("enabled", "count"),
            활성단목록=("tier", lambda x: ", ".join(
                f"{t}단" for t, e in sorted(zip(x, overview_locs.loc[x.index, "enabled"])) if e == 1
            )),
        ).reset_index()
        overview_grp["매대"] = overview_grp.apply(lambda r: f"{r['shelf_type']}-{r['fixture_no']}", axis=1)
        overview_grp["상태"] = overview_grp.apply(
            lambda r: "✅ 전체 활성" if r["활성"] == r["전체"] else f"⚠️ {r['활성']}/{r['전체']}단", axis=1
        )
        st.dataframe(
            overview_grp[["매대", "상태", "활성단목록"]].rename(columns={"활성단목록": "활성 단"}),
            use_container_width=True, hide_index=True,
        )

        st.markdown("---")

        # ── 개별 매대 단 설정 ──
        st.markdown("##### ⚙️ 매대 단 설정 변경")
        cfg_col1, cfg_col2 = st.columns(2)
        with cfg_col1:
            cfg_type = st.selectbox(
                "매대 타입", list(SHELF_CONFIGS.keys()),
                format_func=lambda x: f"{x} ({SHELF_CONFIGS[x]['name']})",
                key="tier_cfg_type",
            )
        with cfg_col2:
            type_fixtures = sorted(all_locs[all_locs["shelf_type"] == cfg_type]["fixture_no"].unique())
            cfg_fixture = st.selectbox(
                "매대 번호", type_fixtures,
                format_func=lambda x: f"{cfg_type}-{x}",
                key="tier_cfg_fixture",
            )

        # 현재 상태 조회
        tier_status = get_fixture_tier_status(cfg_type, cfg_fixture)
        total_tiers = len(SHELF_CONFIGS[cfg_type]["tiers"])
        current_active = sorted(tier_status[tier_status["enabled"] == 1]["tier"].tolist()) if not tier_status.empty else list(range(1, total_tiers + 1))

        st.markdown(f"**{cfg_type}-{cfg_fixture}** — 활성 {len(current_active)}단 / 전체 {total_tiers}단")

        # 단 옵션 목록 생성
        tier_options = []
        tier_labels = {}
        for i in range(total_tiers):
            tier_num = i + 1
            tier_height = SHELF_CONFIGS[cfg_type]["tiers"][i]
            if tier_height >= 999:
                tier_labels[tier_num] = f"{tier_num}단 (무제한)"
            else:
                tier_labels[tier_num] = f"{tier_num}단 ({tier_height}cm)"
            tier_options.append(tier_num)

        _ms_key = f"tier_ms_{cfg_type}_{cfg_fixture}"

        # 저장 콜백: 버튼 클릭 시 session_state에서 multiselect 값을 읽어 DB 저장
        def _save_tier_cfg():
            _type = st.session_state.get("tier_cfg_type")
            _fix = int(st.session_state.get("tier_cfg_fixture"))
            tiers = [int(t) for t in st.session_state.get(_ms_key, [])]
            set_fixture_tiers_enabled(_type, _fix, tiers)
            st.session_state["_tier_save_msg"] = (
                f"{_type}-{_fix}: {len(tiers)}개 단 활성화 완료"
            )
            if _ms_key in st.session_state:
                del st.session_state[_ms_key]

        st.multiselect(
            "사용할 단 선택",
            options=tier_options,
            default=current_active,
            format_func=lambda x: tier_labels[x],
            key=_ms_key,
        )

        st.button("설정 저장", type="primary", key="btn_save_tier_cfg", on_click=_save_tier_cfg)

        _save_msg = st.session_state.pop("_tier_save_msg", None)
        if _save_msg:
            st.toast(_save_msg)
            st.rerun()


# ======================================================================
# 탭 3: 위치별 성과 분석
# ======================================================================
elif menu == "📊 위치별 성과 분석":
    st.title("📊 위치별 성과 분석")

    col_from, col_to = st.columns(2)
    with col_from:
        analysis_from = st.date_input("분석 시작일", value=date.today() - timedelta(days=30), key="analysis_from")
    with col_to:
        analysis_to = st.date_input("분석 종료일", value=date.today(), key="analysis_to")

    placements = get_current_placements()
    if placements.empty:
        st.warning("배치 데이터가 없습니다. '배치 관리' 탭에서 먼저 배치를 등록해 주세요.")
        st.stop()

    with st.spinner("매출 데이터 로딩 중..."):
        merged = fetch_sales_for_placements(
            date_from=analysis_from.isoformat(),
            date_to=analysis_to.isoformat(),
        )

    if merged.empty:
        st.warning("매출 데이터를 가져올 수 없습니다. Supabase 연결을 확인해 주세요.")
        st.stop()

    days = max(1, (analysis_to - analysis_from).days + 1)
    merged["daily_revenue"] = merged["total_revenue"] / days

    # ── 상품 검색 ──
    _perf_product_names = sorted(merged["product_name"].dropna().unique().tolist())
    _perf_search = st.selectbox(
        "🔍 상품 검색 — 배치도에서 위치 찾기 + 성과 확인",
        _perf_product_names, index=None,
        placeholder="상품명을 입력하세요...",
        key="perf_product_search",
    )
    if _perf_search:
        _pm = merged[merged["product_name"] == _perf_search]
        _perf_hl_fixtures = set()
        _perf_details = []
        for _, r in _pm.iterrows():
            fid = f"{r['shelf_type']}-{int(r['fixture_no'])}"
            _perf_hl_fixtures.add(fid)
            ps = int(r.get("position_start") or 1)
            pe = int(r.get("position_end") or 1)
            pos_str = f"{ps}번" if ps == pe else f"{ps}~{pe}번"
            _perf_details.append({
                "위치": f"{fid} / {int(r['tier'])}단 / {pos_str}",
                "총매출": int(r["total_revenue"]),
                "일평균": int(r["daily_revenue"]),
            })
        st.info(f"📍 **{_perf_search}** — {len(_perf_details)}개 위치에 배치됨")
        st.dataframe(
            pd.DataFrame(_perf_details).style.format({"총매출": "{:,.0f}원", "일평균": "{:,.0f}원"}),
            use_container_width=True, hide_index=True,
        )
        _perf_hl_json = _json.dumps(list(_perf_hl_fixtures), ensure_ascii=False)
    else:
        _perf_hl_json = "[]"

    st.markdown("---")

    # ── fixture 위치 데이터 로드 ──
    _perf_layout = None
    if LAYOUT_FILE.exists():
        try:
            with open(str(LAYOUT_FILE), "r", encoding="utf-8") as f:
                _perf_layout = _json.load(f)
        except Exception:
            pass

    if _perf_layout and _perf_layout.get("fixtures"):
        perf_fixtures = _perf_layout["fixtures"]
        perf_facilities = _perf_layout.get("facilities", [])
    else:
        _pf_df = get_fixture_positions()
        perf_fixtures = []
        for _, row in _pf_df.iterrows():
            perf_fixtures.append({
                "id": f"{row['shelf_type']}-{int(row['fixture_no'])}",
                "type": row["shelf_type"],
                "no": int(row["fixture_no"]),
                "x": float(row["x_pos"]),
                "y": float(row["y_pos"]),
                "orient": row["orientation"],
            })
        perf_facilities = []

    merged["fixture_id"] = merged["shelf_type"] + "-" + merged["fixture_no"].astype(str)

    # ── 매대별 매출 집계 ──
    fx_sales = merged.groupby("fixture_id").agg(
        total_revenue=("total_revenue", "sum"),
        daily_revenue=("daily_revenue", "sum"),
        product_count=("product_name", "nunique"),
    ).reset_index()
    fx_cat = merged.groupby("fixture_id")["erp_category"].agg(
        lambda x: x.value_counts().index[0] if len(x.dropna()) > 0 else ""
    ).reset_index().rename(columns={"erp_category": "top_category"})
    fx_sales = fx_sales.merge(fx_cat, on="fixture_id", how="left")
    max_rev = fx_sales["daily_revenue"].max() if not fx_sales.empty else 1
    fx_sales_map = {r["fixture_id"]: r for _, r in fx_sales.iterrows()}

    # 매대별 상세 데이터
    detail_by_fx = {}
    for fid, grp in merged.groupby("fixture_id"):
        rows = []
        for _, r in grp.iterrows():
            ps = int(r.get("position_start") or 1)
            pe = int(r.get("position_end") or 1)
            pos_str = f"{ps}번" if ps == pe else f"{ps}~{pe}번"
            rows.append({
                "단": f"{int(r['tier'])}단",
                "위치": pos_str,
                "상품": r["product_name"],
                "카테고리": r.get("erp_category", ""),
                "총매출": int(r["total_revenue"]),
                "일평균": int(r["daily_revenue"]),
            })
        detail_by_fx[fid] = rows

    # 카테고리별 색상
    all_cats = sorted(set(v.get("top_category", "") for v in fx_sales_map.values()) - {""})
    _palette = ["#FF6B6B","#4ECDC4","#45B7D1","#96CEB4","#FFEAA7","#DDA0DD","#98D8C8","#F7DC6F","#BB8FCE","#85C1E9","#F0B27A","#82E0AA"]
    cat_colors = {c: _palette[i % len(_palette)] for i, c in enumerate(all_cats)}

    _pf_json = _json.dumps(perf_fixtures, ensure_ascii=False)
    _pfac_json = _json.dumps(perf_facilities, ensure_ascii=False)
    _fx_sales_json = _json.dumps({k: {
        "total_revenue": int(v["total_revenue"]),
        "daily_revenue": int(v["daily_revenue"]),
        "product_count": int(v["product_count"]),
        "top_category": v.get("top_category", ""),
    } for k, v in fx_sales_map.items()}, ensure_ascii=False)
    _max_rev_val = int(max_rev) if max_rev > 0 else 1
    _cat_colors_json = _json.dumps(cat_colors, ensure_ascii=False)

    # ── 성과 맵 HTML ──
    perf_map_html = f"""
    <div id="perf-root" style="width:100%;height:480px;position:relative;background:#1a1a2e;border:1px solid #333;border-radius:8px;overflow:hidden;">
      <div id="perf-legend" style="position:absolute;bottom:8px;left:8px;z-index:10;background:rgba(0,0,0,0.7);color:#fff;padding:8px 12px;border-radius:6px;font-size:11px;max-width:200px;cursor:grab;user-select:none;">
        <div style="font-weight:bold;margin-bottom:4px;">카테고리 범례 <span style="font-weight:normal;color:#888;font-size:9px;">⠿ 드래그</span></div>
        <div id="legend-items"></div>
        <div style="margin-top:6px;border-top:1px solid #555;padding-top:4px;">
          <div style="display:flex;align-items:center;gap:4px;">
            <span style="display:inline-block;width:50px;height:8px;background:linear-gradient(90deg,#2d2d44,#ff4444);border-radius:2px;"></span>
            <span>매출 강도</span>
          </div>
        </div>
      </div>
      <div id="perf-info" style="position:absolute;top:8px;right:8px;z-index:10;background:rgba(0,0,0,0.85);color:#fff;padding:12px 16px;border-radius:8px;font-size:12px;display:none;min-width:240px;box-shadow:0 4px 12px rgba(0,0,0,0.5);">
      </div>
      <svg id="perf-svg" style="display:block;cursor:grab;"></svg>
    </div>
    <script>
    (function() {{
      const STORE_W = {STORE_W}, STORE_H = {STORE_H};
      const fixtures = {_pf_json};
      const facilities = {_pfac_json};
      const fxSales = {_fx_sales_json};
      const maxRev = {_max_rev_val};
      const highlightFixtures = new Set({_perf_hl_json});
      const catColors = {_cat_colors_json};

      const TYPES = {{
        A: {{ name:'기본매대', w:900, d:360 }},
        B: {{ name:'연결매대', w:930, d:360 }},
        C: {{ name:'엔드캡매대', w:636, d:360 }},
      }};
      const FAC_COLORS = {{
        'POS':'#e74c3c','조제실':'#3498db','창고':'#95a5a6',
        '프로모션 존':'#e67e22','대기 공간':'#1abc9c','냉장고':'#2980b9','약품 수납장':'#8e44ad',
      }};

      const svgNS = 'http://www.w3.org/2000/svg';
      const root = document.getElementById('perf-root');
      const svg = document.getElementById('perf-svg');
      const info = document.getElementById('perf-info');
      let scale = 0.055, panX = 20, panY = 20;
      let dragging = false, dsx = 0, dsy = 0, psx = 0, psy = 0;
      let selectedFx = null;

      // 범례
      const legendEl = document.getElementById('legend-items');
      Object.entries(catColors).forEach(([cat, color]) => {{
        const d = document.createElement('div');
        d.style.cssText = 'display:flex;align-items:center;gap:4px;margin:1px 0;';
        d.innerHTML = '<span style="display:inline-block;width:10px;height:10px;border-radius:2px;background:'+color+';flex-shrink:0;"></span><span>'+cat+'</span>';
        legendEl.appendChild(d);
      }});
      if (Object.keys(catColors).length === 0) {{
        legendEl.innerHTML = '<span style="color:#888;">배치 카테고리 없음</span>';
      }}

      // 범례 드래그
      (function() {{
        const legend = document.getElementById('perf-legend');
        const parent = legend.parentElement;
        let ldrag = false, offX = 0, offY = 0;
        legend.addEventListener('mousedown', function(e) {{
          ldrag = true;
          const lr = legend.getBoundingClientRect();
          offX = e.clientX - lr.left;
          offY = e.clientY - lr.top;
          legend.style.cursor = 'grabbing';
          e.stopPropagation();
          e.preventDefault();
        }});
        document.addEventListener('mousemove', function(e) {{
          if (!ldrag) return;
          const pr = parent.getBoundingClientRect();
          legend.style.left = (e.clientX - pr.left - offX) + 'px';
          legend.style.top = (e.clientY - pr.top - offY) + 'px';
          legend.style.bottom = 'auto';
          legend.style.right = 'auto';
        }});
        document.addEventListener('mouseup', function() {{
          if (ldrag) {{ ldrag = false; legend.style.cursor = 'grab'; }}
        }});
      }})();

      function toSVG(mx, my) {{ return [mx * scale + panX, my * scale + panY]; }}

      function intensity(rev) {{ return maxRev > 0 ? Math.min(1, rev / maxRev) : 0; }}

      function heatColor(t) {{
        const r = Math.round(45 + t * 210);
        const g = Math.round(45 + Math.max(0, 0.3 - t) * 80);
        const b = Math.round(68 - t * 40);
        return 'rgb('+r+','+g+','+b+')';
      }}

      function fmt(n) {{ return n.toLocaleString('ko-KR'); }}

      function render() {{
        svg.innerHTML = '';
        const ww = root.clientWidth, wh = root.clientHeight;
        svg.setAttribute('width', ww); svg.setAttribute('height', wh);

        // 배경
        const bg = document.createElementNS(svgNS, 'rect');
        bg.setAttribute('width', ww); bg.setAttribute('height', wh);
        bg.setAttribute('fill', '#1a1a2e');
        svg.appendChild(bg);

        // 매장 외곽
        const [sx, sy] = toSVG(0, 0);
        const sw = STORE_W * scale, sh = STORE_H * scale;
        const store = document.createElementNS(svgNS, 'rect');
        store.setAttribute('x', sx); store.setAttribute('y', sy);
        store.setAttribute('width', sw); store.setAttribute('height', sh);
        store.setAttribute('fill', '#16213e'); store.setAttribute('stroke', '#444');
        store.setAttribute('stroke-width', 1); store.setAttribute('rx', 4);
        svg.appendChild(store);

        // 시설물 (반투명)
        facilities.forEach(fac => {{
          const [fx, fy] = toSVG(fac.x, fac.y);
          const fw = fac.w * scale, fh = fac.h * scale;
          const c = FAC_COLORS[fac.name] || '#666';
          const r = document.createElementNS(svgNS, 'rect');
          r.setAttribute('x', fx); r.setAttribute('y', fy);
          r.setAttribute('width', fw); r.setAttribute('height', fh);
          r.setAttribute('fill', c); r.setAttribute('opacity', 0.12);
          r.setAttribute('stroke', c); r.setAttribute('stroke-width', 0.5);
          r.setAttribute('stroke-opacity', 0.3); r.setAttribute('rx', 3);
          svg.appendChild(r);
          const t = document.createElementNS(svgNS, 'text');
          t.setAttribute('x', fx + fw/2); t.setAttribute('y', fy + fh/2);
          t.setAttribute('text-anchor', 'middle'); t.setAttribute('dominant-baseline', 'central');
          t.setAttribute('font-size', Math.max(6, Math.min(10, fw * 0.12)));
          t.setAttribute('fill', c); t.setAttribute('opacity', 0.4);
          t.setAttribute('font-family', '-apple-system, sans-serif');
          t.textContent = fac.name;
          svg.appendChild(t);
        }});

        // 글로우 필터
        const defs = document.createElementNS(svgNS, 'defs');
        defs.innerHTML = '<filter id="glow"><feGaussianBlur stdDeviation="3" result="b"/><feMerge><feMergeNode in="b"/><feMergeNode in="SourceGraphic"/></feMerge></filter>';
        svg.appendChild(defs);

        // 매대 렌더링
        fixtures.forEach(fx => {{
          const tp = TYPES[fx.type];
          if (!tp) return;
          const dx = (fx.orient === 'V' ? tp.d : tp.w) * scale;
          const dy = (fx.orient === 'V' ? tp.w : tp.d) * scale;
          const [rx, ry] = toSVG(fx.x, fx.y);
          const sales = fxSales[fx.id];
          const isSel = selectedFx === fx.id;
          const isHL = highlightFixtures.has(fx.id);

          const g = document.createElementNS(svgNS, 'g');
          g.style.cursor = 'pointer';

          // 검색 하이라이트: 글로우 테두리
          if (isHL) {{
            const glow = document.createElementNS(svgNS, 'rect');
            glow.setAttribute('x', rx - 4); glow.setAttribute('y', ry - 4);
            glow.setAttribute('width', dx + 8); glow.setAttribute('height', dy + 8);
            glow.setAttribute('rx', 4);
            glow.setAttribute('fill', 'none'); glow.setAttribute('stroke', '#00FF88');
            glow.setAttribute('stroke-width', 3);
            const anim = document.createElementNS(svgNS, 'animate');
            anim.setAttribute('attributeName', 'opacity');
            anim.setAttribute('values', '0.3;1;0.3'); anim.setAttribute('dur', '1.5s');
            anim.setAttribute('repeatCount', 'indefinite');
            glow.appendChild(anim);
            g.appendChild(glow);
          }}

          // 본체
          const rect = document.createElementNS(svgNS, 'rect');
          rect.setAttribute('x', rx); rect.setAttribute('y', ry);
          rect.setAttribute('width', dx); rect.setAttribute('height', dy);
          rect.setAttribute('rx', 2);

          if (sales) {{
            const t = intensity(sales.daily_revenue);
            const catCol = catColors[sales.top_category] || '#888';
            rect.setAttribute('fill', heatColor(t));
            rect.setAttribute('stroke', isHL ? '#00FF88' : (isSel ? '#fff' : catCol));
            rect.setAttribute('stroke-width', isHL ? 3 : (isSel ? 2.5 : 1.2));
            rect.setAttribute('fill-opacity', 0.85);
            if (isSel || isHL) rect.setAttribute('filter', 'url(#glow)');

            // 카테고리 컬러 바
            const bar = document.createElementNS(svgNS, 'rect');
            bar.setAttribute('x', rx); bar.setAttribute('y', ry);
            bar.setAttribute('width', dx);
            bar.setAttribute('height', Math.max(2, dy * 0.07));
            bar.setAttribute('fill', catCol); bar.setAttribute('rx', 2);
            g.appendChild(bar);
          }} else {{
            rect.setAttribute('fill', '#2d2d44');
            rect.setAttribute('stroke', isSel ? '#fff' : '#555');
            rect.setAttribute('stroke-width', isSel ? 2.5 : 0.8);
            rect.setAttribute('fill-opacity', 0.5);
          }}
          g.appendChild(rect);

          // 라벨
          const fs = Math.max(5, Math.min(9, Math.min(dx, dy) * 0.35));
          const lbl = document.createElementNS(svgNS, 'text');
          lbl.setAttribute('x', rx + dx/2);
          lbl.setAttribute('y', ry + dy/2 - (sales ? fs * 0.4 : 0));
          lbl.setAttribute('text-anchor', 'middle'); lbl.setAttribute('dominant-baseline', 'central');
          lbl.setAttribute('font-size', fs); lbl.setAttribute('fill', '#fff');
          lbl.setAttribute('font-weight', isSel ? 'bold' : 'normal');
          lbl.setAttribute('font-family', '-apple-system, sans-serif');
          lbl.textContent = fx.id;
          g.appendChild(lbl);

          // 매출 서브라벨
          if (sales && dy > 16) {{
            const sub = document.createElementNS(svgNS, 'text');
            sub.setAttribute('x', rx + dx/2);
            sub.setAttribute('y', ry + dy/2 + fs * 0.7);
            sub.setAttribute('text-anchor', 'middle'); sub.setAttribute('dominant-baseline', 'central');
            sub.setAttribute('font-size', Math.max(4, fs * 0.65));
            sub.setAttribute('fill', '#bbb');
            sub.setAttribute('font-family', '-apple-system, sans-serif');
            const rev = sales.daily_revenue;
            sub.textContent = rev >= 10000 ? Math.round(rev/10000) + '만' : fmt(rev);
            g.appendChild(sub);
          }}

          g.addEventListener('click', e => {{
            e.stopPropagation();
            selectedFx = (selectedFx === fx.id) ? null : fx.id;
            showDetail(fx.id);
            render();
          }});
          svg.appendChild(g);
        }});
      }}

      function showDetail(fxId) {{
        if (!fxId) {{ info.style.display = 'none'; return; }}
        const s = fxSales[fxId];
        let h = '<div style="font-weight:bold;font-size:15px;margin-bottom:8px;border-bottom:1px solid #555;padding-bottom:4px;">' + fxId + '</div>';
        if (s) {{
          const catCol = catColors[s.top_category] || '#888';
          h += '<div style="margin-bottom:3px;">총매출: <b style="color:#FF6B6B;">' + fmt(s.total_revenue) + '원</b></div>';
          h += '<div style="margin-bottom:3px;">일평균: <b style="color:#4ECDC4;">' + fmt(s.daily_revenue) + '원</b></div>';
          h += '<div style="margin-bottom:3px;">배치 상품: ' + s.product_count + '개</div>';
          h += '<div><span style="display:inline-block;width:8px;height:8px;border-radius:2px;background:'+catCol+';margin-right:4px;"></span>' + (s.top_category || '-') + '</div>';
          h += '<div style="margin-top:6px;color:#999;font-size:10px;">아래 드롭다운에서 상세 확인</div>';
        }} else {{
          h += '<div style="color:#888;">매출 데이터 없음</div>';
        }}
        info.innerHTML = h;
        info.style.display = 'block';
      }}

      // 팬
      svg.addEventListener('mousedown', e => {{
        const tgt = e.target;
        const isFixture = tgt.closest && tgt.closest('g[style*="pointer"]');
        if (!isFixture) {{
          dragging = true; dsx = e.clientX; dsy = e.clientY; psx = panX; psy = panY;
          svg.style.cursor = 'grabbing';
          selectedFx = null; info.style.display = 'none'; render();
        }}
      }});
      svg.addEventListener('mousemove', e => {{
        if (!dragging) return;
        panX = psx + (e.clientX - dsx); panY = psy + (e.clientY - dsy); render();
      }});
      svg.addEventListener('mouseup', () => {{ dragging = false; svg.style.cursor = 'grab'; }});
      svg.addEventListener('mouseleave', () => {{ dragging = false; svg.style.cursor = 'grab'; }});
      svg.addEventListener('wheel', e => {{
        if (!e.ctrlKey && !e.metaKey) return;  // 일반 스크롤은 페이지 스크롤로 통과
        e.preventDefault();
        const rc = svg.getBoundingClientRect();
        const mx = e.clientX - rc.left, my = e.clientY - rc.top;
        const f = e.deltaY < 0 ? 1.15 : 1/1.15;
        const ns = Math.max(0.02, Math.min(0.2, scale * f));
        panX = mx - (mx - panX) * (ns / scale);
        panY = my - (my - panY) * (ns / scale);
        scale = ns; render();
      }}, {{ passive: false }});

      render();
    }})();
    </script>
    """
    st.caption("💡 맵 줌: **Ctrl+스크롤** (Mac: ⌘+스크롤) | 드래그로 이동 | 매대 클릭으로 상세 보기")
    import streamlit.components.v1 as components
    components.html(perf_map_html, height=500, scrolling=False)

    # ── 맵 아래: 매대 선택 + 상세 ──
    st.markdown("---")

    # 모든 매대 ID 목록 (배치도 기반)
    _all_fx_ids = sorted(set(
        [f["id"] for f in perf_fixtures] + merged["fixture_id"].unique().tolist()
    ))
    sel_fx = st.selectbox("매대 선택 (클릭하여 상세 보기)", ["전체 요약"] + _all_fx_ids, key="perf_fx_select")

    if sel_fx == "전체 요약":
        col_a, col_b = st.columns(2)
        with col_a:
            st.markdown("**매대 타입별 일평균 매출**")
            type_stats = merged.groupby("shelf_type").agg(
                avg_revenue=("daily_revenue", "mean"),
                total_revenue=("total_revenue", "sum"),
                count=("product_name", "count"),
            ).reset_index()
            type_stats["타입"] = type_stats["shelf_type"].map(lambda x: f"{x} ({SHELF_CONFIGS[x]['name']})")
            fig_type = px.bar(type_stats, x="타입", y="avg_revenue",
                              color="shelf_type", text_auto=",.0f",
                              labels={"avg_revenue": "일평균 매출 (원)"})
            fig_type.update_layout(showlegend=False, height=300)
            st.plotly_chart(fig_type, use_container_width=True)
        with col_b:
            st.markdown("**단(Tier)별 일평균 매출**")
            tier_stats = merged.groupby("tier").agg(
                avg_revenue=("daily_revenue", "mean"),
            ).reset_index()
            tier_stats["단"] = tier_stats["tier"].astype(str) + "단"
            fig_tier = px.bar(tier_stats, y="단", x="avg_revenue", orientation="h",
                               text_auto=",.0f", color="avg_revenue",
                               color_continuous_scale="YlOrRd",
                               labels={"avg_revenue": "일평균 매출 (원)"})
            fig_tier.update_layout(yaxis=dict(autorange="reversed"),
                                    coloraxis_showscale=False, height=300)
            st.plotly_chart(fig_tier, use_container_width=True)

        st.markdown("---")
        ranked = merged[merged["total_revenue"] > 0].sort_values("total_revenue", ascending=False)
        col_top, col_bottom = st.columns(2)
        with col_top:
            st.markdown("**TOP 10 매출 선반**")
            top10 = ranked.head(10)[["display_label", "product_name", "erp_category", "total_revenue", "daily_revenue"]].copy()
            top10.columns = ["위치", "상품", "카테고리", "총매출", "일평균매출"]
            st.dataframe(top10.style.format({"총매출": "{:,.0f}", "일평균매출": "{:,.0f}"}),
                         use_container_width=True, hide_index=True)
        with col_bottom:
            st.markdown("**BOTTOM 10 매출 선반**")
            bottom10 = ranked.tail(10)[["display_label", "product_name", "erp_category", "total_revenue", "daily_revenue"]].copy()
            bottom10.columns = ["위치", "상품", "카테고리", "총매출", "일평균매출"]
            st.dataframe(bottom10.style.format({"총매출": "{:,.0f}", "일평균매출": "{:,.0f}"}),
                         use_container_width=True, hide_index=True)
    else:
        # 선택된 매대의 원본 데이터 (merged에서 추출)
        fx_merged = merged[merged["fixture_id"] == sel_fx]
        if fx_merged.empty:
            st.info(f"{sel_fx}에 배치된 상품이 없습니다.")
        else:
            # KPI
            _fx_total = fx_merged["total_revenue"].sum()
            _fx_daily = fx_merged["daily_revenue"].sum()
            _fx_prods = fx_merged["product_name"].nunique()
            kpi1, kpi2, kpi3 = st.columns(3)
            kpi1.metric("총매출", f"{_fx_total:,.0f}원")
            kpi2.metric("일평균 매출", f"{_fx_daily:,.0f}원")
            kpi3.metric("배치 상품수", f"{_fx_prods}개")

            # 매대 정면도
            _stype = sel_fx.split("-")[0]
            _cfg = SHELF_CONFIGS.get(_stype, {})
            _n_tiers = len(_cfg.get("tiers", [5]))
            _shelf_width_cm = _cfg.get("width", 90)
            _DEFAULT_WIDTH_CM = 8.0  # 치수 미등록 상품 기본 가로

            _view_mode = st.radio(
                "정면도 보기 방식",
                ["칸으로 보기", "길이로 보기"],
                horizontal=True,
                key="shelf_view_mode",
                help="칸으로 보기: position 기반 균일 칸 | 길이로 보기: 상품 가로(cm) 비율 반영",
            )

            # ── 공통 데이터 준비 ──
            _all_daily = [float(r["daily_revenue"]) for _, r in fx_merged.iterrows()]

            _max_daily = max(_all_daily) if _all_daily else 1

            if _view_mode == "칸으로 보기":
                # ── 칸으로 보기 (position 기반 균일 칸) ──
                _max_pos = 1
                for _, r in fx_merged.iterrows():
                    pe = int(r.get("position_end") or 1)
                    if pe > _max_pos:
                        _max_pos = pe
                _max_pos = max(_max_pos, 6)

                _grid = {}
                for _, r in fx_merged.iterrows():
                    tier = int(r["tier"])
                    ps = int(r.get("position_start") or 1)
                    pe = int(r.get("position_end") or 1)
                    daily = float(r["daily_revenue"])
                    for pos in range(ps, pe + 1):
                        _grid[(tier, pos)] = {
                            "name": r["product_name"],
                            "category": r.get("erp_category", "") or "",
                            "daily": daily,
                            "total": float(r["total_revenue"]),
                            "span_start": ps, "span_end": pe,
                        }

                _grid_json_data = {}
                for (tier, pos), v in _grid.items():
                    _grid_json_data[f"{tier}-{pos}"] = {
                        "name": v["name"], "category": v["category"],
                        "daily": int(v["daily"]), "total": int(v["total"]),
                        "span_start": v["span_start"], "span_end": v["span_end"],
                    }

                _shelf_detail_html = f"""
                <div id="shelf-detail-root" style="width:100%;background:#1a1a2e;border:1px solid #333;border-radius:8px;overflow-x:auto;padding:16px;">
                  <div style="color:#fff;font-size:16px;font-weight:bold;margin-bottom:12px;font-family:-apple-system,sans-serif;">
                    {sel_fx} 매대 정면도
                    <span style="font-size:12px;color:#888;font-weight:normal;margin-left:8px;">
                      ({_cfg.get('name','')}, {_shelf_width_cm}cm, {_n_tiers}단)
                    </span>
                  </div>
                  <svg id="shelf-detail-svg"></svg>
                  <div id="shelf-tooltip" style="position:fixed;display:none;background:rgba(0,0,0,0.9);color:#fff;padding:10px 14px;border-radius:6px;font-size:12px;pointer-events:none;z-index:9999;max-width:280px;box-shadow:0 4px 12px rgba(0,0,0,0.5);font-family:-apple-system,sans-serif;"></div>
                </div>
                <script>
                (function() {{
                  const nTiers = {_n_tiers};
                  const maxPos = {_max_pos};
                  const maxDaily = {int(_max_daily) if _max_daily > 0 else 1};
                  const grid = {_json.dumps(_grid_json_data, ensure_ascii=False)};
                  const tierHeights = {_json.dumps(_cfg.get('tiers', []))};
                  const catColors = {_cat_colors_json};
                  const svgNS = 'http://www.w3.org/2000/svg';
                  const svg = document.getElementById('shelf-detail-svg');
                  const tooltip = document.getElementById('shelf-tooltip');
                  const cellW = 90, cellH = 70, labelW = 60, padX = 10, padY = 10;
                  const totalW = labelW + maxPos * cellW + padX * 2;
                  const totalH = nTiers * cellH + padY * 2 + 30;
                  svg.setAttribute('width', totalW); svg.setAttribute('height', totalH); svg.style.display = 'block';
                  function heatColor(t) {{ if(t<=0) return '#2d2d44'; return 'rgb('+Math.round(45+t*210)+','+Math.round(45+Math.max(0,0.3-t)*80)+','+Math.round(68-t*40)+')'; }}
                  function fmt(n) {{ return n.toLocaleString('ko-KR'); }}
                  const bg = document.createElementNS(svgNS,'rect'); bg.setAttribute('width',totalW); bg.setAttribute('height',totalH); bg.setAttribute('fill','#1a1a2e'); bg.setAttribute('rx',6); svg.appendChild(bg);
                  for (let p=1;p<=maxPos;p++) {{ const t=document.createElementNS(svgNS,'text'); t.setAttribute('x',padX+labelW+(p-1)*cellW+cellW/2); t.setAttribute('y',padY+12); t.setAttribute('text-anchor','middle'); t.setAttribute('font-size',10); t.setAttribute('fill','#888'); t.setAttribute('font-family','-apple-system,sans-serif'); t.textContent=p+'번'; svg.appendChild(t); }}
                  const rendered = new Set();
                  for (let tier=nTiers;tier>=1;tier--) {{
                    const rowIdx=nTiers-tier, y=padY+24+rowIdx*cellH, th=tierHeights[tier-1], thLabel=th>=999?'무제한':th+'cm';
                    const lbl=document.createElementNS(svgNS,'text'); lbl.setAttribute('x',padX+labelW-6); lbl.setAttribute('y',y+cellH/2); lbl.setAttribute('text-anchor','end'); lbl.setAttribute('dominant-baseline','central'); lbl.setAttribute('font-size',11); lbl.setAttribute('fill','#ccc'); lbl.setAttribute('font-family','-apple-system,sans-serif'); lbl.textContent=tier+'단'; svg.appendChild(lbl);
                    const lblS=document.createElementNS(svgNS,'text'); lblS.setAttribute('x',padX+labelW-6); lblS.setAttribute('y',y+cellH/2+13); lblS.setAttribute('text-anchor','end'); lblS.setAttribute('dominant-baseline','central'); lblS.setAttribute('font-size',8); lblS.setAttribute('fill','#666'); lblS.setAttribute('font-family','-apple-system,sans-serif'); lblS.textContent=thLabel; svg.appendChild(lblS);
                    for (let pos=1;pos<=maxPos;pos++) {{
                      const key=tier+'-'+pos, data=grid[key], x=padX+labelW+(pos-1)*cellW;
                      if (data && rendered.has(data.span_start+'-'+data.span_end+'-'+tier+'-'+data.name)) continue;
                      const g=document.createElementNS(svgNS,'g');
                      if (data) {{
                        const spanW=(data.span_end-data.span_start+1)*cellW, spanX=padX+labelW+(data.span_start-1)*cellW;
                        const intensity=Math.min(1,data.daily/maxDaily), catCol=catColors[data.category]||'#888';
                        rendered.add(data.span_start+'-'+data.span_end+'-'+tier+'-'+data.name);
                        const rect=document.createElementNS(svgNS,'rect'); rect.setAttribute('x',spanX+1); rect.setAttribute('y',y+1); rect.setAttribute('width',spanW-2); rect.setAttribute('height',cellH-2); rect.setAttribute('fill',heatColor(intensity)); rect.setAttribute('stroke',catCol); rect.setAttribute('stroke-width',1.5); rect.setAttribute('rx',4); rect.setAttribute('fill-opacity',0.85); g.appendChild(rect);
                        const bar=document.createElementNS(svgNS,'rect'); bar.setAttribute('x',spanX+1); bar.setAttribute('y',y+1); bar.setAttribute('width',spanW-2); bar.setAttribute('height',3); bar.setAttribute('fill',catCol); bar.setAttribute('rx',4); g.appendChild(bar);
                        const mc=Math.max(3,Math.floor(spanW/9)); let nameText=data.name.length>mc?data.name.substring(0,mc-1)+'..':data.name;
                        const nt=document.createElementNS(svgNS,'text'); nt.setAttribute('x',spanX+spanW/2); nt.setAttribute('y',y+cellH/2-8); nt.setAttribute('text-anchor','middle'); nt.setAttribute('dominant-baseline','central'); nt.setAttribute('font-size',Math.min(10,Math.max(7,spanW/nameText.length*0.85))); nt.setAttribute('fill','#fff'); nt.setAttribute('font-family','-apple-system,sans-serif'); nt.textContent=nameText; g.appendChild(nt);
                        const rt=document.createElementNS(svgNS,'text'); rt.setAttribute('x',spanX+spanW/2); rt.setAttribute('y',y+cellH/2+10); rt.setAttribute('text-anchor','middle'); rt.setAttribute('dominant-baseline','central'); rt.setAttribute('font-size',9); rt.setAttribute('fill','#4ECDC4'); rt.setAttribute('font-weight','bold'); rt.setAttribute('font-family','-apple-system,sans-serif'); rt.textContent=data.daily>=10000?Math.round(data.daily/10000)+'만/일':fmt(data.daily)+'/일'; g.appendChild(rt);
                        if (spanW>60) {{ const ct=document.createElementNS(svgNS,'text'); ct.setAttribute('x',spanX+spanW/2); ct.setAttribute('y',y+cellH/2+24); ct.setAttribute('text-anchor','middle'); ct.setAttribute('dominant-baseline','central'); ct.setAttribute('font-size',7); ct.setAttribute('fill','#999'); ct.setAttribute('font-family','-apple-system,sans-serif'); ct.textContent=data.category.length>8?data.category.substring(0,7)+'..':data.category; g.appendChild(ct); }}
                        g.style.cursor='pointer';
                        g.addEventListener('mouseenter',(e)=>{{ tooltip.innerHTML='<div style="font-weight:bold;margin-bottom:4px;">'+data.name+'</div><div>위치: '+tier+'단 '+(data.span_start===data.span_end?data.span_start+'번':data.span_start+'~'+data.span_end+'번')+'</div><div>카테고리: '+(data.category||'-')+'</div><div style="margin-top:4px;">총매출: <b style="color:#FF6B6B;">'+fmt(data.total)+'원</b></div><div>일평균: <b style="color:#4ECDC4;">'+fmt(data.daily)+'원</b></div>'; tooltip.style.display='block'; }});
                        g.addEventListener('mousemove',(e)=>{{ tooltip.style.left=(e.clientX+12)+'px'; tooltip.style.top=(e.clientY-10)+'px'; }});
                        g.addEventListener('mouseleave',()=>{{ tooltip.style.display='none'; }});
                      }} else {{
                        const rect=document.createElementNS(svgNS,'rect'); rect.setAttribute('x',x+1); rect.setAttribute('y',y+1); rect.setAttribute('width',cellW-2); rect.setAttribute('height',cellH-2); rect.setAttribute('fill','#1e1e36'); rect.setAttribute('stroke','#333'); rect.setAttribute('stroke-width',0.5); rect.setAttribute('rx',4); rect.setAttribute('fill-opacity',0.5); g.appendChild(rect);
                      }}
                      svg.appendChild(g);
                    }}
                    const line=document.createElementNS(svgNS,'line'); line.setAttribute('x1',padX+labelW); line.setAttribute('y1',y+cellH); line.setAttribute('x2',padX+labelW+maxPos*cellW); line.setAttribute('y2',y+cellH); line.setAttribute('stroke','#333'); line.setAttribute('stroke-width',0.5); svg.appendChild(line);
                  }}
                }})();
                </script>
                """
                components.html(_shelf_detail_html, height=_n_tiers * 70 + 80, scrolling=True)

            else:
                # ── 길이로 보기 (width cm 비율) ──
                _dims_all = get_all_dimensions()
                _width_map = {}
                if not _dims_all.empty:
                    _width_map = _dims_all.set_index("product_name")["width"].dropna().to_dict()
                _tier_items = {}
                for _, r in fx_merged.iterrows():
                    tier = int(r["tier"])
                    pname = r["product_name"]
                    w = _width_map.get(pname)
                    has_width = w is not None and w > 0
                    width_cm = w if has_width else _DEFAULT_WIDTH_CM
                    if tier not in _tier_items:
                        _tier_items[tier] = []
                    _tier_items[tier].append({
                        "name": pname, "category": r.get("erp_category", "") or "",
                        "daily": float(r["daily_revenue"]), "total": float(r["total_revenue"]),
                        "width_cm": round(width_cm, 1), "has_width": has_width,
                    })
                _tier_data_json = {}
                _tier_util = {}
                for tier, items in _tier_items.items():
                    _tier_data_json[str(tier)] = items
                    used = sum(it["width_cm"] for it in items)
                    _tier_util[str(tier)] = round(used / _shelf_width_cm * 100, 1)

                _shelf_detail_html = f"""
                <div id="shelf-detail-root" style="width:100%;background:#1a1a2e;border:1px solid #333;border-radius:8px;overflow-x:auto;padding:16px;">
                  <div style="color:#fff;font-size:16px;font-weight:bold;margin-bottom:12px;font-family:-apple-system,sans-serif;">
                    {sel_fx} 매대 정면도
                    <span style="font-size:12px;color:#888;font-weight:normal;margin-left:8px;">
                      ({_cfg.get('name','')}, {_shelf_width_cm}cm, {_n_tiers}단)
                    </span>
                  </div>
                  <svg id="shelf-detail-svg"></svg>
                  <div id="shelf-tooltip" style="position:fixed;display:none;background:rgba(0,0,0,0.9);color:#fff;padding:10px 14px;border-radius:6px;font-size:12px;pointer-events:none;z-index:9999;max-width:280px;box-shadow:0 4px 12px rgba(0,0,0,0.5);font-family:-apple-system,sans-serif;"></div>
                </div>
                <script>
                (function() {{
                  const nTiers = {_n_tiers};
                  const shelfWidthCm = {_shelf_width_cm};
                  const maxDaily = {int(_max_daily) if _max_daily > 0 else 1};
                  const tierData = {_json.dumps(_tier_data_json, ensure_ascii=False)};
                  const tierUtil = {_json.dumps(_tier_util, ensure_ascii=False)};
                  const tierHeights = {_json.dumps(_cfg.get('tiers', []))};
                  const catColors = {_cat_colors_json};
                  const svgNS = 'http://www.w3.org/2000/svg';
                  const svg = document.getElementById('shelf-detail-svg');
                  const tooltip = document.getElementById('shelf-tooltip');
                  const shelfPxWidth = 700;
                  let maxTierCm = shelfWidthCm;
                  for (let t=1;t<=nTiers;t++) {{ const items=tierData[String(t)]||[]; const totalCm=items.reduce((s,it)=>s+it.width_cm,0); if(totalCm>maxTierCm) maxTierCm=totalCm; }}
                  const pxPerCm = shelfPxWidth / maxTierCm;
                  const cellH=70, labelW=60, utilW=50, padX=10, padY=10;
                  const totalW=labelW+shelfPxWidth+utilW+padX*2, totalH=nTiers*cellH+padY*2+20;
                  svg.setAttribute('width',totalW); svg.setAttribute('height',totalH); svg.style.display='block';
                  function heatColor(t) {{ if(t<=0) return '#2d2d44'; return 'rgb('+Math.round(45+t*210)+','+Math.round(45+Math.max(0,0.3-t)*80)+','+Math.round(68-t*40)+')'; }}
                  function fmt(n) {{ return n.toLocaleString('ko-KR'); }}
                  const bg=document.createElementNS(svgNS,'rect'); bg.setAttribute('width',totalW); bg.setAttribute('height',totalH); bg.setAttribute('fill','#1a1a2e'); bg.setAttribute('rx',6); svg.appendChild(bg);
                  for (let tier=nTiers;tier>=1;tier--) {{
                    const rowIdx=nTiers-tier, y=padY+8+rowIdx*cellH, th=tierHeights[tier-1], thLabel=th>=999?'무제한':th+'cm';
                    const items=tierData[String(tier)]||[], util=tierUtil[String(tier)]||0;
                    const lbl=document.createElementNS(svgNS,'text'); lbl.setAttribute('x',padX+labelW-6); lbl.setAttribute('y',y+cellH/2); lbl.setAttribute('text-anchor','end'); lbl.setAttribute('dominant-baseline','central'); lbl.setAttribute('font-size',11); lbl.setAttribute('fill','#ccc'); lbl.setAttribute('font-family','-apple-system,sans-serif'); lbl.textContent=tier+'단'; svg.appendChild(lbl);
                    const lblS=document.createElementNS(svgNS,'text'); lblS.setAttribute('x',padX+labelW-6); lblS.setAttribute('y',y+cellH/2+13); lblS.setAttribute('text-anchor','end'); lblS.setAttribute('dominant-baseline','central'); lblS.setAttribute('font-size',8); lblS.setAttribute('fill','#666'); lblS.setAttribute('font-family','-apple-system,sans-serif'); lblS.textContent=thLabel; svg.appendChild(lblS);
                    const shelfBg=document.createElementNS(svgNS,'rect'); shelfBg.setAttribute('x',padX+labelW); shelfBg.setAttribute('y',y+1); shelfBg.setAttribute('width',shelfPxWidth); shelfBg.setAttribute('height',cellH-2); shelfBg.setAttribute('fill','#1e1e36'); shelfBg.setAttribute('stroke','#333'); shelfBg.setAttribute('stroke-width',0.5); shelfBg.setAttribute('rx',4); shelfBg.setAttribute('fill-opacity',0.5); svg.appendChild(shelfBg);
                    let offsetX=padX+labelW;
                    for (let i=0;i<items.length;i++) {{
                      const item=items[i], itemPxW=Math.max(20,item.width_cm*pxPerCm);
                      const intensity=Math.min(1,item.daily/maxDaily), catCol=catColors[item.category]||'#888';
                      const g=document.createElementNS(svgNS,'g');
                      const rect=document.createElementNS(svgNS,'rect'); rect.setAttribute('x',offsetX+1); rect.setAttribute('y',y+1); rect.setAttribute('width',itemPxW-2); rect.setAttribute('height',cellH-2); rect.setAttribute('fill',heatColor(intensity)); rect.setAttribute('stroke',catCol); rect.setAttribute('stroke-width',1.5); rect.setAttribute('rx',4); rect.setAttribute('fill-opacity',0.85); g.appendChild(rect);
                      const bar=document.createElementNS(svgNS,'rect'); bar.setAttribute('x',offsetX+1); bar.setAttribute('y',y+1); bar.setAttribute('width',itemPxW-2); bar.setAttribute('height',3); bar.setAttribute('fill',catCol); bar.setAttribute('rx',4); g.appendChild(bar);
                      const mc=Math.max(2,Math.floor(itemPxW/9)); let nameText=item.name.length>mc?item.name.substring(0,mc-1)+'..':item.name; if(!item.has_width) nameText+='*';
                      const nt=document.createElementNS(svgNS,'text'); nt.setAttribute('x',offsetX+itemPxW/2); nt.setAttribute('y',y+cellH/2-8); nt.setAttribute('text-anchor','middle'); nt.setAttribute('dominant-baseline','central'); nt.setAttribute('font-size',Math.min(10,Math.max(6,itemPxW/nameText.length*0.85))); nt.setAttribute('fill','#fff'); nt.setAttribute('font-family','-apple-system,sans-serif'); nt.textContent=nameText; g.appendChild(nt);
                      if (itemPxW>35) {{ const rt=document.createElementNS(svgNS,'text'); rt.setAttribute('x',offsetX+itemPxW/2); rt.setAttribute('y',y+cellH/2+8); rt.setAttribute('text-anchor','middle'); rt.setAttribute('dominant-baseline','central'); rt.setAttribute('font-size',8); rt.setAttribute('fill','#4ECDC4'); rt.setAttribute('font-weight','bold'); rt.setAttribute('font-family','-apple-system,sans-serif'); rt.textContent=item.daily>=10000?Math.round(item.daily/10000)+'만/일':fmt(item.daily)+'/일'; g.appendChild(rt); }}
                      if (itemPxW>30) {{ const wt=document.createElementNS(svgNS,'text'); wt.setAttribute('x',offsetX+itemPxW/2); wt.setAttribute('y',y+cellH/2+22); wt.setAttribute('text-anchor','middle'); wt.setAttribute('dominant-baseline','central'); wt.setAttribute('font-size',7); wt.setAttribute('fill',item.has_width?'#999':'#e67e22'); wt.setAttribute('font-family','-apple-system,sans-serif'); wt.textContent=item.has_width?item.width_cm+'cm':'(수치 필요)'; g.appendChild(wt); }}
                      g.style.cursor='pointer';
                      g.addEventListener('mouseenter',(e)=>{{ tooltip.innerHTML='<div style="font-weight:bold;margin-bottom:4px;">'+item.name+'</div><div>카테고리: '+(item.category||'-')+'</div><div>가로: '+(item.has_width?item.width_cm+'cm':'<span style="color:#e67e22;">(수치 필요) 기본 '+item.width_cm+'cm 적용</span>')+'</div><div style="margin-top:4px;">총매출: <b style="color:#FF6B6B;">'+fmt(item.total)+'원</b></div><div>일평균: <b style="color:#4ECDC4;">'+fmt(item.daily)+'원</b></div>'; tooltip.style.display='block'; }});
                      g.addEventListener('mousemove',(e)=>{{ tooltip.style.left=(e.clientX+12)+'px'; tooltip.style.top=(e.clientY-10)+'px'; }});
                      g.addEventListener('mouseleave',()=>{{ tooltip.style.display='none'; }});
                      svg.appendChild(g); offsetX+=itemPxW;
                    }}
                    const utilText=document.createElementNS(svgNS,'text'); utilText.setAttribute('x',padX+labelW+shelfPxWidth+8); utilText.setAttribute('y',y+cellH/2); utilText.setAttribute('text-anchor','start'); utilText.setAttribute('dominant-baseline','central'); utilText.setAttribute('font-size',11); utilText.setAttribute('font-weight','bold'); utilText.setAttribute('font-family','-apple-system,sans-serif'); utilText.setAttribute('fill',util>=80?'#4ECDC4':util>=50?'#FFEAA7':'#FF6B6B'); utilText.textContent=items.length>0?util+'%':'-'; svg.appendChild(utilText);
                    const line=document.createElementNS(svgNS,'line'); line.setAttribute('x1',padX+labelW); line.setAttribute('y1',y+cellH); line.setAttribute('x2',padX+labelW+shelfPxWidth); line.setAttribute('y2',y+cellH); line.setAttribute('stroke','#333'); line.setAttribute('stroke-width',0.5); svg.appendChild(line);
                  }}
                  const legendY=padY+8+nTiers*cellH+4;
                  const legend=document.createElementNS(svgNS,'text'); legend.setAttribute('x',padX+labelW); legend.setAttribute('y',legendY); legend.setAttribute('font-size',9); legend.setAttribute('fill','#888'); legend.setAttribute('font-family','-apple-system,sans-serif'); legend.textContent='* 가로 수치 미등록 (기본 8cm 적용)'; svg.appendChild(legend);
                }})();
                </script>
                """
                components.html(_shelf_detail_html, height=_n_tiers * 70 + 60, scrolling=True)

            # 상세 테이블
            st.markdown("---")
            detail_rows = detail_by_fx.get(sel_fx, [])
            if detail_rows:
                detail_df = pd.DataFrame(detail_rows).sort_values(["단", "위치"])
                st.dataframe(
                    detail_df.style.format({"총매출": "{:,.0f}", "일평균": "{:,.0f}"}),
                    use_container_width=True, hide_index=True,
                )

    # ── 배치 이력 분석 (하위 섹션) ──
    st.divider()
    st.markdown("## 📅 배치 이력 분석")

    all_placements = get_all_placements()
    if all_placements.empty:
        st.info("배치 이력이 없습니다.")
    else:
        tab_loc, tab_prod, tab_timeline = st.tabs([
            "같은 자리 비교", "같은 상품 비교", "배치 타임라인"
        ])

        with tab_loc:
            st.subheader("같은 자리, 다른 상품 — 일평균 매출 비교")
            locations = get_all_locations()
            loc_options = locations[["id", "display_label"]].to_dict("records")
            sel_loc = st.selectbox("선반 위치 선택", loc_options,
                                   format_func=lambda x: x["display_label"], key="history_loc")

            if sel_loc:
                history = get_placement_history(sel_loc["id"])
                if history.empty:
                    st.info("이 위치에 배치 이력이 없습니다.")
                else:
                    st.dataframe(
                        history[["product_name", "erp_category", "start_date", "end_date", "notes"]],
                        use_container_width=True, hide_index=True)

                    chart_data = []
                    for _, h in history.iterrows():
                        s_date = str(h["start_date"])
                        e_date = str(h["end_date"]) if pd.notna(h["end_date"]) else date.today().isoformat()
                        sales = fetch_sales_for_placement_history(h["product_name"], s_date, e_date)
                        daily_rev = sales["total_revenue"] / max(1, sales["days"])
                        chart_data.append({
                            "상품": h["product_name"][:15], "기간": f"{s_date} ~ {e_date}",
                            "일평균매출": daily_rev,
                        })

                    if chart_data:
                        chart_df = pd.DataFrame(chart_data)
                        fig = px.bar(chart_df, x="상품", y="일평균매출", color="기간",
                                     title=f"{sel_loc['display_label']} — 배치별 일평균 매출", text_auto=",.0f")
                        st.plotly_chart(fig, use_container_width=True)

        with tab_prod:
            st.subheader("같은 상품, 다른 위치 — 일평균 매출 비교")
            if not all_placements.empty:
                product_names = sorted(all_placements["product_name"].unique().tolist())
                sel_prod = st.selectbox("상품 선택", product_names, key="history_prod")

                if sel_prod:
                    prod_history = get_product_placement_history(sel_prod)
                    if prod_history.empty:
                        st.info("이 상품의 배치 이력이 없습니다.")
                    else:
                        st.dataframe(prod_history[["display_label", "start_date", "end_date", "notes"]],
                                     use_container_width=True, hide_index=True)

                        chart_data = []
                        for _, h in prod_history.iterrows():
                            s_date = str(h["start_date"])
                            e_date = str(h["end_date"]) if pd.notna(h["end_date"]) else date.today().isoformat()
                            sales = fetch_sales_for_placement_history(sel_prod, s_date, e_date)
                            daily_rev = sales["total_revenue"] / max(1, sales["days"])
                            chart_data.append({"위치": h["display_label"], "기간": f"{s_date} ~ {e_date}",
                                               "일평균매출": daily_rev})

                        if chart_data:
                            chart_df = pd.DataFrame(chart_data)
                            fig = px.bar(chart_df, x="위치", y="일평균매출", color="기간",
                                         title=f"{sel_prod} — 위치별 일평균 매출 비교", text_auto=",.0f")
                            st.plotly_chart(fig, use_container_width=True)

        with tab_timeline:
            st.subheader("배치 타임라인")
            if not all_placements.empty:
                gantt_df = all_placements.copy()
                gantt_df["start"] = pd.to_datetime(gantt_df["start_date"])
                gantt_df["end"] = pd.to_datetime(gantt_df["end_date"].fillna(date.today().isoformat()))

                gantt_type = st.selectbox("매대 타입 필터", ["전체"] + list(SHELF_CONFIGS.keys()), key="gantt_type")
                if gantt_type != "전체":
                    gantt_df = gantt_df[gantt_df["shelf_type"] == gantt_type]

                if gantt_df.empty:
                    st.info("해당 조건의 배치 이력이 없습니다.")
                else:
                    gantt_df = gantt_df.head(50)
                    fig = px.timeline(gantt_df, x_start="start", x_end="end", y="display_label",
                                      color="erp_category", hover_name="product_name",
                                      title="배치 타임라인 (Gantt Chart)",
                                      labels={"display_label": "선반 위치", "erp_category": "카테고리"})
                    fig.update_layout(height=max(400, len(gantt_df) * 25 + 100),
                                      yaxis=dict(autorange="reversed"))
                    st.plotly_chart(fig, use_container_width=True)


# ======================================================================
# SKU 치수 관리
# ======================================================================
elif menu == "📐 SKU 치수 관리":
    st.title("📐 SKU 치수 관리")

    dims = get_all_dimensions()

    tab_overview, tab_edit, tab_recommend, tab_predict = st.tabs([
        "치수 현황", "치수 입력/수정", "배치 추천", "선반 수요 예측"
    ])

    # ── 치수 현황 ──
    with tab_overview:
        if dims.empty:
            st.warning("치수 데이터가 없습니다. '치수 입력/수정' 탭에서 데이터를 추가해 주세요.")
        else:
            c1, c2, c3, c4, c5 = st.columns(5)
            c1.metric("등록 SKU", f"{len(dims)}개")
            tall_n = len(dims[dims["size_class"] == "tall"])
            med_n = len(dims[dims["size_class"] == "medium"])
            short_n = len(dims[dims["size_class"] == "short"])
            c2.metric("키 큰 (>23cm)", f"{tall_n}개")
            c3.metric("중간 (15~23cm)", f"{med_n}개")
            c4.metric("작은 (<=15cm)", f"{short_n}개")
            dual_rate = dims["dual_row"].mean() * 100
            c5.metric("2열 가능 비율", f"{dual_rate:.1f}%")

            st.markdown("---")

            col_hist, col_pie = st.columns(2)
            with col_hist:
                fig_h = px.histogram(dims, x="height", nbins=30, title="높이(cm) 분포",
                                     labels={"height": "높이 (cm)", "count": "SKU 수"},
                                     color_discrete_sequence=["#4ECDC4"])
                fig_h.add_vline(x=15, line_dash="dash", line_color="orange",
                                annotation_text="15cm (작은/중간 경계)")
                fig_h.add_vline(x=23, line_dash="dash", line_color="red",
                                annotation_text="23cm (중간/키큰 경계)")
                st.plotly_chart(fig_h, use_container_width=True)

            with col_pie:
                size_counts = dims["size_class"].value_counts().reset_index()
                size_counts.columns = ["분류", "수량"]
                label_map = {"tall": "키 큰 (>23cm)", "medium": "중간 (15~23cm)", "short": "작은 (<=15cm)"}
                size_counts["분류"] = size_counts["분류"].map(label_map)
                fig_pie = px.pie(size_counts, values="수량", names="분류", title="사이즈 분류 비율",
                                 color_discrete_sequence=["#FF6B6B", "#4ECDC4", "#45B7D1"])
                st.plotly_chart(fig_pie, use_container_width=True)

            if dims["width"].notna().any():
                fig_scatter = px.scatter(
                    dims.dropna(subset=["width", "height"]),
                    x="width", y="height", color="size_class", hover_name="product_name",
                    title="가로 x 높이 분포",
                    labels={"width": "가로 (cm)", "height": "높이 (cm)", "size_class": "분류"},
                    color_discrete_map={"tall": "#FF6B6B", "medium": "#4ECDC4", "short": "#45B7D1"},
                )
                fig_scatter.add_hline(y=23, line_dash="dot", line_color="gray",
                                      annotation_text="25cm 선반 한계 (여유 2cm)")
                st.plotly_chart(fig_scatter, use_container_width=True)

            st.subheader("치수 데이터")
            filter_size = st.selectbox("사이즈 필터",
                                       ["전체", "키 큰 (>23cm)", "중간 (15~23cm)", "작은 (<=15cm)"],
                                       key="dims_filter")
            show_dims = dims.copy()
            if "키 큰" in filter_size:
                show_dims = show_dims[show_dims["size_class"] == "tall"]
            elif "중간" in filter_size:
                show_dims = show_dims[show_dims["size_class"] == "medium"]
            elif "작은" in filter_size:
                show_dims = show_dims[show_dims["size_class"] == "short"]

            st.dataframe(
                show_dims[["product_name", "width", "height", "depth", "size_class", "dual_row"]].rename(
                    columns={"product_name": "상품명", "width": "가로", "height": "높이",
                             "depth": "깊이", "size_class": "분류", "dual_row": "2열가능"}
                ),
                use_container_width=True, hide_index=True,
            )
            st.caption(f"표시: {len(show_dims)}개 / 전체: {len(dims)}개")

    # ── 치수 입력/수정 ──
    with tab_edit:
        st.subheader("개별 입력/수정")

        col_e1, col_e2 = st.columns(2)
        with col_e1:
            products_df = load_product_list()
            if not products_df.empty:
                all_names = sorted(products_df["name"].dropna().unique().tolist())
                edit_product = st.selectbox("상품 선택", all_names, key="dim_edit_product")
            else:
                edit_product = st.text_input("상품명", key="dim_edit_product_text")

            existing = get_dimension(edit_product) if edit_product else None
            default_w = existing["width"] if existing and existing.get("width") else 0.0
            default_h = existing["height"] if existing and existing.get("height") else 0.0
            default_d = existing["depth"] if existing and existing.get("depth") else 0.0

        with col_e2:
            _dk = edit_product or "_none"
            edit_w = st.number_input("가로 (cm)", value=float(default_w), min_value=0.0, step=0.1, key=f"dim_w_{_dk}")
            edit_h = st.number_input("높이 (cm)", value=float(default_h), min_value=0.0, step=0.1, key=f"dim_h_{_dk}")
            edit_d = st.number_input("깊이 (cm)", value=float(default_d), min_value=0.0, step=0.1, key=f"dim_d_{_dk}")

        if existing:
            st.info(f"기존 데이터: {existing.get('width')}x{existing.get('height')}x{existing.get('depth')}cm — 분류: {existing.get('size_class')}")

        if st.button("저장", type="primary", key="btn_dim_save"):
            if edit_product and edit_h > 0:
                upsert_dimension(edit_product, edit_w if edit_w > 0 else None, edit_h, edit_d if edit_d > 0 else None)
                st.success(f"저장 완료! {edit_product} — {edit_w}x{edit_h}x{edit_d}cm")
                st.rerun()
            else:
                st.warning("상품명과 높이를 입력해 주세요.")

        # ── 치수 미기입 제품 목록 ──
        st.markdown("---")
        st.subheader("치수 미기입 제품")
        st.caption("현재 매대에 배치되어 있지만 치수가 등록되지 않은 상품입니다.")

        placements = get_current_placements()
        all_dims = get_all_dimensions()
        if not placements.empty:
            placed_names = set(placements["product_name"].dropna().unique())
            dim_names = set(all_dims["product_name"].unique()) if not all_dims.empty else set()
            missing = sorted(placed_names - dim_names)

            if missing:
                st.warning(f"치수 미기입 제품: **{len(missing)}개**")
                selected_missing = st.selectbox(
                    "제품 선택 후 아래에서 치수를 입력하세요",
                    missing, key="dim_missing_select",
                )
                # 선택된 제품의 배치 위치 표시
                if selected_missing:
                    locs = placements[placements["product_name"] == selected_missing]
                    loc_strs = [f"{r.get('shelf_type','')}-{r.get('fixture_no','')}/{r.get('tier','')}단"
                                for _, r in locs.iterrows()]
                    st.info(f"📍 배치 위치: {', '.join(loc_strs)}")

                    mc1, mc2 = st.columns(2)
                    with mc1:
                        mw = st.number_input("가로 (cm)", value=0.0, min_value=0.0, step=0.1, key=f"miss_w_{selected_missing}")
                        mh = st.number_input("높이 (cm)", value=0.0, min_value=0.0, step=0.1, key=f"miss_h_{selected_missing}")
                        md = st.number_input("깊이 (cm)", value=0.0, min_value=0.0, step=0.1, key=f"miss_d_{selected_missing}")
                    with mc2:
                        st.markdown("<br>", unsafe_allow_html=True)
                        if st.button("저장", type="primary", key="btn_miss_save"):
                            if mh > 0:
                                upsert_dimension(selected_missing, mw if mw > 0 else None, mh, md if md > 0 else None)
                                st.success(f"저장 완료! {selected_missing} — {mw}x{mh}x{md}cm")
                                st.rerun()
                            else:
                                st.warning("높이는 필수 입력입니다.")
            else:
                st.success("모든 배치 제품의 치수가 등록되어 있습니다! ✅")
        else:
            st.info("현재 배치된 제품이 없습니다.")

        st.markdown("---")
        st.subheader("일괄 업로드 (Excel/CSV)")
        st.caption("컬럼: product_name(상품명), width(가로), height(높이), depth(깊이)")

        dim_upload = st.file_uploader("파일 업로드", type=["csv", "xlsx"], key="dim_bulk_upload")
        if dim_upload:
            if dim_upload.name.endswith(".csv"):
                udf = pd.read_csv(dim_upload)
            else:
                udf = pd.read_excel(dim_upload, engine="openpyxl")

            col_remap = {}
            for c in udf.columns:
                cl = c.strip()
                if cl in ("상품명", "product_name", "name"):
                    col_remap[c] = "product_name"
                elif cl in ("가로", "width"):
                    col_remap[c] = "width"
                elif cl in ("높이", "세로", "height"):
                    col_remap[c] = "height"
                elif cl in ("깊이", "폭", "depth"):
                    col_remap[c] = "depth"
            udf = udf.rename(columns=col_remap)

            valid_udf = udf[udf["product_name"].notna() & udf["height"].notna()]
            st.dataframe(valid_udf.head(10), use_container_width=True, hide_index=True)
            st.caption(f"유효 행: {len(valid_udf)}개")

            if st.button("일괄 등록/갱신", type="primary", key="btn_dim_bulk"):
                records = valid_udf.to_dict("records")
                count = bulk_upsert_dimensions(records)
                st.success(f"일괄 등록 완료! {count}건")
                st.rerun()

    # ── 배치 추천 ──
    with tab_recommend:
        st.subheader("상품 치수 기반 배치 위치 추천")
        st.caption("상품 높이와 선반 높이를 비교하여 적합한 빈 선반을 추천합니다.")

        if dims.empty:
            st.warning("치수 데이터가 없습니다.")
        else:
            dim_products = sorted(dims["product_name"].tolist())
            rec_product = st.selectbox("상품 선택", dim_products, key="rec_product")

            if rec_product:
                dim_info = get_dimension(rec_product)
                if dim_info:
                    st.info(
                        f"**{rec_product}** — "
                        f"가로 {dim_info.get('width', '?')}cm x "
                        f"높이 {dim_info['height']}cm x "
                        f"깊이 {dim_info.get('depth', '?')}cm — "
                        f"분류: {dim_info['size_class']}"
                    )

                recs = recommend_locations(rec_product, top_n=10)
                if recs:
                    rec_df = pd.DataFrame(recs)
                    st.markdown("**추천 위치 (적합도 순)**")

                    chart_df = rec_df.head(10)
                    fig_rec = px.bar(chart_df, x="display_label", y="fit_score",
                                     color="fit_score", color_continuous_scale="RdYlGn",
                                     title="배치 적합도 점수",
                                     labels={"display_label": "선반 위치", "fit_score": "적합도"},
                                     text_auto=".0f")
                    fig_rec.update_layout(coloraxis_showscale=False, xaxis_tickangle=-45)
                    st.plotly_chart(fig_rec, use_container_width=True)

                    st.dataframe(
                        rec_df[["display_label", "shelf_type", "tier_height", "height_waste", "fit_score"]].rename(
                            columns={"display_label": "위치", "shelf_type": "타입",
                                     "tier_height": "선반높이(cm)", "height_waste": "여유공간(cm)",
                                     "fit_score": "적합도"}
                        ),
                        use_container_width=True, hide_index=True,
                    )
                else:
                    st.warning("이 상품이 들어갈 수 있는 빈 선반이 없습니다.")

    # ── 선반 수요 예측 ──
    with tab_predict:
        st.subheader("선반 수요 예측")
        st.caption("현재 치수 데이터 기반으로 필요한 매대 수를 예측합니다.")

        if dims.empty:
            st.warning("치수 데이터가 없습니다.")
        else:
            pred = predict_shelf_demand()

            if "error" in pred:
                st.error(pred["error"])
            else:
                c1, c2, c3 = st.columns(3)
                c1.metric("총 등록 SKU", f"{pred['total_products']}개")
                c2.metric("필요 선반 수", f"{pred['shelves_needed']['total']}개")
                surplus = pred["surplus"]
                c3.metric("현재 보유 대비",
                          f"{'+' if surplus >= 0 else ''}{surplus}개",
                          delta=f"{'여유' if surplus >= 0 else '부족'}",
                          delta_color="normal" if surplus >= 0 else "inverse")

                st.markdown("---")

                col_size, col_shelf = st.columns(2)

                with col_size:
                    st.markdown("##### 사이즈별 상품 수")
                    size_data = pd.DataFrame([
                        {"분류": "키 큰 (>23cm)", "상품수": pred["size_counts"]["tall"],
                         "적합 선반": "A타입 5단 (높이 무제한)"},
                        {"분류": "중간 (15~23cm)", "상품수": pred["size_counts"]["medium"],
                         "적합 선반": "모든 타입 1~4단 (25cm)"},
                        {"분류": "작은 (<=15cm)", "상품수": pred["size_counts"]["short"],
                         "적합 선반": "모든 타입 1~4단 (25cm)"},
                    ])
                    st.dataframe(size_data, use_container_width=True, hide_index=True)

                    st.caption(
                        f"평균 가로: {pred['avg_width']}cm | "
                        f"2열 가능: {pred['dual_row_rate']}% | "
                        f"선반당 평균 수용: {pred['per_shelf_avg']}개"
                    )

                    if pred.get("tall_note"):
                        st.info(pred["tall_note"])

                with col_shelf:
                    st.markdown("##### 필요 선반 수 vs 보유")
                    a_unlimited = SHELF_CONFIGS["A"]["count"]  # A타입 5단
                    normal_shelves = pred["current_total"] - a_unlimited
                    shelf_data = pd.DataFrame([
                        {"구분": "높이 무제한 (A 5단)", "필요": pred["shelves_needed"]["tall_unlimited"],
                         "보유": a_unlimited},
                        {"구분": "일반 25cm 선반", "필요": pred["shelves_needed"]["normal_25cm"],
                         "보유": normal_shelves},
                    ])
                    shelf_data["과부족"] = shelf_data["보유"] - shelf_data["필요"]
                    st.dataframe(shelf_data, use_container_width=True, hide_index=True)

                st.markdown("---")

                st.markdown("##### 필요 매대 수 추정")
                fix_data = pd.DataFrame([
                    {"타입": f"A ({SHELF_CONFIGS['A']['name']})", "필요": pred["fixtures_needed"]["A"],
                     "현재": SHELF_CONFIGS["A"]["count"],
                     "과부족": SHELF_CONFIGS["A"]["count"] - pred["fixtures_needed"]["A"]},
                    {"타입": f"B ({SHELF_CONFIGS['B']['name']})", "필요": pred["fixtures_needed"]["B"],
                     "현재": SHELF_CONFIGS["B"]["count"],
                     "과부족": SHELF_CONFIGS["B"]["count"] - pred["fixtures_needed"]["B"]},
                    {"타입": f"C ({SHELF_CONFIGS['C']['name']})", "필요": pred["fixtures_needed"]["C"],
                     "현재": SHELF_CONFIGS["C"]["count"],
                     "과부족": SHELF_CONFIGS["C"]["count"] - pred["fixtures_needed"]["C"]},
                ])
                st.dataframe(fix_data, use_container_width=True, hide_index=True)

                # 시각화
                fig_compare = go.Figure()
                types = ["A", "B", "C"]
                needed = [pred["fixtures_needed"][t] for t in types]
                current = [SHELF_CONFIGS[t]["count"] for t in types]
                fig_compare.add_trace(go.Bar(name="필요", x=types, y=needed, marker_color="#FF6B6B"))
                fig_compare.add_trace(go.Bar(name="현재 보유", x=types, y=current, marker_color="#4ECDC4"))
                fig_compare.update_layout(barmode="group", title="타입별 매대 수: 필요 vs 보유",
                                          xaxis_title="매대 타입", yaxis_title="매대 수 (대)")
                st.plotly_chart(fig_compare, use_container_width=True)

                # 신규 SKU 시뮬레이션
                st.markdown("---")
                st.markdown("##### 신규 SKU 추가 시뮬레이션")

                sim_col1, sim_col2, sim_col3 = st.columns(3)
                with sim_col1:
                    sim_tall = st.number_input("키 큰 상품 추가", min_value=0, value=0, key="sim_tall")
                with sim_col2:
                    sim_med = st.number_input("중간 상품 추가", min_value=0, value=0, key="sim_med")
                with sim_col3:
                    sim_short = st.number_input("작은 상품 추가", min_value=0, value=0, key="sim_short")

                if sim_tall + sim_med + sim_short > 0:
                    sim_products = []
                    for i in range(sim_tall):
                        sim_products.append({"product_name": f"sim_tall_{i}", "height": 25.0, "width": 8.0, "depth": 10.0, "dual_row": 1})
                    for i in range(sim_med):
                        sim_products.append({"product_name": f"sim_med_{i}", "height": 18.0, "width": 8.0, "depth": 10.0, "dual_row": 1})
                    for i in range(sim_short):
                        sim_products.append({"product_name": f"sim_short_{i}", "height": 12.0, "width": 8.0, "depth": 10.0, "dual_row": 1})

                    sim_df = pd.DataFrame(sim_products)
                    pred_sim = predict_shelf_demand(extra_products=sim_df)

                    delta_shelves = pred_sim["shelves_needed"]["total"] - pred["shelves_needed"]["total"]

                    sc1, sc2, sc3 = st.columns(3)
                    sc1.metric("시뮬 총 SKU", f"{pred_sim['total_products']}개",
                               delta=f"+{sim_tall + sim_med + sim_short}")
                    sc2.metric("시뮬 필요 선반", f"{pred_sim['shelves_needed']['total']}개",
                               delta=f"+{delta_shelves}")
                    sc3.metric("시뮬 과부족", f"{pred_sim['surplus']}개",
                               delta_color="normal" if pred_sim["surplus"] >= 0 else "inverse")


# ======================================================================
# 탭 6: 교차판매 분석
# ======================================================================
elif menu == "🛒 교차판매 분석":
    st.title("🛒 교차판매 분석")
    st.caption("동일 주문 내 함께 구매된 상품을 분석하여 매대 인접 배치 의사결정을 지원합니다.")

    from basket_analysis import (
        prepare_basket_data,
        compute_cooccurrence,
        get_cross_sell_candidates,
        get_category_heatmap_data,
        get_category_cross_sell,
        get_products_by_category_pair,
        generate_placement_suggestions,
    )

    # ── 기간 선택 ──
    cs_col1, cs_col2 = st.sidebar.columns(2)
    with cs_col1:
        cs_date_from = st.date_input(
            "시작일", value=date.today() - timedelta(days=90), key="cs_date_from"
        )
    with cs_col2:
        cs_date_to = st.date_input(
            "종료일", value=date.today(), key="cs_date_to"
        )

    # ── 캐시된 바스켓 분석 ──
    @st.cache_data(ttl=3600)
    def _cached_basket_analysis(d_from: str, d_to: str):
        items = prepare_basket_data(d_from, d_to)
        if items.empty:
            return items, pd.DataFrame(), pd.DataFrame()
        product_cooc = compute_cooccurrence(items, level="product")
        category_cooc = compute_cooccurrence(items, level="category")
        return items, product_cooc, category_cooc

    with st.spinner("주문 데이터 분석 중..."):
        items_df, product_cooc, category_cooc = _cached_basket_analysis(
            str(cs_date_from), str(cs_date_to)
        )

    if items_df.empty:
        st.warning("해당 기간에 주문 데이터가 없습니다.")
    else:
        tab_category, tab_product, tab_placement = st.tabs([
            "📊 카테고리 교차분석", "🔍 상품별 교차판매", "🗺️ 배치 제안"
        ])

        # ━━━━ 탭 1: 상품별 교차판매 ━━━━
        with tab_product:
            # 상품 목록
            all_products = sorted(items_df["product_name"].unique().tolist())
            selected_product = st.selectbox(
                "상품 검색", all_products, key="cs_product_select"
            )

            if selected_product:
                # KPI
                target_orders = items_df[items_df["product_name"] == selected_product]["order_id"].nunique()
                total_orders = items_df["order_id"].nunique()

                # 해당 상품이 포함된 주문의 평균 바스켓 크기
                target_order_ids = items_df[items_df["product_name"] == selected_product]["order_id"].unique()
                basket_sizes = items_df[items_df["order_id"].isin(target_order_ids)].groupby("order_id")["product_name"].nunique()
                avg_basket = basket_sizes.mean() if len(basket_sizes) > 0 else 0

                # 교차판매 후보
                cross_df = get_cross_sell_candidates(
                    items_df, selected_product,
                    cooccurrence_df=product_cooc,
                    top_n=15, min_count=3,
                )

                k1, k2, k3 = st.columns(3)
                k1.metric("포함 주문 수", f"{target_orders:,}건",
                          help=f"전체 {total_orders:,}건 중")
                k2.metric("평균 바스켓 크기", f"{avg_basket:.1f}개")
                k3.metric("교차판매 상품 수", f"{len(cross_df)}개",
                          help="min_count ≥ 3 기준")

                st.markdown("---")

                if cross_df.empty:
                    st.info("동시구매 횟수 3회 이상인 상품이 없습니다. 기간을 늘려보세요.")
                else:
                    # 바차트: top 10 by lift
                    chart_data = cross_df.head(10).copy()
                    chart_data["label"] = chart_data["product"].apply(
                        lambda x: x[:20] + "..." if len(x) > 20 else x
                    )

                    fig_bar = px.bar(
                        chart_data, x="lift", y="label",
                        orientation="h",
                        color="lift",
                        color_continuous_scale="RdYlGn",
                        title=f"'{selected_product}' 교차판매 Top 10 (Lift 기준)",
                        labels={"lift": "Lift", "label": "상품"},
                        text="count",
                    )
                    fig_bar.update_layout(
                        yaxis=dict(autorange="reversed"),
                        coloraxis_showscale=False,
                        height=400,
                    )
                    fig_bar.update_traces(texttemplate="%{text}회", textposition="outside")
                    st.plotly_chart(fig_bar, use_container_width=True)

                    # 테이블
                    st.subheader("교차판매 상세")
                    display_df = cross_df.copy()
                    display_df["confidence"] = (display_df["confidence"] * 100).round(1)
                    display_df = display_df.rename(columns={
                        "product": "상품명",
                        "category": "카테고리",
                        "count": "동시구매 횟수",
                        "confidence": "Confidence(%)",
                        "lift": "Lift",
                    })
                    st.dataframe(display_df, use_container_width=True, hide_index=True)

                    st.caption(
                        "**Lift 해석**: >1 양의 연관 (함께 구매 경향), =1 무관, <1 음의 연관 | "
                        "**Confidence**: 타깃 상품 구매 시 해당 상품도 구매할 확률"
                    )

        # ━━━━ 탭 2: 카테고리 교차분석 ━━━━
        with tab_category:
            # 지표 선택
            metric_choice = st.radio(
                "지표 선택", ["Confidence", "Lift"],
                horizontal=True, key="cs_metric",
            )
            metric_key = "lift" if metric_choice == "Lift" else "confidence"
            heatmap_matrix = get_category_heatmap_data(items_df, metric=metric_key)

            if heatmap_matrix.empty:
                st.info("카테고리 교차분석 데이터가 없습니다.")
            else:
                if metric_key == "confidence":
                    st.subheader("카테고리 × 카테고리 Confidence Heatmap")
                    st.info(
                        "**Confidence란?** "
                        "\"A를 산 고객 중 B도 같이 산 비율\"입니다.\n\n"
                        "예) 감기약→비타민 = 25%이면, 감기약 산 고객 4명 중 1명은 비타민도 함께 구매했다는 뜻입니다.\n\n"
                        "**읽는 법**: 행(↓)이 기준 카테고리, 열(→)이 함께 구매한 카테고리입니다. "
                        "값이 높을수록 함께 구매하는 비율이 높으므로, 가까이 배치하면 효과적입니다."
                    )
                else:
                    st.subheader("카테고리 × 카테고리 Lift Heatmap")
                    st.info(
                        "**Lift란?** "
                        "\"두 카테고리가 우연히 같이 팔릴 확률 대비, 실제로 얼마나 더 같이 팔리는가\"입니다.\n\n"
                        "예) Lift = 2.0이면, 우연히 기대되는 것보다 2배 더 자주 함께 구매된다는 뜻입니다.\n\n"
                        "**읽는 법**: 1보다 크면(초록) 서로 끌어주는 관계, 1보다 작으면(빨강) 오히려 따로 사는 경향입니다. "
                        "Lift는 방향 구분 없이 대칭입니다."
                    )
                st.caption("👆 히트맵 칸을 **클릭**하면 교차구매 상품을 확인할 수 있습니다.")

                cats_list = heatmap_matrix.columns.tolist()
                n_cats = len(cats_list)

                # 지표별 색상/기준값 설정
                if metric_key == "lift":
                    colorscale = "RdYlGn"
                    zmid = 1.0
                    fmt = ".2f"
                    caption = "Lift > 1 (초록): 함께 구매 경향 | Lift < 1 (빨강): 따로 구매 경향 | Lift = 1: 무관"
                else:
                    colorscale = "YlOrRd"
                    zmid = None
                    fmt = ".0%"
                    caption = "값이 높을수록 (진한 빨강) 행 카테고리 구매 시 열 카테고리를 함께 구매하는 비율이 높음"

                # 텍스트 포맷
                if metric_key == "confidence":
                    text_vals = (heatmap_matrix.values * 100).round(1)
                    text_template = "%{text}%"
                else:
                    text_vals = np.round(heatmap_matrix.values, 2)
                    text_template = "%{text}"

                # Heatmap(시각) + 투명 Scatter(클릭용) 결합
                fig_heat = go.Figure()

                fig_heat.add_trace(go.Heatmap(
                    z=heatmap_matrix.values,
                    x=cats_list,
                    y=heatmap_matrix.index.tolist(),
                    colorscale=colorscale,
                    zmid=zmid,
                    text=text_vals,
                    texttemplate=text_template,
                    hoverinfo="skip",
                    showscale=True,
                ))

                # 투명 Scatter (클릭 캡처용)
                sc_x, sc_y, sc_custom = [], [], []
                for i, row_cat in enumerate(heatmap_matrix.index.tolist()):
                    for j, col_cat in enumerate(cats_list):
                        sc_x.append(col_cat)
                        sc_y.append(row_cat)
                        sc_custom.append([row_cat, col_cat])
                fig_heat.add_trace(go.Scatter(
                    x=sc_x, y=sc_y,
                    mode="markers",
                    marker=dict(size=max(18, min(50, 500 // n_cats)),
                                symbol="square", opacity=0),
                    hovertemplate="행: %{y}<br>열: %{x}<extra></extra>",
                    customdata=sc_custom,
                    showlegend=False,
                ))

                fig_heat.update_layout(
                    height=max(500, n_cats * 35),
                    xaxis_tickangle=-45,
                    margin=dict(l=150, b=150),
                )

                event = st.plotly_chart(
                    fig_heat, use_container_width=True,
                    on_select="rerun",
                    selection_mode=("points",),
                    key="cat_heatmap",
                )
                st.caption(caption)

                # 클릭 이벤트에서 카테고리 쌍 추출
                sel_cat_a = None
                sel_cat_b = None
                if event and event.selection and event.selection.points:
                    for pt in event.selection.points:
                        if pt.get("curve_number") == 1:
                            sel_cat_a = pt.get("y")
                            sel_cat_b = pt.get("x")
                            break

                if sel_cat_a and sel_cat_b:
                    cell_val = heatmap_matrix.loc[sel_cat_a, sel_cat_b] if sel_cat_a in heatmap_matrix.index and sel_cat_b in heatmap_matrix.columns else None
                    if metric_key == "lift":
                        lift_color = "🟢" if cell_val and cell_val > 1.1 else ("🔴" if cell_val and cell_val < 0.9 else "⚪")
                    else:
                        lift_color = "🟢" if cell_val and cell_val > 0.3 else ("🔴" if cell_val and cell_val < 0.1 else "⚪")

                    st.markdown("---")
                    st.subheader(f"{sel_cat_a} → {sel_cat_b}" if metric_key == "confidence" else f"{sel_cat_a} × {sel_cat_b}")
                    if cell_val is not None:
                        if metric_key == "confidence":
                            st.markdown(f"Confidence = **{cell_val*100:.1f}%** {lift_color}")
                        else:
                            st.markdown(f"Lift = **{cell_val:.2f}** {lift_color}")

                    pair_products = get_products_by_category_pair(items_df, sel_cat_a, sel_cat_b, top_n=30, min_count=3)

                    if pair_products.empty:
                        # min_count=1로 재조회하여 데이터 자체가 없는지, 3회 미만인지 구분
                        pair_any = get_products_by_category_pair(items_df, sel_cat_a, sel_cat_b, top_n=1, min_count=1)
                        if not pair_any.empty:
                            st.info("동시 구매가 3회 미만입니다.")
                        else:
                            st.info("해당 카테고리 쌍의 교차구매 데이터가 없습니다.")
                    else:
                        dp = pair_products.copy()
                        if metric_key == "confidence":
                            dp["confidence_a"] = (dp["confidence_a"] * 100).round(1)
                            dp["confidence_b"] = (dp["confidence_b"] * 100).round(1)
                            display_pair = dp.rename(columns={
                                "product_a": "상품 A",
                                "cat_a": "카테고리 A",
                                "product_b": "상품 B",
                                "cat_b": "카테고리 B",
                                "count": "동시구매 횟수",
                                "confidence_a": "Conf A→B(%)",
                                "confidence_b": "Conf B→A(%)",
                                "lift": "Lift",
                            })
                        else:
                            display_pair = dp.drop(columns=["confidence_a", "confidence_b"], errors="ignore").rename(columns={
                                "product_a": "상품 A",
                                "cat_a": "카테고리 A",
                                "product_b": "상품 B",
                                "cat_b": "카테고리 B",
                                "count": "동시구매 횟수",
                                "lift": "Lift",
                            })
                        st.dataframe(display_pair, use_container_width=True, hide_index=True)

        # ━━━━ 탭 3: 배치 제안 ━━━━
        with tab_placement:
            st.subheader("교차판매 기반 배치 제안")
            st.caption("교차판매 상위 상품들의 현재 매대 위치를 확인하고, 인접 배치 여부를 제안합니다.")

            # 상품 선택
            all_products_p = sorted(items_df["product_name"].unique().tolist())
            sel_product_p = st.selectbox(
                "상품 검색", all_products_p, key="cs_placement_product"
            )

            if sel_product_p:
                # 교차판매 후보
                cross_p = get_cross_sell_candidates(
                    items_df, sel_product_p,
                    cooccurrence_df=product_cooc,
                    top_n=10, min_count=3,
                )

                if cross_p.empty:
                    st.info("교차판매 데이터가 부족합니다. 기간을 늘리거나 다른 상품을 선택하세요.")
                else:
                    placements = get_current_placements()
                    fixture_pos = get_fixture_positions()

                    # 선택 상품 현재 위치 표시
                    target_match = placements[placements["product_name"] == sel_product_p]
                    if not target_match.empty:
                        tr = target_match.iloc[0]
                        pos_start = tr.get("position_start", "")
                        pos_end = tr.get("position_end", "")
                        pos_str = f"{pos_start}번" if pos_start == pos_end else f"{pos_start}~{pos_end}번"
                        st.info(
                            f"📍 **{sel_product_p}** 현재 위치: "
                            f"**{tr['display_label']}** ({pos_str} 위치)"
                        )
                    else:
                        st.warning(f"📍 **{sel_product_p}**은(는) 현재 매대에 배치되어 있지 않습니다.")

                    suggestions = generate_placement_suggestions(
                        cross_p, placements, fixture_pos, sel_product_p
                    )

                    if suggestions.empty:
                        st.warning("배치 정보를 가져올 수 없습니다.")
                    else:
                        # 인접 배치 권장 수 (is_adjacent가 False인 항목)
                        placed = suggestions[suggestions["is_adjacent"].notna()]
                        n_not_adjacent = (placed["is_adjacent"] == False).sum() if not placed.empty else 0
                        s1, s2 = st.columns(2)
                        s1.metric("교차판매 상위 상품", f"{len(suggestions)}개")
                        s2.metric("인접 배치 권장", f"{n_not_adjacent}개",
                                  help="인접 매대에 위치하지 않은 상품")

                        st.markdown("---")

                        # 테이블
                        disp_sug = suggestions.copy()
                        disp_sug["인접배치"] = disp_sug["is_adjacent"].apply(
                            lambda x: "✅ 인접" if x is True else ("⚠️ 권장" if x is False else "-")
                        )

                        disp_sug = disp_sug.rename(columns={
                            "product": "교차판매 상품",
                            "category": "카테고리",
                            "lift": "Lift",
                            "count": "동시구매",
                            "target_location": "타깃 위치",
                            "product_location": "상품 위치",
                        })

                        st.dataframe(
                            disp_sug[["교차판매 상품", "카테고리", "Lift", "동시구매",
                                      "타깃 위치", "상품 위치", "인접배치"]],
                            use_container_width=True, hide_index=True,
                        )

                        st.caption(
                            "✅ 인접: 같은 매대 또는 바로 옆 매대에 배치 | "
                            "⚠️ 권장: 인접 매대가 아니므로 가까이 배치 시 교차판매 효과 기대 | "
                            "실제 배치 변경은 '✏️ 배치 관리' 메뉴에서 수행하세요."
                        )

# ======================================================================
# 쇼카드 제작
# ======================================================================
elif menu == "🏷️ 쇼카드 제작":
    st.header("🏷️ 쇼카드 제작")
    st.caption("매대 쇼카드를 직접 제작하고 인쇄용 PDF로 다운로드합니다.")

    # ── 상수 ──
    SHOWCARD_COLORS = {
        "진통/해열": "#d6211a", "소화/위장": "#5e9e33", "잇몸/치과": "#e61a40",
        "치질": "#94c96e", "비염/코": "#14a1ad", "눈건강": "#2e2b85",
        "피부/연고": "#e61778", "여성건강": "#a80d82", "간/영양": "#146133",
        "소화효소": "#1c2e6e", "탈모": "#1c9ecc", "관절": "#0a82c2",
    }
    MAX_SPEC_SIZES = 7  # 규격 타입 최대 개수

    def _get_size_keys(specs):
        """specs에서 사이즈 키만 추출 (순서 유지)"""
        order = specs.get("_size_order")
        if order and isinstance(order, list):
            return [k for k in order if k in specs]
        return [k for k in specs if not k.startswith("_")]

    # ── 디자인 규격 기본값 (v2: 사이즈별) ──
    FONT_FAMILIES = [
        "sans-serif",
        "'Noto Sans KR', sans-serif",
        "'Pretendard', sans-serif",
        "serif",
        "monospace",
    ]

    DEFAULT_COMMON = {
        "text_align": "left",
    }

    # ── PDF 규격 기반 기본 템플릿 (수정본 0312) ──
    _SHORT_BASE = {
        "card_width_mm": 70, "card_height_mm": 85,
        "padding_lr_mm": 2.5, "padding_bottom_mm": 31.0,
        "header_height_mm": 12.5, "header_corner_mm": 10.0, "header_body_gap_mm": 9.7,
        "header_font_pt": 15.0, "header_font_weight": "Regular",
        "line1_font_pt": 15.0, "line1_weight": "Regular", "line1_font_family": "sans-serif",
        "line2_font_pt": 21.0, "line2_weight": "Regular", "line2_font_family": "sans-serif",
        "line3_font_pt": 30.0, "line3_weight": "ExtraBold", "line3_font_family": "sans-serif",
    }
    _LONG_BASE = {
        "card_width_mm": 110, "card_height_mm": 78.5,
        "padding_lr_mm": 3.0, "padding_bottom_mm": 31.0,
        "header_height_mm": 13.5, "header_corner_mm": 9.0, "header_body_gap_mm": 9.0,
        "header_font_pt": 15.0, "header_font_weight": "Regular",
        "line1_font_pt": 15.0, "line1_weight": "Regular", "line1_font_family": "sans-serif",
        "line2_font_pt": 20.0, "line2_weight": "Regular", "line2_font_family": "sans-serif",
        "line3_font_pt": 36.0, "line3_weight": "ExtraBold", "line3_font_family": "sans-serif",
    }

    DEFAULT_SIZE_SPEC = dict(_LONG_BASE)  # 새 규격 추가 시 Long 기반 템플릿

    DEFAULT_SPECS = {
        "_common": dict(DEFAULT_COMMON),
        "_uniform": False,
        "_size_order": ["S1", "S2", "L1", "L2", "L3", "L4", "L5"],
        "S1": {**_SHORT_BASE, "card_width_mm": 40},
        "S2": {**_SHORT_BASE, "card_width_mm": 70},
        "L1": {**_LONG_BASE, "card_width_mm": 110},
        "L2": {**_LONG_BASE, "card_width_mm": 130},
        "L3": {**_LONG_BASE, "card_width_mm": 150},
        "L4": {**_LONG_BASE, "card_width_mm": 210},
        "L5": {**_LONG_BASE, "card_width_mm": 240},
    }

    # badge_* → header_* 필드 마이그레이션 헬퍼
    _BADGE_TO_HEADER = {
        "badge_height_mm": "header_height_mm",
        "badge_radius_mm": "header_corner_mm",
        "badge_gap_mm": "header_body_gap_mm",
    }

    def _migrate_badge_fields(spec_dict):
        """구 badge_* 필드를 header_* 로 변환 (in-place)"""
        for old_k, new_k in _BADGE_TO_HEADER.items():
            if old_k in spec_dict and new_k not in spec_dict:
                spec_dict[new_k] = spec_dict.pop(old_k)
            elif old_k in spec_dict:
                del spec_dict[old_k]
        # badge_font_pt 제거 (더 이상 스펙에서 관리하지 않음)
        spec_dict.pop("badge_font_pt", None)
        return spec_dict

    # v2 → v3 마이그레이션
    _V2_REMOVE_SIZE_FIELDS = {"line1_y_pct", "line2_y_pct", "line3_y_pct", "top_ratio_pct", "padding_top_mm"}

    def _migrate_spec_v2_to_v3(spec_dict):
        """v2 → v3: Y% 제거, header_font 추가, corner_radius 제거"""
        if "_common" in spec_dict:
            spec_dict["_common"].pop("corner_radius_mm", None)
        size_keys = _get_size_keys(spec_dict)
        for sz in size_keys:
            if sz in spec_dict and isinstance(spec_dict[sz], dict):
                for f in _V2_REMOVE_SIZE_FIELDS:
                    spec_dict[sz].pop(f, None)
                spec_dict[sz].setdefault("header_font_pt", 15.0)
                spec_dict[sz].setdefault("header_font_weight", "Regular")
        return spec_dict

    # v1 → v2 마이그레이션
    def _migrate_spec_v1_to_v2(old_spec):
        """v1(단일 dict) → v2(사이즈별) 자동 변환"""
        try:
            common = {
                "corner_radius_mm": old_spec.get("corner_radius_mm", 2.5),
                "text_align": old_spec.get("text_align", "left"),
            }
            size_data = {}
            for k in DEFAULT_SIZE_SPEC:
                size_data[k] = old_spec.get(k, DEFAULT_SIZE_SPEC[k])
            for n in [1, 2, 3]:
                ff_key = f"line{n}_font_family"
                if ff_key not in size_data:
                    size_data[ff_key] = "sans-serif"
            if "padding_bottom_mm" not in size_data:
                size_data["padding_bottom_mm"] = 31.0
            _migrate_badge_fields(size_data)
            new_specs = {"_common": common, "_uniform": True, "_size_order": ["S", "M", "L", "XL", "XXL"]}
            sz_widths = {"S": 54, "M": 70, "L": 90, "XL": 110, "XXL": 150}
            for sz in ["S", "M", "L", "XL", "XXL"]:
                sz_data = dict(size_data)
                sz_data["card_width_mm"] = old_spec.get("card_width_mm", sz_widths[sz])
                new_specs[sz] = sz_data
            return new_specs
        except Exception:
            return None

    # session_state에 규격 로드 (v2: 사이즈별)
    if "sc_specs" not in st.session_state:
        _loaded_spec = None
        try:
            sb = _get_sb()
            if sb:
                res = sb.table("showcard_specs").select("*").order("created_at", desc=True).limit(1).execute()
                if res.data:
                    _loaded_spec = res.data[0].get("spec_json", {})
        except Exception:
            pass
        if _loaded_spec and isinstance(_loaded_spec, dict):
            ver = _loaded_spec.get("version", 1)
            if ver >= 2:
                # v2/v3 구조: version 키 제거 후 사용
                v_data = {k: v for k, v in _loaded_spec.items() if k != "version"}
                # 사이즈 키 목록 (동적)
                size_keys = _get_size_keys(v_data)
                if not size_keys:
                    size_keys = ["S", "M", "L", "XL", "XXL"]
                    v_data["_size_order"] = size_keys
                # 각 사이즈 누락 키 보충 + badge→header 마이그레이션
                for sz in size_keys:
                    if sz not in v_data:
                        v_data[sz] = dict(DEFAULT_SIZE_SPEC)
                    else:
                        _migrate_badge_fields(v_data[sz])
                        v_data[sz] = {**DEFAULT_SIZE_SPEC, **v_data[sz]}
                if "_common" not in v_data:
                    v_data["_common"] = dict(DEFAULT_COMMON)
                else:
                    v_data["_common"] = {**DEFAULT_COMMON, **v_data["_common"]}
                if "_uniform" not in v_data:
                    v_data["_uniform"] = True
                if "_size_order" not in v_data:
                    v_data["_size_order"] = size_keys
                # v2 → v3 마이그레이션 (v2인 경우)
                if ver < 3:
                    _migrate_spec_v2_to_v3(v_data)
                st.session_state["sc_specs"] = v_data
            else:
                # v1 → v2 → v3 마이그레이션
                migrated = _migrate_spec_v1_to_v2(_loaded_spec)
                if migrated:
                    _migrate_spec_v2_to_v3(migrated)
                st.session_state["sc_specs"] = migrated if migrated else {k: (dict(v) if isinstance(v, dict) else v) for k, v in DEFAULT_SPECS.items()}
        else:
            import copy
            st.session_state["sc_specs"] = copy.deepcopy(DEFAULT_SPECS)

    all_specs = st.session_state["sc_specs"]

    # ── 디자인 규격 설정 (expander) ──
    with st.expander("⚙️ 디자인 규격 설정 (디자이너용)", expanded=False):
        st.caption("쇼카드의 텍스트 크기, 위치, 여백 등을 규격화합니다. 저장하면 이후 모든 쇼카드에 적용됩니다.")

        common = all_specs["_common"]
        weight_options = ["Light", "Regular", "Medium", "Bold", "ExtraBold", "Black"]

        # 전체 동일 적용 체크박스
        is_uniform = st.checkbox("전체 사이즈에 동일 적용", value=all_specs.get("_uniform", True), key="sp_uniform")
        all_specs["_uniform"] = is_uniform

        # 필수 정의 항목
        _REQUIRED_FIELDS = {
            "card_width_mm": "카드 가로",
            "card_height_mm": "카드 세로",
            "line3_font_pt": "제목/제품명 폰트 크기",
            "padding_bottom_mm": "하단 필수 여백",
        }

        def _render_size_spec_ui(sz, sp):
            """사이즈별 규격 UI 렌더링 (v3: 카드기본/상단부/메인 3섹션)"""
            sfx = sz  # widget key 접미사

            # ── [카드 기본] ──
            st.markdown("**카드 기본**")
            card_c1, card_c2 = st.columns(2)
            with card_c1:
                sp["card_width_mm"] = st.number_input("전체 가로길이 (mm)", 10, 300, int(sp.get("card_width_mm", 90)), key=f"sp_w_{sfx}")
            with card_c2:
                sp["card_height_mm"] = st.number_input("전체 세로길이 (mm)", 10, 120, int(sp["card_height_mm"]), key=f"sp_h_{sfx}")

            st.markdown("---")

            # ── [상단부] ──
            st.markdown("**상단부**")
            hdr_c1, hdr_c2 = st.columns(2)
            with hdr_c1:
                sp["header_corner_mm"] = st.number_input("헤더 모퉁이 (mm)", 0.0, 20.0, float(sp.get("header_corner_mm", 9.0)), 0.5, key=f"sp_hc_{sfx}")
                sp["header_height_mm"] = st.number_input("헤더 높이 (mm)", 2.0, 30.0, float(sp.get("header_height_mm", 13.5)), 0.5, key=f"sp_hh_{sfx}")
            with hdr_c2:
                sp["header_font_pt"] = st.number_input("상단부 텍스트 크기 (pt)", 4.0, 30.0, float(sp.get("header_font_pt", 15.0)), 0.5, key=f"sp_hfp_{sfx}")
                sp["header_font_weight"] = st.selectbox("상단부 텍스트 굵기", weight_options, index=weight_options.index(sp.get("header_font_weight", "Regular")), key=f"sp_hfw_{sfx}")

            st.markdown("---")

            # ── [메인] ──
            st.markdown("**메인**")
            # 메인 세로 길이 (자동 계산, 읽기 전용)
            main_height = sp["card_height_mm"] - sp.get("header_height_mm", 13.5)
            st.caption(f"메인 세로 길이: **{main_height:.1f}** mm (자동 계산)")

            main_c1, main_c2 = st.columns(2)
            with main_c1:
                sp["header_body_gap_mm"] = st.number_input("헤더-본문 간격 (mm)", 0.0, 20.0, float(sp.get("header_body_gap_mm", 9.0)), 0.5, key=f"sp_hg_{sfx}")
                sp["padding_lr_mm"] = st.number_input("안쪽 바디 테두리 (mm)", 0.0, 15.0, float(sp["padding_lr_mm"]), 0.5, key=f"sp_plr_{sfx}")
            with main_c2:
                sp["padding_bottom_mm"] = st.number_input("하단 필수 여백 (mm)", 0.0, 50.0, float(sp.get("padding_bottom_mm", 31.0)), 0.5, key=f"sp_pb_{sfx}")

            st.markdown("**텍스트 규격**")
            txt_c1, txt_c2, txt_c3 = st.columns(3)

            with txt_c1:
                st.markdown("*1줄 (설명)*")
                sp["line1_font_pt"] = st.number_input("폰트 크기 (pt)", 4.0, 30.0, float(sp["line1_font_pt"]), 0.5, key=f"sp_f1_{sfx}")
                cur_ff1 = sp.get("line1_font_family", "sans-serif")
                ff1_idx = FONT_FAMILIES.index(cur_ff1) if cur_ff1 in FONT_FAMILIES else 0
                sp["line1_font_family"] = st.selectbox("폰트 종류", FONT_FAMILIES, index=ff1_idx, key=f"sp_ff1_{sfx}")
                sp["line1_weight"] = st.selectbox("굵기", weight_options, index=weight_options.index(sp.get("line1_weight", "Regular")), key=f"sp_w1_{sfx}")

            with txt_c2:
                st.markdown("*2줄 (소제목)*")
                sp["line2_font_pt"] = st.number_input("폰트 크기 (pt)", 4.0, 30.0, float(sp["line2_font_pt"]), 0.5, key=f"sp_f2_{sfx}")
                cur_ff2 = sp.get("line2_font_family", "sans-serif")
                ff2_idx = FONT_FAMILIES.index(cur_ff2) if cur_ff2 in FONT_FAMILIES else 0
                sp["line2_font_family"] = st.selectbox("폰트 종류", FONT_FAMILIES, index=ff2_idx, key=f"sp_ff2_{sfx}")
                sp["line2_weight"] = st.selectbox("굵기", weight_options, index=weight_options.index(sp.get("line2_weight", "Regular")), key=f"sp_w2_{sfx}")

            with txt_c3:
                st.markdown("*3줄 (제목/제품명)*")
                sp["line3_font_pt"] = st.number_input("폰트 크기 (pt)", 6.0, 40.0, float(sp["line3_font_pt"]), 0.5, key=f"sp_f3_{sfx}")
                cur_ff3 = sp.get("line3_font_family", "sans-serif")
                ff3_idx = FONT_FAMILIES.index(cur_ff3) if cur_ff3 in FONT_FAMILIES else 0
                sp["line3_font_family"] = st.selectbox("폰트 종류", FONT_FAMILIES, index=ff3_idx, key=f"sp_ff3_{sfx}")
                sp["line3_weight"] = st.selectbox("굵기", weight_options, index=weight_options.index(sp.get("line3_weight", "ExtraBold")), key=f"sp_w3_{sfx}")

            common["text_align"] = st.radio("텍스트 정렬", ["center", "left"], format_func=lambda x: {"center": "중앙 정렬", "left": "좌측 정렬"}.get(x), horizontal=True, key=f"sp_align_{sfx}")

            # ── 필수값 검증 ──
            for field, label in _REQUIRED_FIELDS.items():
                val = sp.get(field)
                if val is None or val <= 0:
                    st.markdown(f':red[**{label}** — 정의가 필요합니다.]')

        cur_sizes = _get_size_keys(all_specs)

        if is_uniform:
            # 전체 동일 적용: 첫 번째 사이즈 기준으로 편집
            first_sz = cur_sizes[0] if cur_sizes else "L"
            _render_size_spec_ui("ALL", all_specs[first_sz])
        else:
            # 사이즈별 개별 설정 (동적 탭)
            size_tabs = st.tabs(cur_sizes)
            for tab, sz in zip(size_tabs, cur_sizes):
                with tab:
                    _render_size_spec_ui(sz, all_specs[sz])

        # ── 규격 추가 / 삭제 / 이름 변경 ──
        st.markdown("---")
        st.markdown("**규격 타입 관리**")
        mgmt_c1, mgmt_c2, mgmt_c3 = st.columns(3)
        with mgmt_c1:
            st.markdown("*추가*")
            if len(cur_sizes) >= MAX_SPEC_SIZES:
                st.caption(f"최대 {MAX_SPEC_SIZES}개까지 가능합니다.")
            else:
                new_name = st.text_input("새 규격 이름", placeholder="예: 미니, 와이드", key="sp_new_name")
                if st.button("➕ 규격 추가", use_container_width=True, key="sp_add_size"):
                    name = new_name.strip()
                    if not name:
                        st.warning("이름을 입력해주세요.")
                    elif name in all_specs or name.startswith("_"):
                        st.warning(f"'{name}' 이름은 이미 존재하거나 사용할 수 없습니다.")
                    else:
                        all_specs[name] = dict(DEFAULT_SIZE_SPEC)
                        if "_size_order" not in all_specs:
                            all_specs["_size_order"] = list(cur_sizes)
                        all_specs["_size_order"].append(name)
                        st.session_state["sc_specs"] = all_specs
                        st.rerun()
        with mgmt_c2:
            st.markdown("*이름 변경*")
            rename_target = st.selectbox("변경할 규격", cur_sizes, key="sp_rename_target")
            rename_new = st.text_input("새 이름", placeholder="변경할 이름 입력", key="sp_rename_new")
            if st.button("✏️ 이름 변경", use_container_width=True, key="sp_rename"):
                new_n = rename_new.strip()
                if not new_n:
                    st.warning("새 이름을 입력해주세요.")
                elif new_n == rename_target:
                    st.info("동일한 이름입니다.")
                elif new_n in all_specs or new_n.startswith("_"):
                    st.warning(f"'{new_n}' 이름은 이미 존재하거나 사용할 수 없습니다.")
                else:
                    # 데이터 이동
                    all_specs[new_n] = all_specs.pop(rename_target)
                    # _size_order 업데이트
                    if "_size_order" in all_specs:
                        order = all_specs["_size_order"]
                        idx = order.index(rename_target) if rename_target in order else -1
                        if idx >= 0:
                            order[idx] = new_n
                    st.session_state["sc_specs"] = all_specs
                    st.rerun()
        with mgmt_c3:
            st.markdown("*삭제*")
            if len(cur_sizes) <= 1:
                st.caption("최소 1개 규격은 유지해야 합니다.")
            else:
                del_target = st.selectbox("삭제할 규격", cur_sizes, key="sp_del_target")
                if st.button("🗑️ 규격 삭제", use_container_width=True, key="sp_del_size"):
                    if del_target in all_specs:
                        del all_specs[del_target]
                        if "_size_order" in all_specs and del_target in all_specs["_size_order"]:
                            all_specs["_size_order"].remove(del_target)
                        st.session_state["sc_specs"] = all_specs
                        st.rerun()

        st.markdown("---")
        save_col1, save_col2 = st.columns(2)
        with save_col1:
            if st.button("💾 규격 저장", use_container_width=True, key="sp_save"):
                try:
                    cur_sz_keys = _get_size_keys(all_specs)
                    # uniform이면 첫 번째 사이즈 값을 전체에 복사
                    if all_specs["_uniform"]:
                        base = dict(all_specs[cur_sz_keys[0]])
                        for sz in cur_sz_keys:
                            all_specs[sz] = dict(base)
                    save_data = {"version": 3}
                    save_data["_common"] = dict(common)
                    save_data["_uniform"] = all_specs["_uniform"]
                    save_data["_size_order"] = cur_sz_keys
                    for sz in cur_sz_keys:
                        save_data[sz] = dict(all_specs[sz])
                    sb = _get_sb()
                    if sb:
                        sb.table("showcard_specs").insert({"spec_json": save_data}).execute()
                    st.session_state["sc_specs"] = all_specs
                    st.success("규격 저장 완료! 이후 모든 쇼카드에 적용됩니다.")
                except Exception as e:
                    st.session_state["sc_specs"] = all_specs
                    st.warning(f"로컬 저장 완료 (DB 저장 실패: {e})")
        with save_col2:
            if st.button("↩️ 기본값으로 초기화", use_container_width=True, key="sp_reset"):
                import copy
                st.session_state["sc_specs"] = copy.deepcopy(DEFAULT_SPECS)
                st.rerun()

    def _get_showcard_color(category: str) -> str:
        if not category:
            return "#5e9e33"
        for key, color in SHOWCARD_COLORS.items():
            if key in category or category in key:
                return color
        return "#5e9e33"

    def _recommend_size(width_cm: float) -> str:
        """진열폭(cm)에 가장 가까운 사이즈 추천 — 동적 사이즈 대응"""
        sz_keys = _get_size_keys(all_specs)
        if not sz_keys:
            return "L"
        # 각 사이즈의 card_width_mm와 진열폭 비교, 가장 가까운 것 선택
        width_mm = width_cm * 10
        best, best_diff = sz_keys[0], float("inf")
        for sz in sz_keys:
            w = all_specs[sz].get("card_width_mm", 90)
            diff = abs(w - width_mm)
            if diff < best_diff:
                best, best_diff = sz, diff
        return best

    def _escape_xml(s: str) -> str:
        return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")

    def _auto_fit(text: str, max_w: float, max_fs: float, min_fs: float = 8) -> float:
        char_w = 0.55
        for fs in range(int(max_fs), int(min_fs) - 1, -1):
            if len(text) * fs * char_w <= max_w:
                return fs
        return min_fs

    # mm → px 변환 (scale 기준)
    def _mm2px(mm_val, scale=3):
        return mm_val * scale

    def _weight_to_svg(w):
        return {"Light": "300", "Regular": "400", "Medium": "500", "Bold": "700", "ExtraBold": "800", "Black": "900"}.get(w, "400")

    def _calc_text_positions(h_px, size_spec, l1, l2, l3):
        """mm 기반 텍스트 Y 위치 자동 배치"""
        content_top = _mm2px(size_spec.get("header_height_mm", 13.5) + size_spec.get("header_body_gap_mm", 9.0))
        content_bottom = h_px - _mm2px(size_spec.get("padding_bottom_mm", 31.0))
        available = max(content_bottom - content_top, 1)
        has_l1 = bool(l1)
        has_l2 = bool(l2)
        # l3 always present
        if has_l1 and has_l2:
            # 3줄
            y1 = content_top + available * 0.25
            y2 = content_top + available * 0.50
            y3 = content_top + available * 0.75
        elif has_l2:
            # 2줄 (l2+l3)
            y1 = 0
            y2 = content_top + available * 0.35
            y3 = content_top + available * 0.70
        elif has_l1:
            # 2줄 (l1+l3)
            y1 = content_top + available * 0.35
            y2 = 0
            y3 = content_top + available * 0.70
        else:
            # 1줄 (l3만)
            y1, y2 = 0, 0
            y3 = content_top + available * 0.55
        return y1, y2, y3

    def _badge_svg(badge_type: str, w_px: float, size_spec: dict, common_spec: dict) -> str:
        """필(pill) 형태 배지 생성 — 배지는 규격에서 제외, 하드코딩 기본값 사용"""
        if badge_type == "none":
            return ""
        badges = {
            "동일성분": {"text": "동일성분", "text2": "↓ 저렴해요"},
            "유사성분": {"text": "유사성분", "text2": "↓ 저렴해요"},
            "업그레이드": {"text": "↑ 업그레이드", "text2": None},
        }
        b = badges.get(badge_type)
        if not b:
            return ""
        # 배지 고정값 (규격에서 제외됨)
        _BADGE_FONT_PT = 11.0
        _BADGE_HEIGHT_MM = 5.5
        _BADGE_RADIUS_MM = 3.0
        _BADGE_GAP_MM = 2.0
        fs = _mm2px(_BADGE_FONT_PT * 0.35)
        char_w = fs * 0.55
        bh = _mm2px(_BADGE_HEIGHT_MM)
        br = _mm2px(_BADGE_RADIUS_MM)
        gap = _mm2px(_BADGE_GAP_MM)
        pad = _mm2px(size_spec["padding_lr_mm"])
        pad_top = _mm2px(2.0)  # 고정 오프셋
        # 전체 배지 폭 계산 (위치 결정용)
        text1_w = len(b["text"]) * char_w
        bw = text1_w + fs * 1.6
        total_w = bw
        bw2 = 0
        if b["text2"]:
            text2_w = len(b["text2"]) * char_w
            bw2 = text2_w + fs * 1.6
            total_w += gap + bw2
        x, y = pad, pad_top  # 배지는 항상 좌측 상단 고정
        # 첫 번째 배지
        svg = (f'<rect x="{x}" y="{y}" width="{bw}" height="{bh}" rx="{br}" fill="rgba(0,0,0,0.35)"/>'
               f'<text x="{x + bw/2}" y="{y + bh*0.72}" text-anchor="middle" fill="white" '
               f'font-size="{fs}" font-weight="700" font-family="sans-serif">{_escape_xml(b["text"])}</text>')
        # 두 번째 배지
        if b["text2"]:
            x2 = x + bw + gap
            svg += (f'<rect x="{x2}" y="{y}" width="{bw2}" height="{bh}" rx="{br}" fill="rgba(255,255,255,0.25)"/>'
                    f'<text x="{x2 + bw2/2}" y="{y + bh*0.72}" text-anchor="middle" fill="white" '
                    f'font-size="{fs}" font-weight="700" font-family="sans-serif">{_escape_xml(b["text2"])}</text>')
        return svg

    def _gen_design_a(w_px, h_px, bg, badge, l1, l2, l3, size_spec, common_spec, header_text=""):
        """디자인 A: 단색 배경 — 규격 기반 (v3: 상단만 라운드, mm 기반 배치)"""
        r = _mm2px(size_spec.get("header_corner_mm", 10.0))
        pad = _mm2px(size_spec["padding_lr_mm"])
        fs1 = _auto_fit(l1 or "", w_px - pad*2, _mm2px(size_spec["line1_font_pt"] * 0.35), 8)
        fs2 = _auto_fit(l2 or "", w_px - pad*2, _mm2px(size_spec["line2_font_pt"] * 0.35), 7)
        fs3 = _auto_fit(l3, w_px - pad*2, _mm2px(size_spec["line3_font_pt"] * 0.35), 10)
        w1 = _weight_to_svg(size_spec["line1_weight"])
        w2 = _weight_to_svg(size_spec["line2_weight"])
        w3 = _weight_to_svg(size_spec["line3_weight"])
        ff1 = size_spec.get("line1_font_family", "sans-serif")
        ff2 = size_spec.get("line2_font_family", "sans-serif")
        ff3 = size_spec.get("line3_font_family", "sans-serif")
        align = common_spec["text_align"]
        anchor = "middle" if align == "center" else "start"
        tx = w_px / 2 if align == "center" else pad
        badge_el = _badge_svg(badge, w_px, size_spec, common_spec)
        # 상단만 라운드, 하단 직각
        bg_path = f"M{r},0 L{w_px-r},0 Q{w_px},0 {w_px},{r} L{w_px},{h_px} L0,{h_px} L0,{r} Q0,0 {r},0 Z"
        # 상단부 텍스트
        header_el = ""
        if header_text:
            top_h = _mm2px(size_spec.get("header_height_mm", 12.5))
            hdr_fs = _mm2px(size_spec.get("header_font_pt", 15.0) * 0.35)
            hdr_w = _weight_to_svg(size_spec.get("header_font_weight", "Regular"))
            header_y = top_h * 0.65
            header_el = (f'<text x="{tx}" y="{header_y}" text-anchor="{anchor}" fill="white" '
                         f'font-size="{hdr_fs}" font-weight="{hdr_w}" font-family="sans-serif" opacity="0.9">{_escape_xml(header_text)}</text>')
        # mm 기반 텍스트 Y 위치
        y1, y2, y3 = _calc_text_positions(h_px, size_spec, l1, l2, l3)
        lines = []
        if l1:
            lines.append(f'<text x="{tx}" y="{y1}" text-anchor="{anchor}" fill="white" '
                         f'font-size="{fs1}" font-weight="{w1}" font-family="{ff1}" opacity="0.92">{_escape_xml(l1)}</text>')
        if l2:
            lines.append(f'<text x="{tx}" y="{y2}" text-anchor="{anchor}" fill="white" '
                         f'font-size="{fs2}" font-weight="{w2}" font-family="{ff2}" opacity="0.85">{_escape_xml(l2)}</text>')
        lines.append(f'<text x="{tx}" y="{y3}" text-anchor="{anchor}" fill="white" '
                     f'font-size="{fs3}" font-weight="{w3}" font-family="{ff3}">{_escape_xml(l3)}</text>')
        return (f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {w_px} {h_px}" width="{w_px}" height="{h_px}">'
                f'<path d="{bg_path}" fill="{bg}"/>'
                f'{badge_el}{header_el}'
                f'{"".join(lines)}'
                f'</svg>')

    def _is_light(hex_color):
        h = hex_color.lstrip("#")
        if len(h) == 3:
            h = "".join(c*2 for c in h)
        try:
            r_, g_, b_ = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
            return (r_ * 299 + g_ * 587 + b_ * 114) / 1000 > 128
        except Exception:
            return True

    def _gen_design_b(w_px, h_px, bg, badge, l1, l2, l3, size_spec, common_spec, top_color=None, bot_color=None, header_text=""):
        """디자인 B: 상하단 컬러 구분 — 규격 기반 (v3: 상단만 라운드, mm 기반 배치)"""
        r = _mm2px(size_spec.get("header_corner_mm", 10.0))
        pad = _mm2px(size_spec["padding_lr_mm"])
        top_h = _mm2px(size_spec.get("header_height_mm", 13.5))
        tc = top_color or "#FFFFFF"
        bc = bot_color or bg
        bot_text_fill = "#333" if _is_light(bc) else "#FFFFFF"
        fs1 = _auto_fit(l1 or "", w_px - pad*2, _mm2px(size_spec["line1_font_pt"] * 0.35), 8)
        fs2 = _auto_fit(l2 or "", w_px - pad*2, _mm2px(size_spec["line2_font_pt"] * 0.35), 7)
        fs3 = _auto_fit(l3, w_px - pad*2, _mm2px(size_spec["line3_font_pt"] * 0.35), 10)
        w1 = _weight_to_svg(size_spec["line1_weight"])
        w2 = _weight_to_svg(size_spec["line2_weight"])
        w3 = _weight_to_svg(size_spec["line3_weight"])
        ff1 = size_spec.get("line1_font_family", "sans-serif")
        ff2 = size_spec.get("line2_font_family", "sans-serif")
        ff3 = size_spec.get("line3_font_family", "sans-serif")
        align = common_spec["text_align"]
        anchor = "middle" if align == "center" else "start"
        tx = w_px / 2 if align == "center" else pad
        # 배지 (상단 영역 안) — 하드코딩 기본값 사용
        badge_el = ""
        if badge != "none":
            _BADGE_FONT_PT = 11.0
            _BADGE_HEIGHT_MM = 5.5
            _BADGE_RADIUS_MM = 3.0
            b_fs = _mm2px(_BADGE_FONT_PT * 0.35)
            b_bh = _mm2px(_BADGE_HEIGHT_MM)
            b_br = _mm2px(_BADGE_RADIUS_MM)
            char_w = b_fs * 0.55
            btxt = "↑ 업그레이드" if badge == "업그레이드" else badge
            bw = len(btxt) * char_w + b_fs * 1.6
            bx, by = pad, _mm2px(2.0)  # 고정 오프셋
            badge_el = (f'<rect x="{bx}" y="{by}" width="{bw}" height="{b_bh}" rx="{b_br}" fill="{bc}" opacity="0.85"/>'
                        f'<text x="{bx + bw/2}" y="{by + b_bh*0.72}" text-anchor="middle" fill="{bot_text_fill}" '
                        f'font-size="{b_fs}" font-weight="700" font-family="sans-serif">{_escape_xml(btxt)}</text>')
        # 상단부 텍스트
        header_el = ""
        if header_text:
            hdr_fs = _mm2px(size_spec.get("header_font_pt", 15.0) * 0.35)
            hdr_w = _weight_to_svg(size_spec.get("header_font_weight", "Regular"))
            top_text_fill = "#333" if _is_light(tc) else "#FFFFFF"
            header_y = top_h * 0.65
            header_el = (f'<text x="{tx}" y="{header_y}" text-anchor="{anchor}" fill="{top_text_fill}" '
                         f'font-size="{hdr_fs}" font-weight="{hdr_w}" font-family="sans-serif">{_escape_xml(header_text)}</text>')
        # mm 기반 텍스트 Y 위치
        y1, y2, y3 = _calc_text_positions(h_px, size_spec, l1, l2, l3)
        lines = []
        if l1:
            lines.append(f'<text x="{tx}" y="{y1}" text-anchor="{anchor}" fill="{bot_text_fill}" '
                         f'font-size="{fs1}" font-weight="{w1}" font-family="{ff1}" opacity="0.92">{_escape_xml(l1)}</text>')
        if l2:
            lines.append(f'<text x="{tx}" y="{y2}" text-anchor="{anchor}" fill="{bot_text_fill}" '
                         f'font-size="{fs2}" font-weight="{w2}" font-family="{ff2}" opacity="0.85">{_escape_xml(l2)}</text>')
        lines.append(f'<text x="{tx}" y="{y3}" text-anchor="{anchor}" fill="{bot_text_fill}" '
                     f'font-size="{fs3}" font-weight="{w3}" font-family="{ff3}">{_escape_xml(l3)}</text>')
        # 상단만 라운드, 하단 직각 (전체 배경)
        bg_path = f"M{r},0 L{w_px-r},0 Q{w_px},0 {w_px},{r} L{w_px},{h_px} L0,{h_px} L0,{r} Q0,0 {r},0 Z"
        top_path = (f'M{r},0 L{w_px-r},0 Q{w_px},0 {w_px},{r} '
                    f'L{w_px},{top_h} L0,{top_h} L0,{r} Q0,0 {r},0 Z')
        return (f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {w_px} {h_px}" width="{w_px}" height="{h_px}">'
                f'<path d="{bg_path}" fill="{bc}"/>'
                f'<path d="{top_path}" fill="{tc}"/>'
                f'{badge_el}{header_el}'
                f'{"".join(lines)}'
                f'</svg>')

    def _svg_to_pdf_bytes(svg_str: str, w_mm: float, h_mm: float) -> bytes:
        """SVG → PDF 변환 (reportlab 사용)"""
        from io import BytesIO
        try:
            from svglib.svglib import renderSVG
            from reportlab.graphics import renderPDF
            drawing = renderSVG(BytesIO(svg_str.encode("utf-8")))
            # 실제 mm → 포인트 (1mm = 2.8346pt)
            pt_w = (w_mm + 4) * 2.8346  # +4 for bleed
            pt_h = (h_mm + 4) * 2.8346
            drawing.width = pt_w
            drawing.height = pt_h
            buf = BytesIO()
            renderPDF.drawToFile(drawing, buf, fmt="PDF")
            buf.seek(0)
            return buf.read()
        except ImportError:
            pass
        # Fallback: 간단한 PDF (SVG를 HTML 경유)
        try:
            import subprocess, tempfile, os
            with tempfile.NamedTemporaryFile(suffix=".svg", delete=False, mode="w") as f:
                f.write(svg_str)
                svg_path = f.name
            pdf_path = svg_path.replace(".svg", ".pdf")
            # cairosvg 시도
            import cairosvg
            cairosvg.svg2pdf(url=svg_path, write_to=pdf_path)
            with open(pdf_path, "rb") as pf:
                data = pf.read()
            os.unlink(svg_path)
            os.unlink(pdf_path)
            return data
        except ImportError:
            pass
        # 최종 Fallback: 순수 reportlab로 텍스트만
        try:
            from reportlab.lib.pagesizes import mm
            from reportlab.pdfgen import canvas as rl_canvas
            buf = BytesIO()
            c = rl_canvas.Canvas(buf, pagesize=((w_mm + 4) * mm, (h_mm + 4) * mm))
            c.drawString(10, 10, "쇼카드 — SVG 렌더링 라이브러리를 설치해주세요 (pip install cairosvg)")
            c.save()
            buf.seek(0)
            return buf.read()
        except ImportError:
            return b""

    # ── 데이터 로드 ──
    all_dims_df = get_all_dimensions()
    if not all_dims_df.empty and "product_name" in all_dims_df.columns:
        dims_dict = {row["product_name"]: row.to_dict() for _, row in all_dims_df.iterrows()}
    else:
        dims_dict = {}
    all_products_raw_df = get_current_placements()
    if not all_products_raw_df.empty and "product_name" in all_products_raw_df.columns:
        all_products_raw = all_products_raw_df.to_dict("records")
        product_names = sorted(set(p["product_name"] for p in all_products_raw if p.get("product_name")))
    else:
        all_products_raw = []
        product_names = []

    # ── 상품 선택 ──
    st.subheader("1️⃣ 상품 선택")
    sc_product = st.selectbox("상품 검색", [""] + product_names, index=0, key="sc_product")

    if sc_product:
        p_info = next((p for p in all_products_raw if p["product_name"] == sc_product), {})
        category = p_info.get("erp_category", "")
        dim = dims_dict.get(sc_product, {})

        col_info1, col_info2 = st.columns(2)
        with col_info1:
            st.info(f"📦 **{sc_product}**")
        with col_info2:
            if dim:
                st.info(f"📐 가로 {dim.get('width', '?')}cm × 높이 {dim.get('height', '?')}cm")
            else:
                st.warning("치수 미등록")

        # ── 사이즈 & 색상 ──
        st.subheader("2️⃣ 사이즈 & 색상")
        col_sz, col_clr1, col_clr2, col_clr3 = st.columns([2, 1, 1, 1])

        with col_sz:
            size_options = _get_size_keys(all_specs)
            rec_size = _recommend_size(dim.get("width", 7)) if dim.get("width") else size_options[0]
            if rec_size not in size_options:
                rec_size = size_options[0]
            def _sz_label(sz):
                sp = all_specs.get(sz, {})
                w = sp.get("card_width_mm", 90)
                h = sp.get("card_height_mm", 65)
                return f"{sz} ({w}×{h}mm)"
            sc_size = st.selectbox(
                "사이즈",
                size_options,
                index=size_options.index(rec_size),
                format_func=_sz_label,
                key="sc_size",
            )

        default_color = _get_showcard_color(category)
        with col_clr1:
            sc_color = st.color_picker("단색 배경", default_color, key="sc_color")
        with col_clr2:
            sc_top_color = st.color_picker("상단 색상", "#FFFFFF", key="sc_top_color")
        with col_clr3:
            sc_bot_color = st.color_picker("하단 색상", default_color, key="sc_bot_color")

        # ── 뱃지 ──
        st.subheader("3️⃣ 뱃지 타입")
        sc_badge = st.radio(
            "뱃지",
            ["none", "동일성분", "유사성분", "업그레이드"],
            format_func=lambda x: {"none": "없음", "동일성분": "동일성분 + 저렴해요", "유사성분": "유사성분 + 저렴해요", "업그레이드": "업그레이드"}.get(x, x),
            horizontal=True,
            key="sc_badge",
        )

        # ── 워딩 입력 ──
        st.subheader("4️⃣ 워딩 입력")
        sc_header_text = st.text_input("상단부 텍스트", placeholder="예: 마트약국 추천", key="sc_header_text")
        col_w1, col_w2 = st.columns(2)
        with col_w1:
            sc_line1 = st.text_input("1줄: 소구 포인트", placeholder="예: 같은 성분, 더 저렴하게", key="sc_l1")
            sc_line2 = st.text_input("2줄: 부가 설명 (선택)", placeholder="예: 속 쓰림엔", key="sc_l2")
            sc_line3 = st.text_input("3줄: 제품명", value=sc_product, key="sc_l3")

        # AI 워딩 제안
        with col_w2:
            st.markdown("**✨ AI 워딩 제안**")
            if st.button("AI 카피 생성", key="sc_ai_btn", use_container_width=True):
                with st.spinner("AI가 카피를 생성하고 있어요..."):
                    try:
                        import anthropic
                        client = anthropic.Anthropic()
                        prompt = (
                            f"마트약국 쇼카드 카피라이터로서, 다음 원본 워딩을 기반으로 2가지 대안을 제안하세요.\n"
                            f"오프라인 매장 쇼카드용으로 짧고 임팩트 있게 작성. 한 줄은 최대 15자 내외.\n\n"
                            f"제품명: {sc_product}\n카테고리: {category or '미분류'}\n뱃지: {sc_badge}\n"
                            f"원본 워딩:\n1줄: {sc_line1}\n2줄: {sc_line2 or '(없음)'}\n3줄: {sc_line3}\n\n"
                            f"기존 쇼카드 예시:\n- \"같은 성분, 더 저렴하게\" + \"노바손\"\n"
                            f"- \"속 쓰림엔\" + \"타이센\"\n- \"잇몸 튼튼\" + \"치렉스정\"\n\n"
                            f"JSON으로만 응답:\n"
                            f'{{"variantA":{{"line1":"임팩트카피","line2":"부가설명","line3":"{sc_product}"}},'
                            f'"variantB":{{"line1":"설득력카피","line2":"효능강조","line3":"{sc_product}"}}}}'
                        )
                        resp = client.messages.create(
                            model="claude-haiku-4-5-20251001",
                            max_tokens=300,
                            messages=[{"role": "user", "content": prompt}],
                        )
                        import json
                        ai_text = resp.content[0].text.strip()
                        # JSON 추출
                        if "{" in ai_text:
                            ai_text = ai_text[ai_text.index("{"):ai_text.rindex("}") + 1]
                        ai_result = json.loads(ai_text)
                        st.session_state["sc_ai_result"] = ai_result
                    except ImportError:
                        st.warning("anthropic 패키지 미설치. `pip install anthropic` 후 ANTHROPIC_API_KEY 환경변수 설정 필요")
                        # 폴백
                        st.session_state["sc_ai_result"] = {
                            "variantA": {"line1": sc_line1[:8] + "!" if sc_line1 else "추천 제품", "line2": "가성비 최고", "line3": sc_product},
                            "variantB": {"line1": {"동일성분": "같은 성분, 더 저렴하게", "유사성분": "비슷한 효과, 합리적 가격", "업그레이드": "한 단계 업그레이드"}.get(sc_badge, "약사 추천"), "line2": category or "", "line3": sc_product},
                        }
                    except Exception as e:
                        st.error(f"AI 생성 실패: {e}")
                        st.session_state["sc_ai_result"] = {
                            "variantA": {"line1": sc_line1[:8] + "!" if sc_line1 else "추천 제품", "line2": "가성비 최고", "line3": sc_product},
                            "variantB": {"line1": "약사 추천", "line2": category or "", "line3": sc_product},
                        }

            if "sc_ai_result" in st.session_state:
                ai = st.session_state["sc_ai_result"]
                st.markdown("---")
                wording_choice = st.radio(
                    "워딩 선택",
                    ["내가 쓴 워딩", "AI 제안 A (임팩트)", "AI 제안 B (설득력)"],
                    key="sc_wording_choice",
                )
                if wording_choice == "AI 제안 A (임팩트)":
                    va = ai["variantA"]
                    st.caption(f'1줄: {va["line1"]}')
                    st.caption(f'2줄: {va["line2"]}')
                    st.caption(f'3줄: {va["line3"]}')
                elif wording_choice == "AI 제안 B (설득력)":
                    vb = ai["variantB"]
                    st.caption(f'1줄: {vb["line1"]}')
                    st.caption(f'2줄: {vb["line2"]}')
                    st.caption(f'3줄: {vb["line3"]}')
                else:
                    st.caption(f"1줄: {sc_line1} / 2줄: {sc_line2} / 3줄: {sc_line3}")

        # ── 최종 워딩 결정 ──
        final_l1, final_l2, final_l3 = sc_line1, sc_line2, sc_line3
        wording_src = "original"
        if "sc_ai_result" in st.session_state and "sc_wording_choice" in st.session_state:
            ai = st.session_state["sc_ai_result"]
            choice = st.session_state["sc_wording_choice"]
            if choice == "AI 제안 A (임팩트)":
                va = ai["variantA"]
                final_l1, final_l2, final_l3 = va["line1"], va["line2"], va["line3"]
                wording_src = "ai_a"
            elif choice == "AI 제안 B (설득력)":
                vb = ai["variantB"]
                final_l1, final_l2, final_l3 = vb["line1"], vb["line2"], vb["line3"]
                wording_src = "ai_b"

        # ── 디자인 프리뷰 ──
        st.subheader("5️⃣ 디자인 프리뷰")

        size_spec = all_specs[sc_size]
        common_spec = all_specs["_common"]
        w_mm = size_spec.get("card_width_mm", 90)
        h_mm = size_spec["card_height_mm"]
        scale = 3
        w_px, h_px = w_mm * scale, h_mm * scale

        svg_a = _gen_design_a(w_px, h_px, sc_color, sc_badge, final_l1, final_l2, final_l3, size_spec, common_spec, header_text=sc_header_text)
        svg_b = _gen_design_b(w_px, h_px, sc_color, sc_badge, final_l1, final_l2, final_l3, size_spec, common_spec, top_color=sc_top_color, bot_color=sc_bot_color, header_text=sc_header_text)

        design_col1, design_col2 = st.columns(2)
        with design_col1:
            st.markdown("**A. 단색 배경**")
            st.markdown(svg_a, unsafe_allow_html=True)
        with design_col2:
            st.markdown("**B. 상하단 컬러 구분**")
            st.markdown(svg_b, unsafe_allow_html=True)

        sc_design = st.radio("디자인 선택", ["A. 단색 배경", "B. 상하단 컬러 구분"], horizontal=True, key="sc_design_choice")
        design_idx = {"A. 단색 배경": 0, "B. 상하단 컬러 구분": 1}[sc_design]
        selected_svg = [svg_a, svg_b][design_idx]

        # ── 다운로드 ──
        st.subheader("6️⃣ 다운로드")

        pdf_bytes = _svg_to_pdf_bytes(selected_svg, w_mm, h_mm)
        design_label = ["solid", "split"][design_idx]
        filename = f"showcard_{sc_product}_{sc_size}_{design_label}.pdf"

        dl_col1, dl_col2 = st.columns(2)
        with dl_col1:
            if pdf_bytes:
                st.download_button(
                    "📥 PDF 다운로드 (인쇄용)",
                    data=pdf_bytes,
                    file_name=filename,
                    mime="application/pdf",
                    use_container_width=True,
                    key="sc_pdf_dl",
                )
            else:
                st.warning("PDF 생성 라이브러리 없음. `pip install cairosvg` 또는 `pip install svglib reportlab` 설치 필요")

        with dl_col2:
            st.download_button(
                "📥 SVG 다운로드",
                data=selected_svg,
                file_name=filename.replace(".pdf", ".svg"),
                mime="image/svg+xml",
                use_container_width=True,
                key="sc_svg_dl",
            )

        # 이력 저장
        if st.button("💾 이력 저장 & 다운로드 기록", key="sc_save_history", use_container_width=True):
            try:
                save_showcard({
                    "product_name": sc_product,
                    "product_id": p_info.get("product_id"),
                    "category": category,
                    "badge_type": sc_badge,
                    "appeal_text": final_l1,
                    "wording_line1": final_l1,
                    "wording_line2": final_l2,
                    "wording_line3": final_l3,
                    "wording_source": wording_src,
                    "size_class": sc_size,
                    "card_width_mm": w_mm,
                    "card_height_mm": h_mm,
                    "bg_color": sc_color,
                    "selected_design": design_idx + 1,
                })
                st.success("이력 저장 완료!")
            except Exception as e:
                st.error(f"저장 실패: {e}")

    # ── 제작 이력 ──
    st.markdown("---")
    st.subheader("📋 제작 이력")
    history = get_showcard_history(30)
    if history:
        design_names = {1: "단색 배경", 2: "상하단 컬러 구분"}
        hist_data = []
        for h in history:
            hist_data.append({
                "제품명": h.get("product_name", ""),
                "사이즈": h.get("size_class", ""),
                "뱃지": h.get("badge_type", ""),
                "디자인": design_names.get(h.get("selected_design"), ""),
                "워딩1": h.get("wording_line1", ""),
                "워딩2": h.get("wording_line2", ""),
                "워딩3": h.get("wording_line3", ""),
                "제작일": pd.to_datetime(h.get("created_at", "")).strftime("%Y-%m-%d %H:%M") if h.get("created_at") else "",
            })
        st.dataframe(pd.DataFrame(hist_data), use_container_width=True, hide_index=True)
    else:
        st.info("아직 제작 이력이 없습니다.")


# ======================================================================
# 포레온 시뮬레이션
# ======================================================================
elif menu == "🏪 포레온 시뮬레이션":

    import plotly.express as px

    # 포레온 매장 규격 (mm)
    FOREON_W = 27323
    FOREON_H = 7487

    # 포레온 전용 매대 설정 (D 타입 추가)
    FOREON_SHELF_CONFIGS = {
        "A": {"name": "기본매대", "width": 90.0, "tiers": [25, 25, 25, 25, 999]},
        "B": {"name": "연결매대", "width": 93.0, "tiers": [25, 25, 25, 25, 25]},
        "C": {"name": "엔드캡매대", "width": 63.6, "tiers": [25, 25, 25, 25, 25]},
        "D": {"name": "벽면매대", "width": 90.0, "tiers": [25, 25, 25, 25, 25, 25]},
    }

    # ── session_state 초기화 (Supabase Storage에서 로드) ──
    _saved_foreon = None
    try:
        from shelf_data import load_foreon_layout
        _saved_foreon = load_foreon_layout()
    except Exception:
        pass

    if _saved_foreon and _saved_foreon.get("fixtures"):
        st.session_state.foreon_fixtures = _saved_foreon["fixtures"]
        st.session_state.foreon_facilities = _saved_foreon.get("facilities", [])
    elif "foreon_fixtures" not in st.session_state:
        # 파일도 없고 세션에도 없을 때 — 자동 배치로 초기 매대 생성
        def _init_foreon_fixtures(num_a=21, num_b=15, num_c=14):
            fxs = []
            margin_back = 2600
            margin_left = 4200
            shelf_w = 900
            shelf_d = 360
            gap = 40
            gondola_pair_w = shelf_d * 2 + gap
            aisle_w = 1200
            col_pitch = gondola_pair_w + aisle_w
            avail_w = FOREON_W - margin_left - 1800
            avail_h = FOREON_H - 1800 - margin_back
            shelves_per_col = max(1, int(avail_h / shelf_w))
            max_cols = max(1, int(avail_w / col_pitch))
            a_placed = b_placed = c_placed = 0
            ab_total = num_a + num_b
            for ci in range(max_cols):
                if (a_placed + b_placed) >= ab_total:
                    break
                col_x = margin_left + ci * col_pitch
                for side in range(2):
                    if (a_placed + b_placed) >= ab_total:
                        break
                    x = col_x + side * (shelf_d + gap)
                    for row in range(shelves_per_col):
                        if (a_placed + b_placed) >= ab_total:
                            break
                        y = margin_back + row * shelf_w
                        if a_placed < num_a:
                            a_placed += 1
                            fxs.append({"id": f"A-{a_placed}", "type": "A", "no": a_placed,
                                        "x": x, "y": y, "orient": "V", "zone": "", "label": ""})
                        elif b_placed < num_b:
                            b_placed += 1
                            fxs.append({"id": f"B-{b_placed}", "type": "B", "no": b_placed,
                                        "x": x, "y": y, "orient": "V", "zone": "", "label": ""})
                # C 엔드캡
                ecx = col_x + gondola_pair_w // 2 - 636 // 2
                if c_placed < num_c:
                    c_placed += 1
                    fxs.append({"id": f"C-{c_placed}", "type": "C", "no": c_placed,
                                "x": ecx, "y": margin_back - 400, "orient": "H", "zone": "", "label": ""})
                if c_placed < num_c:
                    c_placed += 1
                    fxs.append({"id": f"C-{c_placed}", "type": "C", "no": c_placed,
                                "x": ecx, "y": margin_back + shelves_per_col * shelf_w + 40,
                                "orient": "H", "zone": "", "label": ""})
            return fxs

        st.session_state.foreon_fixtures = _init_foreon_fixtures()
        st.session_state.foreon_facilities = [
            {"id": "fac-1", "name": "입구", "x": FOREON_W - 3500, "y": FOREON_H - 500, "w": 2500, "h": 500, "label": ""},
            {"id": "fac-2", "name": "POS", "x": 400, "y": FOREON_H - 1600, "w": 1800, "h": 1000, "label": ""},
            {"id": "fac-3", "name": "조제실", "x": 200, "y": 200, "w": 3500, "h": 2200, "label": ""},
            {"id": "fac-4", "name": "약품 수납장", "x": 3900, "y": 200, "w": 3000, "h": 1200, "label": ""},
            {"id": "fac-5", "name": "냉장고", "x": FOREON_W - 1500, "y": 200, "w": 1300, "h": 3000, "label": ""},
            {"id": "fac-6", "name": "창고", "x": FOREON_W - 1500, "y": 3400, "w": 1300, "h": 2000, "label": ""},
            {"id": "fac-7", "name": "프로모션 존", "x": 8000, "y": FOREON_H - 1400, "w": 3000, "h": 1000, "label": ""},
            {"id": "fac-8", "name": "대기 공간", "x": 18000, "y": FOREON_H - 1400, "w": 2000, "h": 1000, "label": ""},
        ]

    # 이수점 배치 데이터를 포레온 시뮬레이션 초기값으로 복사
    if "foreon_placements" not in st.session_state:
        _isu_placements = get_current_placements()
        if not _isu_placements.empty:
            st.session_state.foreon_placements = _isu_placements.to_dict("records")
        else:
            st.session_state.foreon_placements = []

    # ── 페이지 렌더링 ──
    st.markdown("# 🏪 포레온 마트약국 매대 배치 시뮬레이션")
    st.caption("포레온 2호점 신규 매장의 매대 배치와 상품 배정을 미리 검증합니다")
    st.divider()

    # ── Step 1: 매장 규격 & 요약 ──
    st.markdown("## Step 1. 매장 규격 & 매대 현황")

    foreon_area = FOREON_W * FOREON_H / 1_000_000
    foreon_pyeong = foreon_area / 3.3058
    _fx_list = st.session_state.foreon_fixtures
    _na = len([f for f in _fx_list if f["type"] == "A"])
    _nb = len([f for f in _fx_list if f["type"] == "B"])
    _nc = len([f for f in _fx_list if f["type"] == "C"])
    _nd = len([f for f in _fx_list if f["type"] == "D"])
    _total_fx = _na + _nb + _nc + _nd
    _total_shelves = (_na + _nb + _nc) * 5 + _nd * 6

    # 수용 가능 SKU 예측: 선반 너비 ÷ (평균 상품 폭 + 여유) × 2열 가능 비율
    _dims_df = get_all_dimensions()
    if not _dims_df.empty and _dims_df["width"].notna().any():
        _avg_product_w = _dims_df["width"].dropna().mean()
        _dual_row_rate = _dims_df["dual_row"].mean() if "dual_row" in _dims_df.columns else 0.97
    else:
        _avg_product_w = 8.0
        _dual_row_rate = 0.97

    _capacity_by_type = {}
    for _stype, _scfg in FOREON_SHELF_CONFIGS.items():
        _shelf_w_cm = _scfg["width"]
        _n_tiers = len(_scfg["tiers"])
        _count = {"A": _na, "B": _nb, "C": _nc, "D": _nd}.get(_stype, 0)
        _per_shelf = max(1, int(_shelf_w_cm / (_avg_product_w + 0.3)))
        _per_shelf_dual = _per_shelf * (1 + _dual_row_rate)
        _type_capacity = int(_count * _n_tiers * _per_shelf_dual)
        _capacity_by_type[_stype] = {"count": _count, "tiers": _n_tiers, "shelves": _count * _n_tiers, "capacity": _type_capacity}
    _total_capacity = sum(v["capacity"] for v in _capacity_by_type.values())
    _current_assigned = len(st.session_state.foreon_placements)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("매장 면적", f"{foreon_area:.1f} m² ({foreon_pyeong:.1f}평)")
    c2.metric("규격 (W×H)", f"{FOREON_W/1000:.1f} × {FOREON_H/1000:.1f} m")
    c3.metric("총 매대", f"{_total_fx}대", help=f"A:{_na} B:{_nb} C:{_nc} D:{_nd}")
    c4.metric("총 선반 (단면)", f"{_total_shelves}개")

    st.markdown("")
    ca, cb, cc = st.columns(3)
    ca.metric("📦 예상 수용 SKU", f"{_total_capacity:,}개", help="선반 너비 ÷ 평균 상품 폭 × 2열 배치율 기준 추정")
    cb.metric("🏷️ 현재 배정 상품", f"{_current_assigned:,}건")
    _fill_pct = round(_current_assigned / _total_capacity * 100, 1) if _total_capacity > 0 else 0
    cc.metric("📊 배정률", f"{_fill_pct}%", delta=f"잔여 {_total_capacity - _current_assigned:,}자리")

    # 타입별 수용력 상세
    with st.expander("📐 매대 타입별 수용력 상세", expanded=False):
        _cap_rows = []
        for _stype in ["A", "B", "C", "D"]:
            _cv = _capacity_by_type.get(_stype)
            if _cv and _cv["count"] > 0:
                _scfg = FOREON_SHELF_CONFIGS[_stype]
                _cap_rows.append({
                    "타입": f"{_stype} ({_scfg['name']})",
                    "매대 수": _cv["count"],
                    "단 수": _cv["tiers"],
                    "총 선반": _cv["shelves"],
                    "선반당 수용": f"~{int(_cv['capacity'] / _cv['shelves'])}개" if _cv["shelves"] > 0 else "-",
                    "예상 수용 SKU": f"{_cv['capacity']:,}개",
                })
        st.dataframe(pd.DataFrame(_cap_rows), hide_index=True, use_container_width=True)
        st.caption(f"※ 평균 상품 폭 {_avg_product_w:.1f}cm, 2열 배치율 {_dual_row_rate*100:.0f}% 기준 추정치")

    st.divider()

    # ── Step 2: 매장 배치도 에디터 (드래그 가능) ──
    st.markdown("## Step 2. 매장 배치도")

    _foreon_fx_json = _json.dumps(st.session_state.foreon_fixtures, ensure_ascii=False)
    _foreon_fac_json = _json.dumps(st.session_state.foreon_facilities, ensure_ascii=False)

    # Supabase Storage 직접 저장용 설정 (배포 환경에서 localhost:8503 불가)
    try:
        from trend_config import SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY
        _sb_url = SUPABASE_URL or ""
        _sb_key = SUPABASE_SERVICE_ROLE_KEY or ""
    except Exception:
        _sb_url, _sb_key = "", ""

    _foreon_editor_html = f"""
    <div id="editor-root" style="width:100%;height:480px;position:relative;background:#f8f8f8;border:1px solid #ddd;border-radius:8px;overflow:hidden;">
      <div id="toolbar" style="height:44px;background:#fff;border-bottom:1px solid #ddd;display:flex;align-items:center;padding:0 10px;gap:6px;font-size:13px;">
        <strong style="font-size:13px;color:#FF8C00;">포레온</strong>
        <button onclick="addFixture('A')" style="padding:4px 10px;border:1px solid #4A90D9;color:#4A90D9;border-radius:4px;background:#fff;cursor:pointer;">+A</button>
        <button onclick="addFixture('B')" style="padding:4px 10px;border:1px solid #50C878;color:#50C878;border-radius:4px;background:#fff;cursor:pointer;">+B</button>
        <button onclick="addFixture('C')" style="padding:4px 10px;border:1px solid #FF8C00;color:#FF8C00;border-radius:4px;background:#fff;cursor:pointer;">+C</button>
        <button onclick="addFixture('D')" style="padding:4px 10px;border:1px solid #9B59B6;color:#9B59B6;border-radius:4px;background:#fff;cursor:pointer;">+D</button>
        <span style="width:1px;height:24px;background:#ddd;margin:0 2px;"></span>
        <select id="addFac" onchange="addFacilityFromSelect(this)" style="padding:3px 6px;border:1px solid #ccc;border-radius:4px;font-size:12px;">
          <option value="">+시설물</option>
          <option value="조제실">조제실</option><option value="창고">창고</option>
          <option value="프로모션 존">프로모션 존</option><option value="냉장고">냉장고</option>
          <option value="약품 수납장">약품 수납장</option><option value="POS">POS</option>
          <option value="대기 공간">대기 공간</option><option value="기타">기타</option>
        </select>
        <span style="width:1px;height:24px;background:#ddd;margin:0 2px;"></span>
        <button onclick="rotateSelected()" style="padding:4px 10px;border:1px solid #ccc;border-radius:4px;background:#fff;cursor:pointer;">회전(R)</button>
        <button onclick="deleteSelected()" style="padding:4px 10px;border:1px solid #e74c3c;color:#e74c3c;border-radius:4px;background:#fff;cursor:pointer;">삭제(Del)</button>
        <button onclick="undoAction()" style="padding:4px 10px;border:1px solid #ccc;border-radius:4px;background:#fff;cursor:pointer;">되돌리기</button>
        <span style="width:1px;height:24px;background:#ddd;margin:0 2px;"></span>
        <label style="font-size:11px;">스냅:</label>
        <select id="snapSel" onchange="snapGrid=+this.value" style="padding:2px 4px;border:1px solid #ccc;border-radius:4px;font-size:11px;">
          <option value="0">없음</option><option value="50">50</option><option value="100" selected>100</option><option value="200">200</option>
        </select>
        <span style="flex:1;"></span>
        <button id="saveBtn" onclick="saveLayout()" style="padding:4px 14px;border:1px solid #FF8C00;background:#FF8C00;color:#fff;border-radius:4px;cursor:pointer;font-weight:bold;">레이아웃 저장</button>
        <span id="statusText" style="font-size:11px;color:#888;margin-left:8px;"></span>
      </div>
      <svg id="svg" style="display:block;"></svg>
    </div>
    <style>
      .fixture {{ cursor: move; }}
      .fixture:hover rect {{ stroke-width: 2.5; }}
      .fixture.selected rect {{ stroke-width: 3; filter: drop-shadow(0 0 4px rgba(0,0,0,0.3)); }}
      .fixture text {{ pointer-events: none; user-select: none; }}
      .facility {{ cursor: move; }}
      .facility:hover rect {{ stroke-width: 2.5; }}
      .facility.selected rect {{ stroke-width: 3; stroke-dasharray: 6,3; filter: drop-shadow(0 0 4px rgba(0,0,0,0.3)); }}
      .facility text {{ pointer-events: none; user-select: none; }}
    </style>
    <script>
    const STORE_W = {FOREON_W}, STORE_H = {FOREON_H};
    const TYPES = {{
      A: {{ name:'기본매대', w:900, d:360, color:'#4A90D9', light:'rgba(74,144,217,0.25)' }},
      B: {{ name:'연결매대', w:930, d:360, color:'#50C878', light:'rgba(80,200,120,0.25)' }},
      C: {{ name:'엔드캡매대', w:636, d:360, color:'#FF8C00', light:'rgba(255,140,0,0.25)' }},
      D: {{ name:'벽면매대', w:900, d:360, color:'#9B59B6', light:'rgba(155,89,182,0.25)' }},
    }};
    const FACILITY_TYPES = {{
      '입구':{{w:2500,h:500,c:'#DDD',border:'#999'}},
      '조제실':{{w:3200,h:2200,c:'#D4E6F1',border:'#5B9BD5'}},
      '창고':{{w:1200,h:3000,c:'#E8E8E8',border:'#999'}},
      '프로모션 존':{{w:2200,h:1800,c:'#FCE4EC',border:'#E91E63'}},
      '냉장고':{{w:1200,h:5000,c:'#B3E5FC',border:'#03A9F4'}},
      '약품 수납장':{{w:3800,h:1600,c:'#F3E5F5',border:'#9C27B0'}},
      'POS':{{w:2200,h:700,c:'#E8D5B7',border:'#A0522D'}},
      '대기 공간':{{w:1600,h:1400,c:'#E8F5E9',border:'#4CAF50'}},
      '기타':{{w:1000,h:1000,c:'#F5F5F5',border:'#757575'}},
    }};

    let fixtures = {_foreon_fx_json};
    let facilities = {_foreon_fac_json};
    let selection = [];
    let snapGrid = 100, undoStack = [];
    let scale = 1, panX = 0, panY = 0;
    let isPanning = false, panStartX = 0, panStartY = 0;
    let isDragging = false, dragItems = [], dragStartPositions = [], dragOffX = 0, dragOffY = 0;
    let placingType = null, placingFacility = null;
    let isMarquee = false, marqueeX0 = 0, marqueeY0 = 0, marqueeX1 = 0, marqueeY1 = 0;
    let isResizing = false, resizeFac = null, resizeHandle = '', resizeStartX = 0, resizeStartY = 0;
    let resizeOrigX = 0, resizeOrigY = 0, resizeOrigW = 0, resizeOrigH = 0;
    const HANDLE_SIZE = 8;

    function isSelected(type, id) {{ return selection.some(s => s.type === type && s.id === id); }}

    const svgNS = 'http://www.w3.org/2000/svg';
    const svg = document.getElementById('svg');
    const root = document.getElementById('editor-root');

    function toSVG(mx, my) {{ return [mx * scale + panX, my * scale + panY]; }}
    function fromSVG(sx, sy) {{ return [(sx - panX) / scale, (sy - panY) / scale]; }}
    function snap(v) {{ return snapGrid > 0 ? Math.round(v / snapGrid) * snapGrid : v; }}

    function addRect(parent, x, y, w, h, stroke, sw, fill) {{
      const r = document.createElementNS(svgNS, 'rect');
      r.setAttribute('x', x); r.setAttribute('y', y);
      r.setAttribute('width', w); r.setAttribute('height', h);
      r.setAttribute('stroke', stroke); r.setAttribute('stroke-width', sw);
      r.setAttribute('fill', fill || 'none'); r.setAttribute('rx', 2);
      parent.appendChild(r);
    }}
    function addText(parent, x, y, text, size, color) {{
      const t = document.createElementNS(svgNS, 'text');
      t.setAttribute('x', x); t.setAttribute('y', y);
      t.setAttribute('text-anchor', 'middle'); t.setAttribute('dominant-baseline', 'central');
      t.setAttribute('font-size', size); t.setAttribute('fill', color);
      t.setAttribute('font-family', '-apple-system, sans-serif');
      t.textContent = text;
      parent.appendChild(t);
    }}

    function render() {{
      svg.innerHTML = '';
      const ww = root.clientWidth, wh = root.clientHeight - 44;
      svg.setAttribute('width', ww); svg.setAttribute('height', wh);

      const bg = document.createElementNS(svgNS, 'rect');
      bg.setAttribute('width', ww); bg.setAttribute('height', wh);
      bg.setAttribute('fill', '#f8f8f8');
      svg.appendChild(bg);

      const [sx, sy] = toSVG(0, 0);
      const sw = STORE_W * scale, sh = STORE_H * scale;
      addRect(svg, sx, sy, sw, sh, '#333', 2, 'rgba(255,255,255,0.9)');

      // 그리드
      if (scale > 0.02) {{
        const gridSize = scale > 0.04 ? 1000 : 2000;
        const gridG = document.createElementNS(svgNS, 'g');
        gridG.setAttribute('opacity', '0.15');
        for (let gx = 0; gx <= STORE_W; gx += gridSize) {{
          const [lx] = toSVG(gx, 0);
          const line = document.createElementNS(svgNS, 'line');
          line.setAttribute('x1', lx); line.setAttribute('y1', sy);
          line.setAttribute('x2', lx); line.setAttribute('y2', sy + sh);
          line.setAttribute('stroke', '#999'); line.setAttribute('stroke-width', 0.5);
          gridG.appendChild(line);
        }}
        for (let gy = 0; gy <= STORE_H; gy += gridSize) {{
          const [, ly] = toSVG(0, gy);
          const line = document.createElementNS(svgNS, 'line');
          line.setAttribute('x1', sx); line.setAttribute('y1', ly);
          line.setAttribute('x2', sx + sw); line.setAttribute('y2', ly);
          line.setAttribute('stroke', '#999'); line.setAttribute('stroke-width', 0.5);
          gridG.appendChild(line);
        }}
        svg.appendChild(gridG);
      }}

      // 시설물
      facilities.forEach(fac => {{
        const g = document.createElementNS(svgNS, 'g');
        const isSel = isSelected('facility', fac.id);
        g.setAttribute('class', 'facility' + (isSel ? ' selected' : ''));
        const ft = FACILITY_TYPES[fac.name] || FACILITY_TYPES['기타'];
        const fw = fac.w * scale, fh = fac.h * scale;
        const [fx, fy] = toSVG(fac.x, fac.y);
        // 투명 히트 영역 (패딩 8px) — 더블클릭 감지 영역 확대
        const hitPad = 8;
        const hitRect = document.createElementNS(svgNS, 'rect');
        hitRect.setAttribute('x', fx - hitPad); hitRect.setAttribute('y', fy - hitPad);
        hitRect.setAttribute('width', fw + hitPad * 2); hitRect.setAttribute('height', fh + hitPad * 2);
        hitRect.setAttribute('fill', 'transparent'); hitRect.setAttribute('stroke', 'none');
        g.appendChild(hitRect);
        addRect(g, fx, fy, fw, fh, isSel ? ft.border : '#aaa', isSel ? 3 : 1, ft.c);
        const facNameSize = Math.max(8, 10 * scale / 0.035);
        const pyeong = (fac.w * fac.h) / 1000000 / 3.3058;
        const pyeongStr = pyeong.toFixed(1) + '평';
        addText(g, fx + fw / 2, fy + fh / 2 - facNameSize * 0.6, fac.label || fac.name, facNameSize, '#555');
        addText(g, fx + fw / 2, fy + fh / 2 + facNameSize * 0.7, pyeongStr, Math.max(7, facNameSize * 0.75), '#888');
        if (isSel && selection.length === 1) {{
          const hs = HANDLE_SIZE;
          const handles = [
            {{ name:'nw',cx:fx,cy:fy,cursor:'nw-resize' }},{{ name:'ne',cx:fx+fw,cy:fy,cursor:'ne-resize' }},
            {{ name:'sw',cx:fx,cy:fy+fh,cursor:'sw-resize' }},{{ name:'se',cx:fx+fw,cy:fy+fh,cursor:'se-resize' }},
            {{ name:'n',cx:fx+fw/2,cy:fy,cursor:'n-resize' }},{{ name:'s',cx:fx+fw/2,cy:fy+fh,cursor:'s-resize' }},
            {{ name:'w',cx:fx,cy:fy+fh/2,cursor:'w-resize' }},{{ name:'e',cx:fx+fw,cy:fy+fh/2,cursor:'e-resize' }},
          ];
          handles.forEach(h => {{
            const hr = document.createElementNS(svgNS, 'rect');
            hr.setAttribute('x', h.cx - hs/2); hr.setAttribute('y', h.cy - hs/2);
            hr.setAttribute('width', hs); hr.setAttribute('height', hs);
            hr.setAttribute('fill', '#fff'); hr.setAttribute('stroke', ft.border);
            hr.setAttribute('stroke-width', 1.5); hr.setAttribute('rx', 2);
            hr.style.cursor = h.cursor;
            hr.addEventListener('mousedown', ev => {{ ev.stopPropagation(); startResize(ev, fac, h.name); }});
            g.appendChild(hr);
          }});
          addText(g, fx + fw/2, fy + fh + 14, Math.round(fac.w) + ' x ' + Math.round(fac.h) + ' mm', 9, '#999');
        }}
        g.addEventListener('mousedown', e => onFacilityMouseDown(e, fac));
        g.addEventListener('dblclick', e => {{ e.stopPropagation(); onFacilityDblClick(fac); }});
        svg.appendChild(g);
      }});

      // 매대
      fixtures.forEach(fx => {{
        const g = document.createElementNS(svgNS, 'g');
        const isSel = isSelected('fixture', fx.id);
        g.setAttribute('class', 'fixture' + (isSel ? ' selected' : ''));
        const t = TYPES[fx.type];
        const dx = (fx.orient === 'V' ? t.d : t.w) * scale;
        const dy = (fx.orient === 'V' ? t.w : t.d) * scale;
        const [rx, ry] = toSVG(fx.x, fx.y);
        addRect(g, rx, ry, dx, dy, t.color, isSel ? 3 : 1.5, isSel ? t.color : t.light);
        const label = fx.label || fx.id;
        const fontSize = Math.max(6, Math.min(11, Math.min(dx, dy) * 0.4));
        addText(g, rx + dx/2, ry + dy/2, label, fontSize, isSel ? '#fff' : '#333');
        g.addEventListener('mousedown', e => onFixtureMouseDown(e, fx));
        g.addEventListener('dblclick', e => {{ e.stopPropagation(); onFixtureDblClick(fx); }});
        g.style.cursor = 'pointer';
        svg.appendChild(g);
      }});

      if (isMarquee) {{
        const [sx0, sy0] = toSVG(Math.min(marqueeX0,marqueeX1), Math.min(marqueeY0,marqueeY1));
        const mw = Math.abs(marqueeX1 - marqueeX0) * scale;
        const mh = Math.abs(marqueeY1 - marqueeY0) * scale;
        addRect(svg, sx0, sy0, mw, mh, '#4A90D9', 1.5, 'rgba(74,144,217,0.12)');
      }}
      updateStatus();
    }}

    // ── 이벤트 ──
    function saveUndo() {{ undoStack.push(JSON.stringify({{fixtures,facilities}})); if(undoStack.length>50) undoStack.shift(); }}
    function startDragSelected(e) {{
      const rect = root.getBoundingClientRect();
      const [mx,my] = fromSVG(e.clientX-rect.left, e.clientY-rect.top-44);
      dragOffX=mx; dragOffY=my; isDragging=false;
      dragItems=[]; dragStartPositions=[];
      selection.forEach(s => {{
        let item;
        if(s.type==='fixture') item=fixtures.find(f=>f.id===s.id);
        else item=facilities.find(f=>f.id===s.id);
        if(item) {{ dragItems.push(item); dragStartPositions.push({{x:item.x,y:item.y}}); }}
      }});
      saveUndo();
    }}

    function onFixtureMouseDown(e,fx) {{
      e.stopPropagation();
      if(e.shiftKey) {{
        if(isSelected('fixture',fx.id)) selection=selection.filter(s=>!(s.type==='fixture'&&s.id===fx.id));
        else selection.push({{type:'fixture',id:fx.id}});
        render(); return;
      }}
      if(!isSelected('fixture',fx.id)) selection=[{{type:'fixture',id:fx.id}}];
      render(); startDragSelected(e);
    }}
    function onFacilityMouseDown(e,fac) {{
      e.stopPropagation();
      if(e.shiftKey) {{
        if(isSelected('facility',fac.id)) selection=selection.filter(s=>!(s.type==='facility'&&s.id===fac.id));
        else selection.push({{type:'facility',id:fac.id}});
        render(); return;
      }}
      if(!isSelected('facility',fac.id)) selection=[{{type:'facility',id:fac.id}}];
      render(); startDragSelected(e);
    }}

    function getSelectionBBox() {{
      let minX=Infinity,minY=Infinity,maxX=-Infinity,maxY=-Infinity;
      selection.forEach(s => {{
        let item,w,h;
        if(s.type==='fixture') {{
          item=fixtures.find(f=>f.id===s.id); if(!item) return;
          const t=TYPES[item.type]; w=item.orient==='V'?t.d:t.w; h=item.orient==='V'?t.w:t.d;
        }} else {{
          item=facilities.find(f=>f.id===s.id); if(!item) return; w=item.w; h=item.h;
        }}
        minX=Math.min(minX,item.x); minY=Math.min(minY,item.y);
        maxX=Math.max(maxX,item.x+w); maxY=Math.max(maxY,item.y+h);
      }});
      return {{x0:minX,y0:minY,x1:maxX,y1:maxY}};
    }}

    svg.addEventListener('mousedown', e => {{
      if(placingType) {{
        const rect=root.getBoundingClientRect();
        const[mx,my]=fromSVG(e.clientX-rect.left,e.clientY-rect.top-44);
        placeNewFixture(snap(mx),snap(my)); return;
      }}
      if(placingFacility) {{
        const rect=root.getBoundingClientRect();
        const[mx,my]=fromSVG(e.clientX-rect.left,e.clientY-rect.top-44);
        placeNewFacility(snap(mx),snap(my)); return;
      }}
      const rect=root.getBoundingClientRect();
      const[mx,my]=fromSVG(e.clientX-rect.left,e.clientY-rect.top-44);
      if(e.shiftKey) {{
        isMarquee=true; marqueeX0=mx;marqueeY0=my;marqueeX1=mx;marqueeY1=my;
      }} else if(selection.length>0) {{
        const bbox=getSelectionBBox(); const pad=300;
        if(mx>=bbox.x0-pad&&mx<=bbox.x1+pad&&my>=bbox.y0-pad&&my<=bbox.y1+pad) {{
          dragOffX=mx;dragOffY=my;isDragging=false;
          dragItems=[];dragStartPositions=[];
          selection.forEach(s=>{{
            let item;
            if(s.type==='fixture') item=fixtures.find(f=>f.id===s.id);
            else item=facilities.find(f=>f.id===s.id);
            if(item) {{dragItems.push(item);dragStartPositions.push({{x:item.x,y:item.y}});}}
          }});
          saveUndo();
        }} else {{
          selection=[]; render();
          isPanning=true; panStartX=e.clientX-rect.left-panX; panStartY=e.clientY-rect.top-44-panY;
        }}
      }} else {{
        isPanning=true; panStartX=e.clientX-rect.left-panX; panStartY=e.clientY-rect.top-44-panY;
      }}
    }});

    window.addEventListener('mousemove', e => {{
      const rect=root.getBoundingClientRect();
      const[mx,my]=fromSVG(e.clientX-rect.left,e.clientY-rect.top-44);
      if(isResizing&&resizeFac) {{
        const dx=snap(mx-resizeStartX),dy=snap(my-resizeStartY);
        const MIN=200; const h=resizeHandle;
        let nX=resizeOrigX,nY=resizeOrigY,nW=resizeOrigW,nH=resizeOrigH;
        if(h.includes('e')) nW=Math.max(MIN,resizeOrigW+dx);
        if(h.includes('w')) {{nW=Math.max(MIN,resizeOrigW-dx);nX=resizeOrigX+resizeOrigW-nW;}}
        if(h.includes('s')) nH=Math.max(MIN,resizeOrigH+dy);
        if(h.includes('n')) {{nH=Math.max(MIN,resizeOrigH-dy);nY=resizeOrigY+resizeOrigH-nH;}}
        resizeFac.x=nX;resizeFac.y=nY;resizeFac.w=nW;resizeFac.h=nH;
        render();
      }}
      if(dragItems.length>0) {{
        isDragging=true;
        const dx=snap(mx-dragOffX),dy=snap(my-dragOffY);
        dragItems.forEach((item,i)=>{{
          item.x=Math.max(0,Math.min(STORE_W-100,dragStartPositions[i].x+dx));
          item.y=Math.max(0,Math.min(STORE_H-100,dragStartPositions[i].y+dy));
        }});
        render();
      }}
      if(isMarquee) {{
        marqueeX1=mx;marqueeY1=my;
        const x0=Math.min(marqueeX0,marqueeX1),x1=Math.max(marqueeX0,marqueeX1);
        const y0=Math.min(marqueeY0,marqueeY1),y1=Math.max(marqueeY0,marqueeY1);
        selection=[];
        fixtures.forEach(fx=>{{
          const t=TYPES[fx.type];
          const fw=fx.orient==='V'?t.d:t.w,fh=fx.orient==='V'?t.w:t.d;
          if(fx.x+fw>x0&&fx.x<x1&&fx.y+fh>y0&&fx.y<y1) selection.push({{type:'fixture',id:fx.id}});
        }});
        facilities.forEach(fac=>{{
          if(fac.x+fac.w>x0&&fac.x<x1&&fac.y+fac.h>y0&&fac.y<y1) selection.push({{type:'facility',id:fac.id}});
        }});
        render();
      }}
      if(isPanning) {{
        panX=e.clientX-rect.left-panStartX; panY=e.clientY-rect.top-44-panStartY;
        render();
      }}
    }});

    window.addEventListener('mouseup',()=>{{ dragItems=[];dragStartPositions=[];isPanning=false;isMarquee=false;isResizing=false;resizeFac=null; }});

    svg.addEventListener('wheel', e => {{
      if(!e.ctrlKey&&!e.metaKey) return;
      e.preventDefault();
      const rect=root.getBoundingClientRect();
      const[mx,my]=fromSVG(e.clientX-rect.left,e.clientY-rect.top-44);
      const factor=e.deltaY<0?1.15:1/1.15;
      const newScale=Math.max(0.01,Math.min(0.15,scale*factor));
      panX=(e.clientX-rect.left)-mx*newScale;
      panY=(e.clientY-rect.top-44)-my*newScale;
      scale=newScale;
      render();
    }}, {{passive:false}});

    document.addEventListener('keydown', e => {{
      if(e.target.tagName==='INPUT'||e.target.tagName==='SELECT') return;
      if(e.key==='Delete'||e.key==='Backspace') {{deleteSelected();e.preventDefault();}}
      if(e.key==='r'||e.key==='R') rotateSelected();
      if((e.key==='z'||e.key==='Z')&&(e.metaKey||e.ctrlKey)) {{undoAction();e.preventDefault();}}
      if((e.key==='a'||e.key==='A')&&(e.metaKey||e.ctrlKey)) {{selectAll();e.preventDefault();}}
      if(e.key==='Escape') {{placingType=null;placingFacility=null;selection=[];render();}}
    }});

    function rotateSelected() {{
      if(selection.length===0) return; saveUndo();
      selection.forEach(s=>{{
        if(s.type==='fixture') {{ const fx=fixtures.find(f=>f.id===s.id); if(fx) fx.orient=fx.orient==='V'?'H':'V'; }}
        else {{ const fac=facilities.find(f=>f.id===s.id); if(fac) {{const tmp=fac.w;fac.w=fac.h;fac.h=tmp;}} }}
      }});
      render();
    }}
    function deleteSelected() {{
      if(selection.length===0) return; saveUndo();
      const fxIds=new Set(selection.filter(s=>s.type==='fixture').map(s=>s.id));
      const facIds=new Set(selection.filter(s=>s.type==='facility').map(s=>s.id));
      fixtures=fixtures.filter(f=>!fxIds.has(f.id));
      facilities=facilities.filter(f=>!facIds.has(f.id));
      selection=[]; render();
    }}
    function startResize(e,fac,handle) {{
      isResizing=true;resizeFac=fac;resizeHandle=handle;
      const rect=root.getBoundingClientRect();
      const[mx,my]=fromSVG(e.clientX-rect.left,e.clientY-rect.top-44);
      resizeStartX=mx;resizeStartY=my;
      resizeOrigX=fac.x;resizeOrigY=fac.y;resizeOrigW=fac.w;resizeOrigH=fac.h;
      saveUndo();
    }}
    function selectAll() {{
      selection=[];
      fixtures.forEach(f=>selection.push({{type:'fixture',id:f.id}}));
      facilities.forEach(f=>selection.push({{type:'facility',id:f.id}}));
      render();
    }}

    function onFixtureDblClick(fx) {{
      // /tmp 파일 대신 Supabase Storage에 선택 상태 저장
      const selData = JSON.stringify({{fixture_id: fx.id, ts: Date.now()}});
      fetch(SB_URL + '/storage/v1/object/layouts/foreon_selected_fx.json', {{
        method:'PUT',
        headers:{{
          'Content-Type':'application/json',
          'apikey': SB_KEY,
          'Authorization': 'Bearer ' + SB_KEY,
          'x-upsert': 'true',
        }},
        body: selData,
      }}).then(()=>{{
        window.parent.postMessage({{type:'streamlit:setComponentValue',value:fx.id}}, '*');
        window.parent.location.hash = 'foreon-detail-' + fx.id;
        window.parent.location.reload();
      }}).catch(()=>{{
        // 폴백: 그냥 리로드
        window.parent.location.hash = 'foreon-detail-' + fx.id;
        window.parent.location.reload();
      }});
    }}

    function onFacilityDblClick(fac) {{
      const newName = prompt('시설물 이름을 입력하세요:', fac.label || fac.name);
      if (newName !== null && newName.trim() !== '') {{
        saveUndo();
        fac.label = newName.trim();
        render();
      }}
    }}

    function addFixture(type) {{
      placingType=type;placingFacility=null;
      document.getElementById('statusText').textContent=type+' 배치 중 — 클릭으로 위치 지정 (Esc 취소)';
    }}
    function placeNewFixture(x,y) {{
      if(!placingType) return; saveUndo();
      const type=placingType;
      const existing=fixtures.filter(f=>f.type===type);
      const no=existing.length>0?Math.max(...existing.map(f=>f.no))+1:1;
      fixtures.push({{id:type+'-'+no,type,no,x,y,orient:'V',zone:'',label:''}});
      placingType=null;
      selection=[{{type:'fixture',id:type+'-'+no}}]; render();
    }}
    function addFacilityFromSelect(sel) {{
      const name=sel.value; if(!name) return; sel.value='';
      placingFacility=name;placingType=null;
      document.getElementById('statusText').textContent=name+' 배치 중 — 클릭으로 위치 지정 (Esc 취소)';
    }}
    function placeNewFacility(x,y) {{
      if(!placingFacility) return; saveUndo();
      const name=placingFacility;
      const ft=FACILITY_TYPES[name]||FACILITY_TYPES['기타'];
      const maxNo=facilities.length>0?Math.max(...facilities.map(f=>parseInt(f.id.split('-')[1])||0)):0;
      const id='fac-'+(maxNo+1);
      facilities.push({{id,name,x,y,w:ft.w,h:ft.h,label:''}});
      placingFacility=null;
      selection=[{{type:'facility',id}}]; render();
    }}

    function undoAction() {{
      if(undoStack.length===0) return;
      const state=JSON.parse(undoStack.pop());
      fixtures=state.fixtures||[];facilities=state.facilities||[];
      selection=[]; render();
    }}

    function updateStatus() {{
      if(!placingType&&!placingFacility) {{
        const a=fixtures.filter(f=>f.type==='A').length;
        const b=fixtures.filter(f=>f.type==='B').length;
        const c=fixtures.filter(f=>f.type==='C').length;
        const d=fixtures.filter(f=>f.type==='D').length;
        let msg='A:'+a+' B:'+b+' C:'+c+' D:'+d+' (총 '+(a+b+c+d)+'대)';
        if(selection.length>0) msg+=' | 선택: '+selection.length+'개';
        document.getElementById('statusText').textContent=msg;
      }}
    }}

    const SB_URL = '{_sb_url}';
    const SB_KEY = '{_sb_key}';

    function saveLayout() {{
      const btn=document.getElementById('saveBtn');
      btn.textContent='저장 중...'; btn.style.background='#888';
      const data=JSON.stringify({{
        fixtures:fixtures.map(f=>({{id:f.id,type:f.type,no:f.no,x:Math.round(f.x),y:Math.round(f.y),orient:f.orient,zone:f.zone||'',label:f.label||''}})),
        facilities:facilities.map(f=>({{id:f.id,name:f.name,x:Math.round(f.x),y:Math.round(f.y),w:Math.round(f.w),h:Math.round(f.h),label:f.label||''}})),
      }});

      // Supabase Storage에 직접 저장 (배포 환경 호환)
      const storageUrl = SB_URL + '/storage/v1/object/layouts/foreon_layout.json';
      fetch(storageUrl, {{
        method:'PUT',
        headers:{{
          'Content-Type':'application/json',
          'apikey': SB_KEY,
          'Authorization': 'Bearer ' + SB_KEY,
          'x-upsert': 'true',
        }},
        body:data,
      }})
      .then(r=>{{
        if(!r.ok) throw new Error('HTTP '+r.status);
        return r.json();
      }})
      .then(res=>{{
        btn.textContent='저장 완료!'; btn.style.background='#27ae60';
        setTimeout(()=>{{btn.textContent='레이아웃 저장';btn.style.background='#FF8C00';}},2500);
      }})
      .catch(err=>{{
        console.error('Save error:', err);
        btn.textContent='저장 실패!'; btn.style.background='#e74c3c';
        setTimeout(()=>{{btn.textContent='레이아웃 저장';btn.style.background='#FF8C00';}},3000);
      }});
    }}

    function fitView() {{
      const ww=root.clientWidth,wh=root.clientHeight-44;
      const sx=(ww-40)/STORE_W,sy=(wh-40)/STORE_H;
      scale=Math.min(sx,sy);
      panX=(ww-STORE_W*scale)/2;
      panY=(wh-STORE_H*scale)/2;
    }}

    fitView(); render();
    window.addEventListener('resize',()=>{{fitView();render();}});
    </script>
    """

    from streamlit.components.v1 import html as st_html

    st.info("드래그: 화면이동 | Shift+드래그: 범위선택 | R: 회전 | Del: 삭제 | Ctrl+스크롤: 확대/축소 | **더블클릭: 매대 상세보기**")
    st_html(_foreon_editor_html, height=540, scrolling=False)

    st.divider()

    # ── Step 3: 매대 상세보기 & 상품 배정 ──
    st.markdown("## Step 3. 매대 상세보기 & 상품 배정")

    _fp = st.session_state.foreon_placements
    _fp_df = pd.DataFrame(_fp) if _fp else pd.DataFrame()

    # 더블클릭으로 선택된 매대 읽기 (Supabase Storage + /tmp 폴백)
    _dblclick_fx = None
    try:
        import requests as _requests
        _sel_url = f"{_sb_url}/storage/v1/object/layouts/foreon_selected_fx.json"
        _sel_headers = {"apikey": _sb_key, "Authorization": f"Bearer {_sb_key}", "Cache-Control": "no-cache"}
        _sel_resp = _requests.get(_sel_url, headers=_sel_headers, timeout=5)
        if _sel_resp.status_code == 200:
            _sel_data = _sel_resp.json()
            _dblclick_fx = _sel_data.get("fixture_id", "")
            if _dblclick_fx:
                # 읽은 후 초기화
                _requests.put(_sel_url, headers={**_sel_headers, "Content-Type": "application/json", "x-upsert": "true"}, data=b'{"fixture_id":""}', timeout=5)
    except Exception:
        pass
    if not _dblclick_fx:
        try:
            with open("/tmp/foreon_selected_fx.txt", "r") as f:
                _dblclick_fx = f.read().strip()
            with open("/tmp/foreon_selected_fx.txt", "w") as f:
                f.write("")
        except Exception:
            pass

    # 더블클릭 결과를 세션에 저장
    if _dblclick_fx:
        st.session_state["foreon_detail_fx"] = _dblclick_fx

    # 매대 선택 드롭다운
    _all_fx_ids = sorted(
        [f["id"] for f in st.session_state.foreon_fixtures],
        key=lambda x: (x[0], int(x.split("-")[1]) if "-" in x else 0)
    )

    _default_idx = 0
    if "foreon_detail_fx" in st.session_state and st.session_state["foreon_detail_fx"] in _all_fx_ids:
        _default_idx = _all_fx_ids.index(st.session_state["foreon_detail_fx"])

    _sel_fx = st.selectbox(
        "매대 선택 (배치도에서 더블클릭 또는 여기서 선택)",
        _all_fx_ids,
        index=_default_idx,
        key="foreon_fx_select",
    )

    if _sel_fx:
        st.session_state["foreon_detail_fx"] = _sel_fx

        # 선택된 매대 정보
        _stype = _sel_fx.split("-")[0]
        _sno = int(_sel_fx.split("-")[1]) if "-" in _sel_fx else 1
        _cfg = FOREON_SHELF_CONFIGS.get(_stype, SHELF_CONFIGS.get(_stype, {}))
        _n_tiers = len(_cfg.get("tiers", [5]))
        _shelf_width_cm = _cfg.get("width", 90)
        _tier_heights = _cfg.get("tiers", [25, 25, 25, 25, 25])

        # 해당 매대의 배정 데이터
        if not _fp_df.empty and "shelf_type" in _fp_df.columns:
            _fx_data = _fp_df[
                (_fp_df["shelf_type"] == _stype) &
                (_fp_df["fixture_no"] == _sno)
            ]
        else:
            _fx_data = pd.DataFrame()

        # KPI
        _fx_prods = len(_fx_data) if not _fx_data.empty else 0
        _fx_tiers_used = _fx_data["tier"].nunique() if not _fx_data.empty and "tier" in _fx_data.columns else 0
        kpi1, kpi2, kpi3 = st.columns(3)
        kpi1.metric("매대", f"{_sel_fx} ({_cfg.get('name', '')})")
        kpi2.metric("배치 상품수", f"{_fx_prods}개")
        kpi3.metric("사용 단수", f"{_fx_tiers_used} / {_n_tiers}단")

        # ── 정면도 (칸으로 보기) ──
        _max_pos = 1
        if not _fx_data.empty:
            for _, r in _fx_data.iterrows():
                pe = int(r.get("position_end") or 1)
                if pe > _max_pos:
                    _max_pos = pe
        _max_pos = max(_max_pos, 6)

        _grid = {}
        if not _fx_data.empty:
            for _, r in _fx_data.iterrows():
                tier = int(r.get("tier", 1))
                ps = int(r.get("position_start") or 1)
                pe = int(r.get("position_end") or 1)
                for pos in range(ps, pe + 1):
                    _grid[(tier, pos)] = {
                        "name": r.get("product_name", ""),
                        "category": r.get("erp_category", "") or "",
                        "span_start": ps, "span_end": pe,
                    }

        _grid_json = {}
        for (tier, pos), v in _grid.items():
            _grid_json[f"{tier}-{pos}"] = {
                "name": v["name"], "category": v["category"],
                "span_start": v["span_start"], "span_end": v["span_end"],
            }

        # 카테고리 색상
        _all_cats = sorted(set(v.get("category", "") for v in _grid.values()) - {""})
        _palette = ["#FF6B6B","#4ECDC4","#45B7D1","#96CEB4","#FFEAA7","#DDA0DD","#98D8C8","#F7DC6F","#BB8FCE","#85C1E9","#F0B27A","#82E0AA"]
        _cat_colors = {c: _palette[i % len(_palette)] for i, c in enumerate(_all_cats)}

        _detail_html = f"""
        <div id="foreon-shelf-root" style="width:100%;background:#1a1a2e;border:1px solid #333;border-radius:8px;overflow-x:auto;padding:16px;">
          <div style="color:#fff;font-size:16px;font-weight:bold;margin-bottom:12px;font-family:-apple-system,sans-serif;">
            {_sel_fx} 매대 정면도
            <span style="font-size:12px;color:#888;font-weight:normal;margin-left:8px;">
              ({_cfg.get('name','')}, {_shelf_width_cm}cm, {_n_tiers}단)
            </span>
          </div>
          <svg id="foreon-shelf-svg"></svg>
          <div id="foreon-shelf-tooltip" style="position:fixed;display:none;background:rgba(0,0,0,0.9);color:#fff;padding:10px 14px;border-radius:6px;font-size:12px;pointer-events:none;z-index:9999;max-width:280px;box-shadow:0 4px 12px rgba(0,0,0,0.5);font-family:-apple-system,sans-serif;"></div>
        </div>
        <script>
        (function() {{
          const nTiers = {_n_tiers};
          const maxPos = {_max_pos};
          const grid = {_json.dumps(_grid_json, ensure_ascii=False)};
          const tierHeights = {_json.dumps(_tier_heights)};
          const catColors = {_json.dumps(_cat_colors, ensure_ascii=False)};
          const svgNS = 'http://www.w3.org/2000/svg';
          const svg = document.getElementById('foreon-shelf-svg');
          const tooltip = document.getElementById('foreon-shelf-tooltip');
          const cellW = 90, cellH = 70, labelW = 60, padX = 10, padY = 10;
          const totalW = labelW + maxPos * cellW + padX * 2;
          const totalH = nTiers * cellH + padY * 2 + 30;
          svg.setAttribute('width', totalW); svg.setAttribute('height', totalH); svg.style.display = 'block';
          function fmt(n) {{ return n.toLocaleString('ko-KR'); }}
          const bg = document.createElementNS(svgNS,'rect'); bg.setAttribute('width',totalW); bg.setAttribute('height',totalH); bg.setAttribute('fill','#1a1a2e'); bg.setAttribute('rx',6); svg.appendChild(bg);
          for (let p=1;p<=maxPos;p++) {{ const t=document.createElementNS(svgNS,'text'); t.setAttribute('x',padX+labelW+(p-1)*cellW+cellW/2); t.setAttribute('y',padY+12); t.setAttribute('text-anchor','middle'); t.setAttribute('font-size',10); t.setAttribute('fill','#888'); t.setAttribute('font-family','-apple-system,sans-serif'); t.textContent=p+'번'; svg.appendChild(t); }}
          const rendered = new Set();
          for (let tier=nTiers;tier>=1;tier--) {{
            const rowIdx=nTiers-tier, y=padY+24+rowIdx*cellH, th=tierHeights[tier-1], thLabel=th>=999?'무제한':th+'cm';
            const lbl=document.createElementNS(svgNS,'text'); lbl.setAttribute('x',padX+labelW-6); lbl.setAttribute('y',y+cellH/2); lbl.setAttribute('text-anchor','end'); lbl.setAttribute('dominant-baseline','central'); lbl.setAttribute('font-size',11); lbl.setAttribute('fill','#ccc'); lbl.setAttribute('font-family','-apple-system,sans-serif'); lbl.textContent=tier+'단'; svg.appendChild(lbl);
            const lblS=document.createElementNS(svgNS,'text'); lblS.setAttribute('x',padX+labelW-6); lblS.setAttribute('y',y+cellH/2+13); lblS.setAttribute('text-anchor','end'); lblS.setAttribute('dominant-baseline','central'); lblS.setAttribute('font-size',8); lblS.setAttribute('fill','#666'); lblS.setAttribute('font-family','-apple-system,sans-serif'); lblS.textContent=thLabel; svg.appendChild(lblS);
            for (let pos=1;pos<=maxPos;pos++) {{
              const key=tier+'-'+pos, data=grid[key], x=padX+labelW+(pos-1)*cellW;
              if (data && rendered.has(data.span_start+'-'+data.span_end+'-'+tier+'-'+data.name)) continue;
              const g=document.createElementNS(svgNS,'g');
              if (data) {{
                const spanW=(data.span_end-data.span_start+1)*cellW, spanX=padX+labelW+(data.span_start-1)*cellW;
                const catCol=catColors[data.category]||'#888';
                rendered.add(data.span_start+'-'+data.span_end+'-'+tier+'-'+data.name);
                const rect=document.createElementNS(svgNS,'rect'); rect.setAttribute('x',spanX+1); rect.setAttribute('y',y+1); rect.setAttribute('width',spanW-2); rect.setAttribute('height',cellH-2); rect.setAttribute('fill','#2d3a5e'); rect.setAttribute('stroke',catCol); rect.setAttribute('stroke-width',1.5); rect.setAttribute('rx',4); rect.setAttribute('fill-opacity',0.85); g.appendChild(rect);
                const bar=document.createElementNS(svgNS,'rect'); bar.setAttribute('x',spanX+1); bar.setAttribute('y',y+1); bar.setAttribute('width',spanW-2); bar.setAttribute('height',3); bar.setAttribute('fill',catCol); bar.setAttribute('rx',4); g.appendChild(bar);
                const mc=Math.max(3,Math.floor(spanW/9)); let nameText=data.name.length>mc?data.name.substring(0,mc-1)+'..':data.name;
                const nt=document.createElementNS(svgNS,'text'); nt.setAttribute('x',spanX+spanW/2); nt.setAttribute('y',y+cellH/2-4); nt.setAttribute('text-anchor','middle'); nt.setAttribute('dominant-baseline','central'); nt.setAttribute('font-size',Math.min(10,Math.max(7,spanW/nameText.length*0.85))); nt.setAttribute('fill','#fff'); nt.setAttribute('font-family','-apple-system,sans-serif'); nt.textContent=nameText; g.appendChild(nt);
                if (spanW>60) {{ const ct=document.createElementNS(svgNS,'text'); ct.setAttribute('x',spanX+spanW/2); ct.setAttribute('y',y+cellH/2+14); ct.setAttribute('text-anchor','middle'); ct.setAttribute('dominant-baseline','central'); ct.setAttribute('font-size',7); ct.setAttribute('fill','#999'); ct.setAttribute('font-family','-apple-system,sans-serif'); ct.textContent=data.category.length>8?data.category.substring(0,7)+'..':data.category; g.appendChild(ct); }}
                g.style.cursor='pointer';
                g.addEventListener('mouseenter',(e)=>{{ tooltip.innerHTML='<div style="font-weight:bold;margin-bottom:4px;">'+data.name+'</div><div>위치: '+tier+'단 '+(data.span_start===data.span_end?data.span_start+'번':data.span_start+'~'+data.span_end+'번')+'</div><div>카테고리: '+(data.category||'-')+'</div>'; tooltip.style.display='block'; }});
                g.addEventListener('mousemove',(e)=>{{ tooltip.style.left=(e.clientX+12)+'px'; tooltip.style.top=(e.clientY-10)+'px'; }});
                g.addEventListener('mouseleave',()=>{{ tooltip.style.display='none'; }});
              }} else {{
                const rect=document.createElementNS(svgNS,'rect'); rect.setAttribute('x',x+1); rect.setAttribute('y',y+1); rect.setAttribute('width',cellW-2); rect.setAttribute('height',cellH-2); rect.setAttribute('fill','#1e1e36'); rect.setAttribute('stroke','#333'); rect.setAttribute('stroke-width',0.5); rect.setAttribute('rx',4); rect.setAttribute('fill-opacity',0.5); g.appendChild(rect);
              }}
              svg.appendChild(g);
            }}
            const line=document.createElementNS(svgNS,'line'); line.setAttribute('x1',padX+labelW); line.setAttribute('y1',y+cellH); line.setAttribute('x2',padX+labelW+maxPos*cellW); line.setAttribute('y2',y+cellH); line.setAttribute('stroke','#333'); line.setAttribute('stroke-width',0.5); svg.appendChild(line);
          }}
        }})();
        </script>
        """
        from streamlit.components.v1 import html as _detail_html_fn
        _detail_html_fn(_detail_html, height=_n_tiers * 70 + 80, scrolling=True)

        # ── 단별 상품 목록 & 배치 변경 ──
        st.markdown("---")
        st.markdown("### 단별 상품 배정")

        for _t in range(_n_tiers, 0, -1):
            _tier_prods = _fx_data[_fx_data["tier"] == _t] if not _fx_data.empty and "tier" in _fx_data.columns else pd.DataFrame()
            _th = _tier_heights[_t - 1]
            _th_label = "무제한" if _th >= 999 else f"{_th}cm"

            with st.expander(f"**{_t}단** ({_th_label}) — {len(_tier_prods)}개 상품", expanded=False):
                if not _tier_prods.empty:
                    _tier_display = _tier_prods[["product_name", "erp_category", "position_start", "position_end"]].copy()
                    _tier_display.columns = ["상품명", "카테고리", "시작 위치", "끝 위치"]
                    st.dataframe(_tier_display, use_container_width=True, hide_index=True)
                else:
                    st.caption("배정된 상품이 없습니다.")

                # 배치 상품 바꾸기
                if st.button(f"배치 상품 바꾸기", key=f"fp_change_{_sel_fx}_{_t}", use_container_width=True):
                    st.session_state[f"foreon_edit_tier_{_sel_fx}_{_t}"] = True

                if st.session_state.get(f"foreon_edit_tier_{_sel_fx}_{_t}", False):
                    st.markdown("---")
                    _product_list = load_product_list()
                    if not _product_list.empty and "name" in _product_list.columns:
                        _product_names = sorted(_product_list["name"].dropna().unique().tolist())
                    else:
                        _product_names = []

                    _new_product = st.selectbox(
                        "배정할 상품",
                        _product_names,
                        index=None,
                        placeholder="상품을 검색/선택하세요...",
                        key=f"fp_prod_{_sel_fx}_{_t}",
                    )

                    _btn_c1, _btn_c2, _btn_c3 = st.columns(3)

                    with _btn_c1:
                        if st.button("추가", type="primary", key=f"fp_add_{_sel_fx}_{_t}", use_container_width=True):
                            if _new_product:
                                _p_info = {}
                                if not _product_list.empty:
                                    _match = _product_list[_product_list["name"] == _new_product]
                                    if not _match.empty:
                                        _p_info = _match.iloc[0].to_dict()
                                new_entry = {
                                    "shelf_type": _stype,
                                    "fixture_no": _sno,
                                    "tier": _t,
                                    "product_name": _new_product,
                                    "product_id": _p_info.get("id", ""),
                                    "erp_category": _p_info.get("erp_category", "기타"),
                                    "position_start": len(_tier_prods) + 1,
                                    "position_end": len(_tier_prods) + 1,
                                }
                                st.session_state.foreon_placements.append(new_entry)
                                st.success(f"'{_new_product}' → {_sel_fx} / {_t}단 배정")
                                st.rerun()
                            else:
                                st.warning("상품을 선택하세요.")

                    with _btn_c2:
                        if st.button("이 단 비우기", key=f"fp_clear_{_sel_fx}_{_t}", use_container_width=True):
                            _before = len(st.session_state.foreon_placements)
                            st.session_state.foreon_placements = [
                                p for p in st.session_state.foreon_placements
                                if not (p.get("shelf_type") == _stype and
                                        p.get("fixture_no") == _sno and
                                        p.get("tier") == _t)
                            ]
                            _removed = _before - len(st.session_state.foreon_placements)
                            if _removed > 0:
                                st.success(f"{_sel_fx} / {_t}단 — {_removed}건 제거")
                                st.rerun()
                            else:
                                st.info("제거할 배정이 없습니다.")

                    with _btn_c3:
                        if st.button("닫기", key=f"fp_close_{_sel_fx}_{_t}", use_container_width=True):
                            st.session_state[f"foreon_edit_tier_{_sel_fx}_{_t}"] = False
                            st.rerun()

                    # 개별 상품 제거
                    if not _tier_prods.empty:
                        st.markdown("**개별 상품 제거:**")
                        for idx, row in _tier_prods.iterrows():
                            _pname = row.get("product_name", "")
                            _col_name, _col_btn = st.columns([4, 1])
                            _col_name.text(_pname)
                            if _col_btn.button("제거", key=f"fp_rm_{_sel_fx}_{_t}_{idx}"):
                                _found = False
                                _new_placements = []
                                for p in st.session_state.foreon_placements:
                                    if (not _found and
                                        p.get("shelf_type") == _stype and
                                        p.get("fixture_no") == _sno and
                                        p.get("tier") == _t and
                                        p.get("product_name") == _pname):
                                        _found = True
                                        continue
                                    _new_placements.append(p)
                                st.session_state.foreon_placements = _new_placements
                                st.rerun()

        st.divider()

        # 전체 초기화
        if st.button("🔄 이수점 데이터로 초기화", key="fp_reset_btn"):
            _isu_reset = get_current_placements()
            if not _isu_reset.empty:
                st.session_state.foreon_placements = _isu_reset.to_dict("records")
                st.success("이수점 배치 데이터로 초기화 완료")
                st.rerun()
            else:
                st.warning("이수점 배치 데이터를 불러올 수 없습니다.")

